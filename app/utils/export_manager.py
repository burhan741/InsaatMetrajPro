"""
Export Yöneticisi
Malzeme listesi ve raporlar için export işlemleri
"""

from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


class ExportManager:
    """Export işlemleri yöneticisi"""
    
    def __init__(self) -> None:
        """Export yöneticisini başlat"""
        pass
    
    def export_to_excel(self, materials: List[Dict[str, Any]], 
                       output_path: Path, proje_adi: str = "") -> bool:
        """
        Malzeme listesini Excel dosyasına export et.
        
        Args:
            materials: Malzeme listesi
            output_path: Çıktı dosya yolu
            proje_adi: Proje adı
            
        Returns:
            bool: Başarı durumu
        """
        try:
            # DataFrame oluştur
            data = []
            for material in materials:
                data.append({
                    'Malzeme Adı': material.get('malzeme_adi', ''),
                    'Miktar': material.get('miktar', 0),
                    'Birim': material.get('birim', ''),
                    'Poz No': material.get('poz_no', ''),
                    'Poz Tanım': material.get('poz_tanim', ''),
                    'Açıklama': material.get('aciklama', '')
                })
            
            df = pd.DataFrame(data)
            
            # Excel'e yaz
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Malzeme Listesi', index=False)
                
                # Stil ayarları
                worksheet = writer.sheets['Malzeme Listesi']
                
                # Sütun genişlikleri
                worksheet.column_dimensions['A'].width = 30
                worksheet.column_dimensions['B'].width = 15
                worksheet.column_dimensions['C'].width = 10
                worksheet.column_dimensions['D'].width = 15
                worksheet.column_dimensions['E'].width = 40
                worksheet.column_dimensions['F'].width = 30
                
                # Başlık satırını kalın yap
                from openpyxl.styles import Font, Alignment
                header_font = Font(bold=True)
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            return True
            
        except Exception as e:
            print(f"Excel export hatası: {e}")
            return False
    
    def export_to_pdf(self, materials: List[Dict[str, Any]], 
                     output_path: Path, proje_adi: str = "", 
                     fire_orani: float = 0.0) -> bool:
        """
        Malzeme listesini PDF dosyasına export et.
        
        Args:
            materials: Malzeme listesi
            output_path: Çıktı dosya yolu
            proje_adi: Proje adı
            fire_orani: Fire oranı
            
        Returns:
            bool: Başarı durumu
        """
        try:
            doc = SimpleDocTemplate(str(output_path), pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Başlık
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1a1a2e'),
                spaceAfter=30,
                alignment=1  # Center
            )
            
            title_text = f"Malzeme Listesi"
            if proje_adi:
                title_text += f" - {proje_adi}"
            story.append(Paragraph(title_text, title_style))
            story.append(Spacer(1, 0.5*cm))
            
            # Bilgiler
            info_style = styles['Normal']
            info_text = f"<b>Oluşturulma Tarihi:</b> {datetime.now().strftime('%d.%m.%Y %H:%M')}<br/>"
            if fire_orani > 0:
                info_text += f"<b>Fire/Atık Oranı:</b> %{fire_orani*100:.2f}<br/>"
            info_text += f"<b>Toplam Malzeme Çeşidi:</b> {len(materials)}"
            story.append(Paragraph(info_text, info_style))
            story.append(Spacer(1, 0.5*cm))
            
            # Tablo verileri
            table_data = [['Malzeme Adı', 'Miktar', 'Birim', 'Poz No']]
            
            for material in materials:
                table_data.append([
                    material.get('malzeme_adi', ''),
                    f"{material.get('miktar', 0):,.2f}",
                    material.get('birim', ''),
                    material.get('poz_no', '')
                ])
            
            # Tablo oluştur
            table = Table(table_data, colWidths=[8*cm, 3*cm, 2*cm, 3*cm])
            table.setStyle(TableStyle([
                # Başlık satırı
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16213e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Veri satırları
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Malzeme adı
                ('ALIGN', (1, 1), (1, -1), 'RIGHT'),  # Miktar
                ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Birim
                ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # Poz No
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ]))
            
            story.append(table)
            
            # PDF oluştur
            doc.build(story)
            return True
            
        except Exception as e:
            print(f"PDF export hatası: {e}")
            return False
    
    def export_supplier_format(self, materials: List[Dict[str, Any]], 
                              output_path: Path, firma_adi: str = "") -> bool:
        """
        Tedarikçi formatında export (basit liste).
        
        Args:
            materials: Malzeme listesi
            output_path: Çıktı dosya yolu
            firma_adi: Firma adı
            
        Returns:
            bool: Başarı durumu
        """
        try:
            # Basit metin formatı
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("=" * 60 + "\n")
                f.write("MALZEME SİPARİŞ LİSTESİ\n")
                if firma_adi:
                    f.write(f"Firma: {firma_adi}\n")
                f.write(f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n")
                f.write("=" * 60 + "\n\n")
                
                for i, material in enumerate(materials, 1):
                    f.write(f"{i}. {material.get('malzeme_adi', '')}\n")
                    f.write(f"   Miktar: {material.get('miktar', 0):,.2f} {material.get('birim', '')}\n")
                    if material.get('poz_no'):
                        f.write(f"   Poz: {material.get('poz_no', '')}\n")
                    f.write("\n")
            
            return True
            
        except Exception as e:
            print(f"Tedarikçi format export hatası: {e}")
            return False
    
    def export_taseron_offers_to_excel(self, offers: List[Dict[str, Any]], 
                                       output_path: Path, proje_adi: str = "") -> bool:
        """
        Taşeron tekliflerini Excel dosyasına export et.
        
        Args:
            offers: Taşeron teklifleri listesi
            output_path: Çıktı dosya yolu
            proje_adi: Proje adı
            
        Returns:
            bool: Başarı durumu
        """
        try:
            # DataFrame oluştur
            data = []
            for offer in offers:
                data.append({
                    'Firma Adı': offer.get('firma_adi', ''),
                    'Poz No': offer.get('poz_no', ''),
                    'Tanım': offer.get('tanim', ''),
                    'Miktar': offer.get('miktar', 0),
                    'Birim': offer.get('birim', ''),
                    'Birim Fiyat': offer.get('fiyat', 0),
                    'Toplam': offer.get('toplam', 0),
                    'Durum': offer.get('durum', 'beklemede'),
                    'Teklif Tarihi': offer.get('teklif_tarihi', '')
                })
            
            df = pd.DataFrame(data)
            
            # Excel'e yaz
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Taşeron Teklifleri', index=False)
                
                # Stil ayarları
                worksheet = writer.sheets['Taşeron Teklifleri']
                
                # Sütun genişlikleri
                worksheet.column_dimensions['A'].width = 25
                worksheet.column_dimensions['B'].width = 15
                worksheet.column_dimensions['C'].width = 40
                worksheet.column_dimensions['D'].width = 12
                worksheet.column_dimensions['E'].width = 10
                worksheet.column_dimensions['F'].width = 15
                worksheet.column_dimensions['G'].width = 15
                worksheet.column_dimensions['H'].width = 15
                worksheet.column_dimensions['I'].width = 20
                
                # Başlık satırını kalın yap
                from openpyxl.styles import Font, Alignment, PatternFill
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color='16213e', end_color='16213e', fill_type='solid')
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Toplam satırı ekle
                if len(data) > 0:
                    total_row = len(data) + 2
                    worksheet.cell(row=total_row, column=1, value="TOPLAM")
                    worksheet.cell(row=total_row, column=7, value=f"=SUM(G2:G{len(data)+1})")
                    worksheet.cell(row=total_row, column=1).font = Font(bold=True)
                    worksheet.cell(row=total_row, column=7).font = Font(bold=True)
            
            return True
            
        except Exception as e:
            print(f"Taşeron Excel export hatası: {e}")
            return False
    
    def export_taseron_offers_to_pdf(self, offers: List[Dict[str, Any]], 
                                    output_path: Path, proje_adi: str = "") -> bool:
        """
        Taşeron tekliflerini PDF dosyasına export et.
        
        Args:
            offers: Taşeron teklifleri listesi
            output_path: Çıktı dosya yolu
            proje_adi: Proje adı
            
        Returns:
            bool: Başarı durumu
        """
        try:
            doc = SimpleDocTemplate(str(output_path), pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Başlık
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1a1a2e'),
                spaceAfter=30,
                alignment=1  # Center
            )
            
            title_text = "Taşeron Teklifleri"
            if proje_adi:
                title_text += f" - {proje_adi}"
            story.append(Paragraph(title_text, title_style))
            story.append(Spacer(1, 0.5*cm))
            
            # Bilgiler
            info_style = styles['Normal']
            info_text = f"<b>Oluşturulma Tarihi:</b> {datetime.now().strftime('%d.%m.%Y %H:%M')}<br/>"
            info_text += f"<b>Toplam Teklif Sayısı:</b> {len(offers)}<br/>"
            
            # Firma bazında toplamlar
            firma_totals = {}
            for offer in offers:
                firma = offer.get('firma_adi', '')
                toplam = offer.get('toplam', 0)
                if firma not in firma_totals:
                    firma_totals[firma] = 0.0
                firma_totals[firma] += toplam
            
            if firma_totals:
                info_text += f"<b>Firma Sayısı:</b> {len(firma_totals)}<br/>"
                en_dusuk_firma = min(firma_totals.items(), key=lambda x: x[1])
                en_yuksek_firma = max(firma_totals.items(), key=lambda x: x[1])
                info_text += f"<b>En Düşük Teklif:</b> {en_dusuk_firma[0]} ({en_dusuk_firma[1]:,.2f} ₺)<br/>"
                info_text += f"<b>En Yüksek Teklif:</b> {en_yuksek_firma[0]} ({en_yuksek_firma[1]:,.2f} ₺)<br/>"
            
            story.append(Paragraph(info_text, info_style))
            story.append(Spacer(1, 0.5*cm))
            
            # Tablo verileri
            table_data = [['Firma', 'Tanım', 'Miktar', 'Birim', 'Fiyat', 'Toplam', 'Durum']]
            
            for offer in offers:
                table_data.append([
                    offer.get('firma_adi', ''),
                    offer.get('tanim', '')[:30] + ('...' if len(offer.get('tanim', '')) > 30 else ''),
                    f"{offer.get('miktar', 0):,.2f}" if offer.get('miktar', 0) > 0 else '-',
                    offer.get('birim', ''),
                    f"{offer.get('fiyat', 0):,.2f} ₺",
                    f"{offer.get('toplam', 0):,.2f} ₺",
                    offer.get('durum', 'beklemede').title()
                ])
            
            # Toplam satırı
            toplam = sum(offer.get('toplam', 0) for offer in offers)
            table_data.append([
                'TOPLAM', '', '', '', '', f"{toplam:,.2f} ₺", ''
            ])
            
            # Tablo oluştur
            table = Table(table_data, colWidths=[4*cm, 6*cm, 2*cm, 1.5*cm, 2.5*cm, 2.5*cm, 2*cm])
            table.setStyle(TableStyle([
                # Başlık satırı
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16213e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Veri satırları
                ('BACKGROUND', (0, 1), (-1, -2), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
                ('ALIGN', (0, 1), (0, -2), 'LEFT'),  # Firma
                ('ALIGN', (1, 1), (1, -2), 'LEFT'),  # Tanım
                ('ALIGN', (2, 1), (2, -2), 'RIGHT'),  # Miktar
                ('ALIGN', (3, 1), (3, -2), 'CENTER'),  # Birim
                ('ALIGN', (4, 1), (5, -2), 'RIGHT'),  # Fiyat, Toplam
                ('ALIGN', (6, 1), (6, -2), 'CENTER'),  # Durum
                ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -2), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
                
                # Toplam satırı
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#c9184a')),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 10),
            ]))
            
            story.append(table)
            
            # PDF oluştur
            doc.build(story)
            return True
            
        except Exception as e:
            print(f"Taşeron PDF export hatası: {e}")
            return False






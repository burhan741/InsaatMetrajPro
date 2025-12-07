"""
Ana Pencere
PyQt6 ile modern kullanÄ±cÄ± arayÃ¼zÃ¼
"""

from typing import Optional, List, Dict, Any
from pathlib import Path
from datetime import datetime
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QTabWidget,
    QTableWidget, QTableWidgetItem, QTreeWidget, QTreeWidgetItem,
    QPushButton, QFileDialog, QMessageBox, QLabel, QLineEdit,
    QHeaderView, QSplitter, QGroupBox, QFormLayout, QDoubleSpinBox,
    QComboBox, QTextEdit, QDialog, QMenu, QCheckBox, QScrollArea,
    QInputDialog
)
from PyQt6.QtCore import Qt, pyqtSignal, QThread, pyqtSlot
from PyQt6.QtGui import QIcon, QFont

from app.core.database import DatabaseManager
from app.core.calculator import Calculator
from app.core.material_calculator import MaterialCalculator
from app.utils.data_loader import (
    initialize_database_data, check_pozlar_loaded,
    initialize_material_data, check_malzemeler_loaded, check_formuller_loaded
)
from app.utils.export_manager import ExportManager
from app.utils.pdf_importer import PDFBirimFiyatImporter
from app.ui.dialogs import MetrajItemDialog, TaseronOfferDialog


class DataLoaderThread(QThread):
    """Arka planda veri yÃ¼kleme thread'i"""
    data_loaded = pyqtSignal(dict)
    poz_question_needed = pyqtSignal()
    
    def __init__(self, db: DatabaseManager) -> None:
        super().__init__()
        self.db = db
    
    def run(self) -> None:
        """Thread Ã§alÄ±ÅŸtÄ±ÄŸÄ±nda"""
        result = {
            'malzemeler_loaded': False,
            'formuller_loaded': False,
            'malzeme_count': 0,
            'formul_count': 0
        }
        
        # PozlarÄ± kontrol et
        if not check_pozlar_loaded(self.db):
            # Poz yÃ¼kleme sorusu iÃ§in sinyal gÃ¶nder
            self.poz_question_needed.emit()
        else:
            # Malzeme ve formÃ¼lleri kontrol et ve yÃ¼kle
            if not check_malzemeler_loaded(self.db) or not check_formuller_loaded(self.db):
                try:
                    material_result = initialize_material_data(self.db, force_reload=False)
                    result['malzemeler_loaded'] = material_result['malzemeler']['success'] > 0
                    result['formuller_loaded'] = material_result['formuller']['success'] > 0
                    result['malzeme_count'] = material_result['malzemeler']['success']
                    result['formul_count'] = material_result['formuller']['success']
                except Exception as e:
                    print(f"Malzeme yÃ¼kleme hatasÄ±: {e}")
            else:
                # Zaten yÃ¼klÃ¼, sayÄ±larÄ± al
                try:
                    with self.db.get_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute("SELECT COUNT(*) as count FROM malzemeler")
                        result['malzeme_count'] = cursor.fetchone()['count']
                        cursor.execute("SELECT COUNT(*) as count FROM malzeme_formulleri")
                        result['formul_count'] = cursor.fetchone()['count']
                except Exception as e:
                    print(f"SayÄ±m hatasÄ±: {e}")
        
        # Sonucu gÃ¶nder
        self.data_loaded.emit(result)


class InitialDataLoaderThread(QThread):
    """Ä°lk aÃ§Ä±lÄ±ÅŸta proje ve diÄŸer verileri yÃ¼kleyen thread"""
    projects_loaded = pyqtSignal(list)
    
    def __init__(self, db: DatabaseManager) -> None:
        super().__init__()
        self.db = db
    
    def run(self) -> None:
        """Thread Ã§alÄ±ÅŸtÄ±ÄŸÄ±nda"""
        try:
            # Projeleri yÃ¼kle
            projects = self.db.get_all_projects()
            self.projects_loaded.emit(projects)
        except Exception as e:
            print(f"Proje yÃ¼kleme hatasÄ±: {e}")
            self.projects_loaded.emit([])


class MainWindow(QMainWindow):
    """Ana uygulama penceresi"""
    
    def __init__(self, splash: Optional[Any] = None, user_type: str = 'muteahhit') -> None:
        """Ana pencereyi baÅŸlat"""
        super().__init__()
        
        self.splash = splash
        self.user_type = user_type  # 'muteahhit' veya 'taseron'
        
        # Core modÃ¼ller (hafif olanlar hemen yÃ¼kle)
        self.db = DatabaseManager()
        self.calculator = Calculator()
        self.export_manager = ExportManager()
        
        # AÄŸÄ±r modÃ¼ller lazy loading ile (sadece gerektiÄŸinde yÃ¼klenecek)
        self._material_calculator: Optional[MaterialCalculator] = None
        
        # UI durumu
        self.current_project_id: Optional[int] = None
        self.current_materials: List[Dict[str, Any]] = []  # Hesaplanan malzemeler
        
        # Sekme lazy loading iÃ§in
        self._tabs_created = {
            'metraj': False,
            'ozet': False,
            'taseron': False,
            'malzeme': False,
            'sablonlar': False,
            'birim_fiyat': False,
            'ihale': False
        }
        
        # ArayÃ¼zÃ¼ oluÅŸtur
        if self.splash:
            self.splash.showMessage(
                "ArayÃ¼z oluÅŸturuluyor...",
                Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignBottom,
                Qt.GlobalColor.white
            )
            from PyQt6.QtWidgets import QApplication
            QApplication.processEvents()
        
        # KullanÄ±cÄ± tipine gÃ¶re arayÃ¼z oluÅŸtur
        if self.user_type == 'taseron':
            from app.ui.taseron_window import TaseronWindow
            # TaÅŸeron penceresini gÃ¶ster, mÃ¼teahhit penceresini gizle
            self.taseron_window = TaseronWindow(self.db, self.splash)
            self.taseron_window.show()
            self.taseron_window.showMaximized()  # Tam ekran aÃ§
            self.hide()  # MÃ¼teahhit penceresini gizle
        else:
            self.init_ui()
            
            # VeritabanÄ± yÃ¼klemelerini async yap (UI'Ä± bloklamadan)
            self.load_data_async()
        
        # Ä°lk aÃ§Ä±lÄ±ÅŸta pozlarÄ± kontrol et ve yÃ¼kle (async - arka planda)
        self.check_and_load_pozlar_async()
        
        if self.splash:
            self.splash.showMessage(
                "HazÄ±rlanÄ±yor...",
                Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignBottom,
                Qt.GlobalColor.white
            )
            from PyQt6.QtWidgets import QApplication
            QApplication.processEvents()
    
    @property
    def material_calculator(self) -> MaterialCalculator:
        """MaterialCalculator'Ä± lazy loading ile yÃ¼kle"""
        if self._material_calculator is None:
            self._material_calculator = MaterialCalculator(self.db)
        return self._material_calculator
        
    def init_ui(self) -> None:
        """ArayÃ¼zÃ¼ baÅŸlat"""
        self.setWindowTitle("InsaatMetrajPro - Ä°nÅŸaat Metraj UygulamasÄ±")
        self.setGeometry(100, 100, 1400, 900)
        
        # Uygulama ikonunu ayarla
        icon_path = Path(__file__).parent.parent.parent / "assets" / "app_icon.ico"
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))
        
        # Merkezi widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Arka plan gÃ¶rseli ayarla (ana pencere iÃ§in - wireframe ÅŸehir)
        bg_path = Path(__file__).parent.parent.parent / "assets" / "wireframe_background.jpg"
        if bg_path.exists():
            try:
                # QLabel ile arka plan gÃ¶rseli ekle (daha gÃ¼venilir yÃ¶ntem)
                from PyQt6.QtWidgets import QLabel
                from PyQt6.QtGui import QPixmap
                
                bg_label = QLabel(central_widget)
                bg_pixmap = QPixmap(str(bg_path))
                if not bg_pixmap.isNull():
                    bg_label.setPixmap(bg_pixmap)
                    bg_label.setScaledContents(True)
                    bg_label.lower()  # En alta gÃ¶nder (arka planda kalsÄ±n)
                    self._bg_label = bg_label  # ReferansÄ± sakla
                else:
                    print("Arka plan gÃ¶rseli yÃ¼klenemedi: QPixmap null")
            except Exception as e:
                print(f"Arka plan gÃ¶rseli yÃ¼kleme hatasÄ±: {e}")
                import traceback
                traceback.print_exc()
        
        # Ana layout
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(5, 5, 5, 5)
        main_layout.setSpacing(5)
        
        # Splitter (bÃ¶lÃ¼nmÃ¼ÅŸ pencere)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        main_layout.addWidget(splitter)
        
        # Sol panel - Proje AÄŸacÄ±
        self.create_sidebar(splitter)
        
        # SaÄŸ panel - Sekmeli yapÄ±
        self.create_tabs(splitter)
        
        # Splitter oranlarÄ±
        splitter.setSizes([250, 1150])
        
        # MenÃ¼ Ã§ubuÄŸu
        self.create_menu_bar()
        
        # Durum Ã§ubuÄŸu
        self.statusBar().showMessage("HazÄ±r")
        
    def create_sidebar(self, parent: QSplitter) -> None:
        """Sol sidebar'Ä± oluÅŸtur"""
        sidebar_widget = QWidget()
        sidebar_layout = QVBoxLayout(sidebar_widget)
        sidebar_layout.setContentsMargins(5, 5, 5, 5)
        
        # BaÅŸlÄ±k
        title = QLabel("Projeler")
        title.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        sidebar_layout.addWidget(title)
        
        # HÄ±zlÄ± Arama
        search_group = QGroupBox("ğŸ” HÄ±zlÄ± Arama")
        search_layout = QVBoxLayout()
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Proje, kalem, poz ara...")
        self.search_input.textChanged.connect(self.on_search_text_changed)
        search_layout.addWidget(self.search_input)
        
        self.search_type_combo = QComboBox()
        self.search_type_combo.addItems(["TÃ¼mÃ¼", "Projeler", "Kalemler", "Pozlar"])
        self.search_type_combo.currentTextChanged.connect(self.on_search_text_changed)
        search_layout.addWidget(self.search_type_combo)
        
        search_group.setLayout(search_layout)
        sidebar_layout.addWidget(search_group)
        
        # Proje aÄŸacÄ±
        self.project_tree = QTreeWidget()
        self.project_tree.setHeaderLabel("Projelerim")
        self.project_tree.setRootIsDecorated(True)
        self.project_tree.setSelectionMode(QTreeWidget.SelectionMode.SingleSelection)
        self.project_tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        # Hem tek tÄ±klama hem Ã§ift tÄ±klama ile seÃ§im
        self.project_tree.itemClicked.connect(self.on_project_selected)
        self.project_tree.itemDoubleClicked.connect(self.on_project_selected)
        self.project_tree.customContextMenuRequested.connect(self.show_project_context_menu)
        sidebar_layout.addWidget(self.project_tree)
        
        # Butonlar
        btn_layout = QVBoxLayout()
        
        btn_new = QPushButton("Yeni Proje")
        btn_new.clicked.connect(self.new_project)
        btn_layout.addWidget(btn_new)
        
        btn_delete = QPushButton("Proje Sil")
        btn_delete.clicked.connect(self.delete_selected_project)
        btn_delete.setStyleSheet("background-color: #c9184a;")
        btn_layout.addWidget(btn_delete)
        
        btn_refresh = QPushButton("Yenile")
        btn_refresh.clicked.connect(self.load_projects)
        btn_layout.addWidget(btn_refresh)
        
        sidebar_layout.addLayout(btn_layout)
        
        # Proje NotlarÄ± bÃ¶lÃ¼mÃ¼
        notes_group = QGroupBox("ğŸ“ Proje NotlarÄ±")
        notes_layout = QVBoxLayout()
        
        self.project_notes_text = QTextEdit()
        self.project_notes_text.setPlaceholderText("Proje notlarÄ±nÄ±zÄ± buraya yazÄ±n...")
        self.project_notes_text.setMaximumHeight(150)
        notes_layout.addWidget(self.project_notes_text)
        
        btn_save_notes = QPushButton("NotlarÄ± Kaydet")
        btn_save_notes.clicked.connect(self.save_project_notes)
        notes_layout.addWidget(btn_save_notes)
        
        notes_group.setLayout(notes_layout)
        sidebar_layout.addWidget(notes_group)
        
        sidebar_layout.addStretch()
        
        parent.addWidget(sidebar_widget)
        
    def create_tabs(self, parent: QSplitter) -> None:
        """Sekmeli yapÄ±yÄ± oluÅŸtur (lazy loading ile)"""
        self.tabs = QTabWidget()
        # Sekme deÄŸiÅŸtiÄŸinde lazy loading ve Ã¶zeti gÃ¼ncelle
        self.tabs.currentChanged.connect(self.on_tab_changed)
        
        # Sadece ilk sekmeyi hemen oluÅŸtur, diÄŸerleri lazy loading ile
        # Sekme 1: Metraj Cetveli (ilk sekme, hemen yÃ¼kle)
        self.create_metraj_tab()
        self._tabs_created['metraj'] = True
        
        # DiÄŸer sekmeler placeholder olarak ekle, lazy loading ile yÃ¼klenecek
        self.tabs.addTab(QWidget(), "Proje Ã–zeti")
        self.tabs.addTab(QWidget(), "TaÅŸeron Analizi")
        self.tabs.addTab(QWidget(), "Malzeme Listesi")
        self.tabs.addTab(QWidget(), "Åablonlar")
        self.tabs.addTab(QWidget(), "Birim Fiyat YÃ¶netimi")
        self.tabs.addTab(QWidget(), "Ä°hale DosyasÄ± HazÄ±rlama")
        
        parent.addWidget(self.tabs)
        
    def create_metraj_tab(self) -> None:
        """Metraj Cetveli sekmesini oluÅŸtur"""
        tab = QWidget()
        main_layout = QVBoxLayout(tab)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Buton barÄ±
        btn_layout = QHBoxLayout()
        
        btn_add = QPushButton("Kalem Ekle")
        btn_add.clicked.connect(self.add_metraj_item)
        btn_layout.addWidget(btn_add)
        
        btn_edit = QPushButton("DÃ¼zenle")
        btn_edit.clicked.connect(self.edit_metraj_item)
        btn_layout.addWidget(btn_edit)
        
        btn_delete = QPushButton("Sil")
        btn_delete.clicked.connect(self.delete_metraj_item)
        btn_layout.addWidget(btn_delete)
        
        btn_layout.addStretch()
        
        # KDV oranÄ± seÃ§imi
        kdv_label = QLabel("KDV:")
        btn_layout.addWidget(kdv_label)
        self.metraj_kdv_rate = QComboBox()
        self.metraj_kdv_rate.addItems(["%1", "%10", "%20"])
        self.metraj_kdv_rate.setCurrentText("%20")
        self.metraj_kdv_rate.currentTextChanged.connect(self.update_malzeme_total)
        btn_layout.addWidget(self.metraj_kdv_rate)
        
        # Toplam etiketi (KDV hariÃ§)
        self.total_label = QLabel("Toplam (KDV HariÃ§): 0.00 â‚º")
        self.total_label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        btn_layout.addWidget(self.total_label)
        
        # KDV dahil toplam
        self.total_kdv_label = QLabel("Toplam (KDV Dahil): 0.00 â‚º")
        self.total_kdv_label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        self.total_kdv_label.setStyleSheet("color: #00BFFF;")
        btn_layout.addWidget(self.total_kdv_label)
        
        main_layout.addLayout(btn_layout)
        
        # Splitter: Ãœstte metraj tablosu, altta malzeme detaylarÄ±
        splitter = QSplitter(Qt.Orientation.Vertical)
        
        # Ãœst panel: Metraj Tablosu
        metraj_widget = QWidget()
        metraj_layout = QVBoxLayout(metraj_widget)
        metraj_layout.setContentsMargins(0, 0, 0, 0)
        
        metraj_title = QLabel("ğŸ“Š Metraj Cetveli")
        metraj_title.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        metraj_layout.addWidget(metraj_title)
        
        self.metraj_table = QTableWidget()
        self.metraj_table.setColumnCount(7)
        self.metraj_table.setHorizontalHeaderLabels([
            "ID", "Poz No", "TanÄ±m", "Miktar", "Birim", "Birim Fiyat", "Toplam"
        ])
        self.metraj_table.setAlternatingRowColors(True)
        self.metraj_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.metraj_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.metraj_table.horizontalHeader().setStretchLastSection(True)
        self.metraj_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )
        # SatÄ±r seÃ§ildiÄŸinde malzeme detaylarÄ±nÄ± gÃ¶ster
        self.metraj_table.itemSelectionChanged.connect(self.on_metraj_item_selected)
        metraj_layout.addWidget(self.metraj_table)
        
        splitter.addWidget(metraj_widget)
        
        # Alt panel: Malzeme DetaylarÄ±
        malzeme_widget = QWidget()
        malzeme_layout = QVBoxLayout(malzeme_widget)
        malzeme_layout.setContentsMargins(0, 0, 0, 0)
        
        malzeme_title_layout = QHBoxLayout()
        malzeme_title = QLabel("ğŸ“¦ SeÃ§ili Ä°ÅŸ Kalemi Ä°Ã§in Gereken Malzemeler")
        malzeme_title.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        malzeme_title_layout.addWidget(malzeme_title)
        malzeme_title_layout.addStretch()
        
        # Fire oranÄ± bilgisi
        self.metraj_fire_info = QLabel("")
        self.metraj_fire_info.setStyleSheet("color: #666; font-size: 9pt;")
        malzeme_title_layout.addWidget(self.metraj_fire_info)
        
        malzeme_layout.addLayout(malzeme_title_layout)
        
        self.metraj_malzeme_table = QTableWidget()
        self.metraj_malzeme_table.setColumnCount(5)
        self.metraj_malzeme_table.setHorizontalHeaderLabels([
            "Malzeme AdÄ±", "Miktar", "Birim", "Birim Fiyat", "Toplam"
        ])
        self.metraj_malzeme_table.setAlternatingRowColors(True)
        self.metraj_malzeme_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        # Sadece birim fiyat sÃ¼tunu dÃ¼zenlenebilir
        self.metraj_malzeme_table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked | QTableWidget.EditTrigger.SelectedClicked)
        self.metraj_malzeme_table.horizontalHeader().setStretchLastSection(True)
        self.metraj_malzeme_table.setColumnWidth(0, 250)
        self.metraj_malzeme_table.setColumnWidth(1, 120)
        self.metraj_malzeme_table.setColumnWidth(2, 80)
        self.metraj_malzeme_table.setColumnWidth(3, 120)
        self.metraj_malzeme_table.setMinimumHeight(200)
        # Birim fiyat deÄŸiÅŸtiÄŸinde toplamÄ± gÃ¼ncelle
        self.metraj_malzeme_table.cellChanged.connect(self.on_malzeme_fiyat_changed)
        malzeme_layout.addWidget(self.metraj_malzeme_table)
        
        # Malzeme toplam etiketi
        self.metraj_malzeme_total = QLabel("Toplam: 0.00 â‚º")
        self.metraj_malzeme_total.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        malzeme_layout.addWidget(self.metraj_malzeme_total)
        
        splitter.addWidget(malzeme_widget)
        
        # Splitter oranlarÄ± (Ã¼st %60, alt %40)
        splitter.setSizes([400, 300])
        
        main_layout.addWidget(splitter)
        
        self.tabs.addTab(tab, "ğŸ“Š Metraj Cetveli")
    
    def create_proje_ozet_tab(self, add_to_tabs: bool = True) -> None:
        """Proje Ã–zeti/Rapor sekmesini oluÅŸtur"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(15)
        
        # Ãœst panel: Ã–zet kartlarÄ±
        cards_layout = QHBoxLayout()
        cards_layout.setSpacing(10)
        
        # Kart 1: Toplam Kalem
        self.ozet_kalem_card = QGroupBox("Toplam Kalem")
        kalem_layout = QVBoxLayout()
        self.ozet_kalem_label = QLabel("0")
        self.ozet_kalem_label.setFont(QFont("Arial", 24, QFont.Weight.Bold))
        self.ozet_kalem_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ozet_kalem_label.setStyleSheet("color: #c9184a;")
        kalem_layout.addWidget(self.ozet_kalem_label)
        self.ozet_kalem_card.setLayout(kalem_layout)
        self.ozet_kalem_card.setMinimumHeight(100)
        cards_layout.addWidget(self.ozet_kalem_card)
        
        # Kart 2: Toplam Maliyet
        self.ozet_maliyet_card = QGroupBox("Toplam Maliyet (KDV HariÃ§)")
        maliyet_layout = QVBoxLayout()
        self.ozet_maliyet_label = QLabel("0.00 â‚º")
        self.ozet_maliyet_label.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        self.ozet_maliyet_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ozet_maliyet_label.setStyleSheet("color: #4CAF50;")
        maliyet_layout.addWidget(self.ozet_maliyet_label)
        self.ozet_maliyet_card.setLayout(maliyet_layout)
        self.ozet_maliyet_card.setMinimumHeight(100)
        cards_layout.addWidget(self.ozet_maliyet_card)
        
        # Kart 3: KDV Dahil
        self.ozet_kdv_card = QGroupBox("KDV Dahil Toplam")
        kdv_layout = QVBoxLayout()
        self.ozet_kdv_label = QLabel("0.00 â‚º")
        self.ozet_kdv_label.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        self.ozet_kdv_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ozet_kdv_label.setStyleSheet("color: #16213e;")
        kdv_layout.addWidget(self.ozet_kdv_label)
        # KDV oranÄ± seÃ§imi
        kdv_rate_layout = QHBoxLayout()
        kdv_rate_layout.addWidget(QLabel("KDV OranÄ±:"))
        self.ozet_kdv_rate = QComboBox()
        self.ozet_kdv_rate.addItems(["%1", "%10", "%20"])
        self.ozet_kdv_rate.setCurrentText("%20")
        self.ozet_kdv_rate.currentTextChanged.connect(self.update_proje_ozet)
        kdv_rate_layout.addWidget(self.ozet_kdv_rate)
        kdv_rate_layout.addStretch()
        kdv_layout.addLayout(kdv_rate_layout)
        self.ozet_kdv_card.setLayout(kdv_layout)
        self.ozet_kdv_card.setMinimumHeight(100)
        cards_layout.addWidget(self.ozet_kdv_card)
        
        # Kart 4: TaÅŸeron Teklif SayÄ±sÄ±
        self.ozet_taseron_card = QGroupBox("TaÅŸeron Teklifleri")
        taseron_layout = QVBoxLayout()
        self.ozet_taseron_label = QLabel("0")
        self.ozet_taseron_label.setFont(QFont("Arial", 24, QFont.Weight.Bold))
        self.ozet_taseron_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ozet_taseron_label.setStyleSheet("color: #00BFFF;")
        taseron_layout.addWidget(self.ozet_taseron_label)
        self.ozet_taseron_card.setLayout(taseron_layout)
        self.ozet_taseron_card.setMinimumHeight(100)
        cards_layout.addWidget(self.ozet_taseron_card)
        
        layout.addLayout(cards_layout)
        
        # Orta panel: Splitter (Grafikler ve Tablolar)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Sol: Kategori DaÄŸÄ±lÄ±mÄ± (Grafik + Tablo)
        kategori_widget = QWidget()
        kategori_layout = QVBoxLayout(kategori_widget)
        kategori_layout.setContentsMargins(0, 0, 0, 0)
        
        kategori_title = QLabel("ğŸ“‹ Kategori BazÄ±nda DaÄŸÄ±lÄ±m")
        kategori_title.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        kategori_layout.addWidget(kategori_title)
        
        # Pie Chart iÃ§in matplotlib widget
        try:
            from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
            from matplotlib.figure import Figure
            import matplotlib.pyplot as plt
            
            self.kategori_figure = Figure(figsize=(5, 4))
            self.kategori_canvas = FigureCanvasQTAgg(self.kategori_figure)
            self.kategori_ax = self.kategori_figure.add_subplot(111)
            self.kategori_canvas.setMinimumHeight(250)
            kategori_layout.addWidget(self.kategori_canvas)
        except ImportError:
            # Matplotlib yoksa placeholder
            placeholder = QLabel("Matplotlib yÃ¼klenmedi. Grafik gÃ¶sterilemiyor.")
            placeholder.setStyleSheet("color: #666; padding: 20px;")
            kategori_layout.addWidget(placeholder)
            self.kategori_canvas = None
        
        # Kategori tablosu
        self.ozet_kategori_table = QTableWidget()
        self.ozet_kategori_table.setColumnCount(3)
        self.ozet_kategori_table.setHorizontalHeaderLabels([
            "Kategori", "Kalem SayÄ±sÄ±", "Toplam Maliyet"
        ])
        self.ozet_kategori_table.setAlternatingRowColors(True)
        self.ozet_kategori_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.ozet_kategori_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.ozet_kategori_table.horizontalHeader().setStretchLastSection(True)
        self.ozet_kategori_table.setColumnWidth(0, 200)
        self.ozet_kategori_table.setColumnWidth(1, 120)
        self.ozet_kategori_table.setMaximumHeight(150)
        kategori_layout.addWidget(self.ozet_kategori_table)
        
        splitter.addWidget(kategori_widget)
        
        # SaÄŸ: En PahalÄ± Kalemler (Bar Chart + Tablo)
        pahali_widget = QWidget()
        pahali_layout = QVBoxLayout(pahali_widget)
        pahali_layout.setContentsMargins(0, 0, 0, 0)
        
        pahali_title = QLabel("ğŸ’° En PahalÄ± 5 Kalem")
        pahali_title.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        pahali_layout.addWidget(pahali_title)
        
        # Bar Chart iÃ§in matplotlib widget
        try:
            from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
            from matplotlib.figure import Figure
            
            self.pahali_figure = Figure(figsize=(5, 4))
            self.pahali_canvas = FigureCanvasQTAgg(self.pahali_figure)
            self.pahali_ax = self.pahali_figure.add_subplot(111)
            self.pahali_canvas.setMinimumHeight(250)
            pahali_layout.addWidget(self.pahali_canvas)
        except ImportError:
            placeholder = QLabel("Matplotlib yÃ¼klenmedi. Grafik gÃ¶sterilemiyor.")
            placeholder.setStyleSheet("color: #666; padding: 20px;")
            pahali_layout.addWidget(placeholder)
            self.pahali_canvas = None
        
        # En pahalÄ± kalemler tablosu
        self.ozet_pahali_table = QTableWidget()
        self.ozet_pahali_table.setColumnCount(3)
        self.ozet_pahali_table.setHorizontalHeaderLabels([
            "Kalem", "Miktar", "Toplam"
        ])
        self.ozet_pahali_table.setAlternatingRowColors(True)
        self.ozet_pahali_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.ozet_pahali_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.ozet_pahali_table.horizontalHeader().setStretchLastSection(True)
        self.ozet_pahali_table.setColumnWidth(0, 250)
        self.ozet_pahali_table.setColumnWidth(1, 100)
        self.ozet_pahali_table.setMaximumHeight(150)
        pahali_layout.addWidget(self.ozet_pahali_table)
        
        splitter.addWidget(pahali_widget)
        
        splitter.setSizes([400, 400])
        layout.addWidget(splitter)
        
        # Alt panel: Ä°statistikler ve DetaylÄ± Analiz
        stats_group = QGroupBox("ğŸ“Š DetaylÄ± Ä°statistikler ve Analiz")
        stats_layout = QVBoxLayout()
        
        # Ä°statistik tablosu
        self.stats_table = QTableWidget()
        self.stats_table.setColumnCount(2)
        self.stats_table.setHorizontalHeaderLabels(["Ä°statistik", "DeÄŸer"])
        self.stats_table.setAlternatingRowColors(True)
        self.stats_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.stats_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.stats_table.horizontalHeader().setStretchLastSection(True)
        self.stats_table.setColumnWidth(0, 300)
        self.stats_table.setMaximumHeight(200)
        stats_layout.addWidget(self.stats_table)
        
        stats_group.setLayout(stats_layout)
        layout.addWidget(stats_group)
        
        # Alt panel: Malzeme ve TaÅŸeron Ã–zeti
        alt_splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Malzeme Ã–zeti
        malzeme_ozet_widget = QWidget()
        malzeme_ozet_layout = QVBoxLayout(malzeme_ozet_widget)
        malzeme_ozet_layout.setContentsMargins(0, 0, 0, 0)
        
        malzeme_ozet_title = QLabel("ğŸ“¦ Malzeme Ã–zeti")
        malzeme_ozet_title.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        malzeme_ozet_layout.addWidget(malzeme_ozet_title)
        
        self.ozet_malzeme_label = QLabel("Malzeme listesi hesaplanmadÄ±.\n'Malzeme Listesi' sekmesinden hesaplayÄ±nÄ±z.")
        self.ozet_malzeme_label.setWordWrap(True)
        self.ozet_malzeme_label.setStyleSheet("color: #666; padding: 10px;")
        malzeme_ozet_layout.addWidget(self.ozet_malzeme_label)
        
        alt_splitter.addWidget(malzeme_ozet_widget)
        
        # TaÅŸeron Ã–zeti
        taseron_ozet_widget = QWidget()
        taseron_ozet_layout = QVBoxLayout(taseron_ozet_widget)
        taseron_ozet_layout.setContentsMargins(0, 0, 0, 0)
        
        taseron_ozet_title = QLabel("ğŸ’¼ TaÅŸeron Ã–zeti")
        taseron_ozet_title.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        taseron_ozet_layout.addWidget(taseron_ozet_title)
        
        self.ozet_taseron_detay_label = QLabel("TaÅŸeron teklif bilgisi yok.")
        self.ozet_taseron_detay_label.setWordWrap(True)
        self.ozet_taseron_detay_label.setStyleSheet("color: #666; padding: 10px;")
        taseron_ozet_layout.addWidget(self.ozet_taseron_detay_label)
        
        alt_splitter.addWidget(taseron_ozet_widget)
        
        alt_splitter.setSizes([400, 400])
        layout.addWidget(alt_splitter)
        
        # Export butonlarÄ±
        export_layout = QHBoxLayout()
        export_layout.addStretch()
        
        btn_export_pdf = QPushButton("PDF Rapor OluÅŸtur")
        btn_export_pdf.clicked.connect(self.export_proje_ozet_pdf)
        export_layout.addWidget(btn_export_pdf)
        
        btn_export_excel = QPushButton("Excel Rapor OluÅŸtur")
        btn_export_excel.clicked.connect(self.export_proje_ozet_excel)
        export_layout.addWidget(btn_export_excel)
        
        layout.addLayout(export_layout)
        
        self.ozet_widget = tab
        if add_to_tabs:
            self.tabs.addTab(tab, "ğŸ“ˆ Proje Ã–zeti")
        
    def create_taseron_tab(self, add_to_tabs: bool = True) -> None:
        """TaÅŸeron Analizi sekmesini oluÅŸtur"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Buton barÄ±
        btn_layout = QHBoxLayout()
        
        btn_add = QPushButton("Teklif Ekle")
        btn_add.clicked.connect(self.add_taseron_offer)
        btn_layout.addWidget(btn_add)
        
        btn_edit = QPushButton("DÃ¼zenle")
        btn_edit.clicked.connect(self.edit_taseron_offer)
        btn_layout.addWidget(btn_edit)
        
        btn_delete = QPushButton("Sil")
        btn_delete.clicked.connect(self.delete_taseron_offer)
        btn_delete.setStyleSheet("background-color: #c9184a;")
        btn_layout.addWidget(btn_delete)
        
        btn_layout.addStretch()
        
        # KDV oranÄ± seÃ§imi
        kdv_label = QLabel("KDV:")
        btn_layout.addWidget(kdv_label)
        self.taseron_kdv_rate = QComboBox()
        self.taseron_kdv_rate.addItems(["%1", "%10", "%20"])
        self.taseron_kdv_rate.setCurrentText("%20")
        self.taseron_kdv_rate.currentTextChanged.connect(self.load_taseron_data)
        btn_layout.addWidget(self.taseron_kdv_rate)
        
        # Toplam etiketleri
        self.taseron_total_label = QLabel("Toplam (KDV HariÃ§): 0.00 â‚º")
        self.taseron_total_label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        btn_layout.addWidget(self.taseron_total_label)
        
        self.taseron_total_kdv_label = QLabel("Toplam (KDV Dahil): 0.00 â‚º")
        self.taseron_total_kdv_label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        self.taseron_total_kdv_label.setStyleSheet("color: #00BFFF;")
        btn_layout.addWidget(self.taseron_total_kdv_label)
        
        btn_compare = QPushButton("KarÅŸÄ±laÅŸtÄ±r")
        btn_compare.clicked.connect(self.compare_offers)
        btn_layout.addWidget(btn_compare)
        
        # Export butonlarÄ±
        btn_export_excel = QPushButton("Excel'e Aktar")
        btn_export_excel.clicked.connect(self.export_taseron_excel)
        btn_layout.addWidget(btn_export_excel)
        
        btn_export_pdf = QPushButton("PDF'e Aktar")
        btn_export_pdf.clicked.connect(self.export_taseron_pdf)
        btn_layout.addWidget(btn_export_pdf)
        
        layout.addLayout(btn_layout)
        
        # Tablo
        self.taseron_table = QTableWidget()
        self.taseron_table.setColumnCount(7)
        self.taseron_table.setHorizontalHeaderLabels([
            "ID", "Firma", "Kalem", "Miktar", "Birim", "Fiyat", "Toplam"
        ])
        self.taseron_table.setAlternatingRowColors(True)
        self.taseron_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.taseron_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.taseron_table.horizontalHeader().setStretchLastSection(True)
        self.taseron_table.setColumnHidden(0, True)  # ID sÃ¼tununu gizle
        layout.addWidget(self.taseron_table)
        
        # KarÅŸÄ±laÅŸtÄ±rma sonuÃ§larÄ± (tablo olarak)
        comparison_group = QGroupBox("Teklif KarÅŸÄ±laÅŸtÄ±rmasÄ±")
        comparison_layout = QVBoxLayout()
        
        self.comparison_table = QTableWidget()
        self.comparison_table.setColumnCount(4)
        self.comparison_table.setHorizontalHeaderLabels([
            "Firma", "Toplam Tutar", "Durum", "Fark (Ortalamadan)"
        ])
        self.comparison_table.setAlternatingRowColors(True)
        self.comparison_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.comparison_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.comparison_table.horizontalHeader().setStretchLastSection(True)
        self.comparison_table.setMaximumHeight(200)
        comparison_layout.addWidget(self.comparison_table)
        
        self.comparison_summary_label = QLabel("")
        self.comparison_summary_label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        comparison_layout.addWidget(self.comparison_summary_label)
        
        comparison_group.setLayout(comparison_layout)
        layout.addWidget(comparison_group)
        
        self.taseron_widget = tab
        if add_to_tabs:
            self.tabs.addTab(tab, "ğŸ’¼ TaÅŸeron Analizi")
    
    def create_malzeme_tab(self, add_to_tabs: bool = True) -> None:
        """Malzeme Listesi sekmesini oluÅŸtur"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Kontrol paneli
        control_group = QGroupBox("Hesaplama AyarlarÄ±")
        control_layout = QFormLayout()
        
        # Fire oranÄ± modu
        self.fire_mode_combo = QComboBox()
        self.fire_mode_combo.addItems([
            "Otomatik (Poz BazlÄ± - LiteratÃ¼r DeÄŸerleri)",
            "Manuel (TÃ¼m Pozlar Ä°Ã§in AynÄ±)"
        ])
        self.fire_mode_combo.currentIndexChanged.connect(self.on_fire_mode_changed)
        control_layout.addRow("Fire OranÄ± Modu:", self.fire_mode_combo)
        
        # Fire oranÄ± (manuel mod iÃ§in)
        self.fire_spin = QDoubleSpinBox()
        self.fire_spin.setMinimum(0.0)
        self.fire_spin.setMaximum(50.0)
        self.fire_spin.setValue(5.0)
        self.fire_spin.setSuffix(" %")
        self.fire_spin.setDecimals(2)
        self.fire_spin.setEnabled(False)  # BaÅŸlangÄ±Ã§ta otomatik mod
        control_layout.addRow("Manuel Fire/AtÄ±k OranÄ±:", self.fire_spin)
        
        # Bilgi etiketi
        self.fire_info_label = QLabel("â„¹ï¸ Otomatik mod: Her poz iÃ§in LiteratÃ¼r/Kitap deÄŸerlerine gÃ¶re fire oranÄ± kullanÄ±lÄ±r.")
        self.fire_info_label.setWordWrap(True)
        self.fire_info_label.setStyleSheet("color: #666; font-size: 9pt;")
        control_layout.addRow("", self.fire_info_label)
        
        # Hesapla butonu
        btn_calculate = QPushButton("Malzemeleri Hesapla")
        btn_calculate.clicked.connect(self.calculate_materials)
        control_layout.addRow("", btn_calculate)
        
        control_group.setLayout(control_layout)
        layout.addWidget(control_group)
        
        # Export butonlarÄ±
        export_group = QGroupBox("Export Ä°ÅŸlemleri")
        export_layout = QHBoxLayout()
        
        btn_export_excel = QPushButton("Excel'e Aktar")
        btn_export_excel.clicked.connect(self.export_materials_excel)
        export_layout.addWidget(btn_export_excel)
        
        btn_export_pdf = QPushButton("PDF'e Aktar")
        btn_export_pdf.clicked.connect(self.export_materials_pdf)
        export_layout.addWidget(btn_export_pdf)
        
        btn_export_supplier = QPushButton("TedarikÃ§i FormatÄ±")
        btn_export_supplier.clicked.connect(self.export_materials_supplier)
        export_layout.addWidget(btn_export_supplier)
        
        export_group.setLayout(export_layout)
        layout.addWidget(export_group)
        
        # Ã–zet bilgiler
        self.material_summary_label = QLabel("Proje seÃ§iniz ve 'Malzemeleri Hesapla' butonuna tÄ±klayÄ±nÄ±z.")
        self.material_summary_label.setFont(QFont("Arial", 10))
        self.material_summary_label.setWordWrap(True)
        layout.addWidget(self.material_summary_label)
        
        # Malzeme tablosu
        self.material_table = QTableWidget()
        self.material_table.setColumnCount(4)
        self.material_table.setHorizontalHeaderLabels([
            "Malzeme AdÄ±", "Miktar", "Birim", "Poz Bilgisi"
        ])
        self.material_table.setAlternatingRowColors(True)
        self.material_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.material_table.horizontalHeader().setStretchLastSection(True)
        self.material_table.setColumnWidth(0, 300)
        self.material_table.setColumnWidth(1, 150)
        self.material_table.setColumnWidth(2, 100)
        layout.addWidget(self.material_table)
        
        self.malzeme_widget = tab
        if add_to_tabs:
            self.tabs.addTab(tab, "ğŸ“¦ Malzeme Listesi")
    
    def on_fire_mode_changed(self, index: int) -> None:
        """Fire oranÄ± modu deÄŸiÅŸtiÄŸinde"""
        if index == 0:  # Otomatik mod
            self.fire_spin.setEnabled(False)
            self.fire_info_label.setText(
                "â„¹ï¸ Otomatik mod: Her poz iÃ§in LiteratÃ¼r/Kitap deÄŸerlerine gÃ¶re fire oranÄ± kullanÄ±lÄ±r.\n"
                "Kaynak: Ä°nÅŸaat Metraj kitaplarÄ± ve TBDY/TS standartlarÄ±na uygun genel kabul gÃ¶rmÃ¼ÅŸ deÄŸerler."
            )
        else:  # Manuel mod
            self.fire_spin.setEnabled(True)
            self.fire_info_label.setText(
                "â„¹ï¸ Manuel mod: TÃ¼m pozlar iÃ§in aynÄ± fire oranÄ± kullanÄ±lÄ±r.\n"
                "Bu deÄŸer poz bazlÄ± otomatik fire oranlarÄ±nÄ± geÃ§ersiz kÄ±lar."
            )
    
    def calculate_materials(self) -> None:
        """Proje iÃ§in malzeme listesini hesapla ve gÃ¶ster"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§iniz.")
            return
        
        try:
            # Fire oranÄ± modunu kontrol et
            fire_mode = self.fire_mode_combo.currentIndex()
            fire_orani_override = None
            
            if fire_mode == 1:  # Manuel mod
                fire_yuzde = self.fire_spin.value()
                fire_orani_override = fire_yuzde / 100.0
            
            # Malzemeleri hesapla (poz bazlÄ± otomatik fire oranlarÄ± kullanÄ±lÄ±r)
            materials = self.material_calculator.calculate_materials_for_project(
                self.current_project_id, fire_orani_override
            )
            
            if not materials:
                QMessageBox.information(
                    self, "Bilgi", 
                    "Bu proje iÃ§in malzeme formÃ¼lÃ¼ bulunamadÄ±.\n"
                    "LÃ¼tfen pozlar iÃ§in malzeme formÃ¼lleri tanÄ±mlayÄ±nÄ±z."
                )
                self.material_table.setRowCount(0)
                self.material_summary_label.setText("Malzeme bulunamadÄ±.")
                return
            
            # Tabloyu doldur
            self.material_table.setRowCount(len(materials))
            
            for row, material in enumerate(materials):
                # Malzeme adÄ±
                item = QTableWidgetItem(material.get('malzeme_adi', ''))
                self.material_table.setItem(row, 0, item)
                
                # Miktar
                miktar = material.get('miktar', 0)
                item = QTableWidgetItem(f"{miktar:,.2f}")
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.material_table.setItem(row, 1, item)
                
                # Birim
                item = QTableWidgetItem(material.get('birim', ''))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.material_table.setItem(row, 2, item)
                
                # Poz bilgisi (hangi pozlardan geldiÄŸi)
                poz_info = material.get('poz_no', '')
                if poz_info:
                    poz_tanim = material.get('poz_tanim', '')
                    poz_miktar = material.get('poz_miktar', 0)
                    poz_birim = material.get('poz_birim', '')
                    poz_info = f"{poz_info} ({poz_tanim[:30]}... - {poz_miktar} {poz_birim})"
                item = QTableWidgetItem(poz_info)
                self.material_table.setItem(row, 3, item)
            
            # Hesaplanan malzemeleri sakla (export iÃ§in)
            self.current_materials = materials
            self.update_proje_ozet()  # Ã–zeti gÃ¼ncelle (malzeme bilgisi iÃ§in)
            
            # Ã–zet bilgi
            toplam_cesit = len(materials)
            toplam_miktar = sum(m.get('miktar', 0) for m in materials)
            
            if fire_mode == 0:
                # Otomatik mod - poz bazlÄ± fire oranlarÄ± kullanÄ±ldÄ±
                summary = (
                    f"Toplam {toplam_cesit} farklÄ± malzeme tÃ¼rÃ¼ hesaplandÄ±.\n"
                    f"Fire oranÄ±: Otomatik (Poz bazlÄ± - LiteratÃ¼r/Kitap deÄŸerleri)"
                )
            else:
                # Manuel mod
                summary = (
                    f"Toplam {toplam_cesit} farklÄ± malzeme tÃ¼rÃ¼ hesaplandÄ±.\n"
                    f"Fire oranÄ±: Manuel %{fire_orani_override*100:.2f} (TÃ¼m pozlar iÃ§in)"
                )
            self.material_summary_label.setText(summary)
            
            self.statusBar().showMessage(f"Malzeme listesi hesaplandÄ±: {toplam_cesit} Ã§eÅŸit")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Malzeme hesaplanÄ±rken bir hata oluÅŸtu:\n{str(e)}")
            print(f"Malzeme hesaplama hatasÄ±: {e}")
    
    def export_materials_excel(self) -> None:
        """Malzeme listesini Excel'e export et"""
        if not self.current_materials:
            QMessageBox.warning(self, "UyarÄ±", "Ã–nce malzeme listesini hesaplayÄ±nÄ±z.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Excel'e Kaydet", "", "Excel DosyalarÄ± (*.xlsx)"
        )
        
        if file_path:
            proje = self.db.get_project(self.current_project_id) if self.current_project_id else None
            proje_adi = proje.get('ad', '') if proje else ''
            
            if self.export_manager.export_to_excel(self.current_materials, Path(file_path), proje_adi):
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"Malzeme listesi Excel'e aktarÄ±ldÄ±:\n{file_path}")
                self.statusBar().showMessage(f"Excel export tamamlandÄ±: {file_path}")
            else:
                QMessageBox.critical(self, "Hata", "Excel export sÄ±rasÄ±nda bir hata oluÅŸtu.")
    
    def export_materials_pdf(self) -> None:
        """Malzeme listesini PDF'e export et"""
        if not self.current_materials:
            QMessageBox.warning(self, "UyarÄ±", "Ã–nce malzeme listesini hesaplayÄ±nÄ±z.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "PDF'e Kaydet", "", "PDF DosyalarÄ± (*.pdf)"
        )
        
        if file_path:
            proje = self.db.get_project(self.current_project_id) if self.current_project_id else None
            proje_adi = proje.get('ad', '') if proje else ''
            fire_orani = self.fire_spin.value() / 100.0
            
            if self.export_manager.export_to_pdf(self.current_materials, Path(file_path), proje_adi, fire_orani):
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"Malzeme listesi PDF'e aktarÄ±ldÄ±:\n{file_path}")
                self.statusBar().showMessage(f"PDF export tamamlandÄ±: {file_path}")
            else:
                QMessageBox.critical(self, "Hata", "PDF export sÄ±rasÄ±nda bir hata oluÅŸtu.")
    
    def export_materials_supplier(self) -> None:
        """Malzeme listesini tedarikÃ§i formatÄ±nda export et"""
        if not self.current_materials:
            QMessageBox.warning(self, "UyarÄ±", "Ã–nce malzeme listesini hesaplayÄ±nÄ±z.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "TedarikÃ§i FormatÄ±na Kaydet", "", "Metin DosyalarÄ± (*.txt)"
        )
        
        if file_path:
            if self.export_manager.export_supplier_format(self.current_materials, Path(file_path)):
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"Malzeme listesi tedarikÃ§i formatÄ±na aktarÄ±ldÄ±:\n{file_path}")
                self.statusBar().showMessage(f"TedarikÃ§i format export tamamlandÄ±: {file_path}")
            else:
                QMessageBox.critical(self, "Hata", "Export sÄ±rasÄ±nda bir hata oluÅŸtu.")
        
    def create_sablonlar_tab(self, add_to_tabs: bool = True) -> None:
        """Åablonlar sekmesini oluÅŸtur"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Buton barÄ±
        btn_layout = QHBoxLayout()
        
        # Åablon oluÅŸturma butonlarÄ±
        btn_create_empty = QPushButton("â• BoÅŸ Åablon OluÅŸtur")
        btn_create_empty.clicked.connect(self.create_empty_template)
        btn_layout.addWidget(btn_create_empty)
        
        btn_create_from_project = QPushButton("ğŸ“‹ Projeden Åablon OluÅŸtur")
        btn_create_from_project.clicked.connect(self.create_template_from_project)
        btn_layout.addWidget(btn_create_from_project)
        
        btn_create_from_category = QPushButton("ğŸ“‚ Kategori BazlÄ± Åablon OluÅŸtur")
        btn_create_from_category.clicked.connect(self.create_template_from_categories)
        btn_layout.addWidget(btn_create_from_category)
        
        btn_create_from_preset = QPushButton("â­ HazÄ±r Åablonlardan OluÅŸtur")
        btn_create_from_preset.clicked.connect(self.create_template_from_preset)
        btn_layout.addWidget(btn_create_from_preset)
        
        btn_layout.addStretch()
        
        # Åablon iÅŸlem butonlarÄ±
        btn_create_project = QPushButton("ğŸš€ Åablondan Proje OluÅŸtur")
        btn_create_project.clicked.connect(self.create_project_from_template)
        btn_layout.addWidget(btn_create_project)
        
        btn_edit_template = QPushButton("âœï¸ Åablon DÃ¼zenle")
        btn_edit_template.clicked.connect(self.edit_template)
        btn_layout.addWidget(btn_edit_template)
        
        btn_copy_template = QPushButton("ğŸ“‹ Åablon Kopyala")
        btn_copy_template.clicked.connect(self.copy_template)
        btn_layout.addWidget(btn_copy_template)
        
        btn_refresh = QPushButton("ğŸ”„ Yenile")
        btn_refresh.clicked.connect(self.load_templates)
        btn_layout.addWidget(btn_refresh)
        
        btn_delete = QPushButton("ğŸ—‘ï¸ Åablon Sil")
        btn_delete.clicked.connect(self.delete_template)
        btn_delete.setStyleSheet("background-color: #c9184a;")
        btn_layout.addWidget(btn_delete)
        
        layout.addLayout(btn_layout)
        
        # Åablon listesi
        self.template_table = QTableWidget()
        self.template_table.setColumnCount(4)
        self.template_table.setHorizontalHeaderLabels([
            "Åablon AdÄ±", "AÃ§Ä±klama", "OluÅŸturulma Tarihi", "Kalem SayÄ±sÄ±"
        ])
        self.template_table.setAlternatingRowColors(True)
        self.template_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.template_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.template_table.horizontalHeader().setStretchLastSection(True)
        self.template_table.setColumnWidth(0, 250)
        self.template_table.setColumnWidth(1, 300)
        self.template_table.setColumnWidth(2, 150)
        self.template_table.itemDoubleClicked.connect(self.view_template_items)
        self.template_table.itemClicked.connect(self.view_template_items)
        layout.addWidget(self.template_table)
        
        # Åablon kalemleri (seÃ§ili ÅŸablon iÃ§in)
        items_group = QGroupBox("Åablon Kalemleri")
        items_layout = QVBoxLayout()
        
        # Kalem iÅŸlem butonlarÄ±
        items_btn_layout = QHBoxLayout()
        btn_add_item = QPushButton("â• Kalem Ekle")
        btn_add_item.clicked.connect(self.add_template_item)
        items_btn_layout.addWidget(btn_add_item)
        
        btn_edit_item = QPushButton("âœï¸ Kalem DÃ¼zenle")
        btn_edit_item.clicked.connect(self.edit_template_item)
        items_btn_layout.addWidget(btn_edit_item)
        
        btn_delete_item = QPushButton("ğŸ—‘ï¸ Kalem Sil")
        btn_delete_item.clicked.connect(self.delete_template_item)
        btn_delete_item.setStyleSheet("background-color: #c9184a;")
        items_btn_layout.addWidget(btn_delete_item)
        
        items_btn_layout.addStretch()
        items_layout.addLayout(items_btn_layout)
        
        self.template_items_table = QTableWidget()
        self.template_items_table.setColumnCount(8)
        self.template_items_table.setHorizontalHeaderLabels([
            "ID", "Poz No", "TanÄ±m", "Kategori", "Miktar", "Birim", "Birim Fiyat", "Toplam"
        ])
        self.template_items_table.setAlternatingRowColors(True)
        self.template_items_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.template_items_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.template_items_table.horizontalHeader().setStretchLastSection(True)
        self.template_items_table.setColumnHidden(0, True)  # ID kolonunu gizle
        items_layout.addWidget(self.template_items_table)
        
        items_group.setLayout(items_layout)
        layout.addWidget(items_group)
        
        # SeÃ§ili ÅŸablon ID'sini sakla
        self.current_template_id: Optional[int] = None
        
        self.sablonlar_widget = tab
        if add_to_tabs:
            self.tabs.addTab(tab, "ğŸ“‹ Åablonlar")
    
    def create_birim_fiyat_tab(self, add_to_tabs: bool = True) -> None:
        """Birim Fiyat YÃ¶netimi sekmesini oluÅŸtur"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Ãœst panel: Butonlar ve arama
        top_layout = QHBoxLayout()
        
        btn_add = QPushButton("Fiyat Ekle")
        btn_add.clicked.connect(self.add_birim_fiyat)
        top_layout.addWidget(btn_add)
        
        btn_refresh = QPushButton("Yenile")
        btn_refresh.clicked.connect(self.load_birim_fiyatlar)
        top_layout.addWidget(btn_refresh)
        
        btn_edit_fiyat = QPushButton("âœï¸ FiyatÄ± DÃ¼zelt")
        btn_edit_fiyat.clicked.connect(self.edit_birim_fiyat)
        top_layout.addWidget(btn_edit_fiyat)
        
        top_layout.addStretch()
        
        # Filtre
        filter_label = QLabel("Filtre:")
        top_layout.addWidget(filter_label)
        
        self.fiyat_filter_combo = QComboBox()
        self.fiyat_filter_combo.addItems(["TÃ¼mÃ¼", "Sadece Aktif"])
        self.fiyat_filter_combo.setCurrentText("Sadece Aktif")
        self.fiyat_filter_combo.currentTextChanged.connect(self.load_birim_fiyatlar)
        top_layout.addWidget(self.fiyat_filter_combo)
        
        layout.addLayout(top_layout)
        
        # Splitter: Sol tarafta fiyat listesi, saÄŸ tarafta detaylar
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Sol: Birim fiyat listesi
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)
        
        list_title = QLabel("ğŸ’° Birim Fiyat Listesi")
        list_title.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        left_layout.addWidget(list_title)
        
        self.birim_fiyat_table = QTableWidget()
        self.birim_fiyat_table.setColumnCount(6)
        self.birim_fiyat_table.setHorizontalHeaderLabels([
            "Poz No", "Poz TanÄ±mÄ±", "Birim Fiyat", "Tarih", "Kaynak", "Aktif"
        ])
        self.birim_fiyat_table.setAlternatingRowColors(True)
        self.birim_fiyat_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.birim_fiyat_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.birim_fiyat_table.horizontalHeader().setStretchLastSection(True)
        self.birim_fiyat_table.setColumnWidth(0, 120)
        self.birim_fiyat_table.setColumnWidth(1, 300)
        self.birim_fiyat_table.setColumnWidth(2, 120)
        self.birim_fiyat_table.setColumnWidth(3, 120)
        self.birim_fiyat_table.itemDoubleClicked.connect(self.view_fiyat_gecmisi)
        left_layout.addWidget(self.birim_fiyat_table)
        
        splitter.addWidget(left_widget)
        
        # SaÄŸ: Fiyat geÃ§miÅŸi ve karÅŸÄ±laÅŸtÄ±rma
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)
        
        detail_title = QLabel("ğŸ“Š Fiyat GeÃ§miÅŸi ve KarÅŸÄ±laÅŸtÄ±rma")
        detail_title.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        right_layout.addWidget(detail_title)
        
        # Fiyat geÃ§miÅŸi tablosu
        self.fiyat_gecmisi_table = QTableWidget()
        self.fiyat_gecmisi_table.setColumnCount(5)
        self.fiyat_gecmisi_table.setHorizontalHeaderLabels([
            "Tarih", "Birim Fiyat", "Kaynak", "AÃ§Ä±klama", "Aktif"
        ])
        self.fiyat_gecmisi_table.setAlternatingRowColors(True)
        self.fiyat_gecmisi_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.fiyat_gecmisi_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.fiyat_gecmisi_table.horizontalHeader().setStretchLastSection(True)
        right_layout.addWidget(self.fiyat_gecmisi_table)
        
        # KarÅŸÄ±laÅŸtÄ±rma Ã¶zeti
        self.fiyat_karsilastirma_label = QLabel("Bir fiyat seÃ§in veya Ã§ift tÄ±klayÄ±n")
        self.fiyat_karsilastirma_label.setWordWrap(True)
        self.fiyat_karsilastirma_label.setStyleSheet("padding: 10px; background-color: #f5f5f5; border: 1px solid #ddd;")
        right_layout.addWidget(self.fiyat_karsilastirma_label)
        
        splitter.addWidget(right_widget)
        
        splitter.setSizes([500, 500])
        layout.addWidget(splitter)
        
        self.birim_fiyat_widget = tab
        if add_to_tabs:
            self.tabs.addTab(tab, "ğŸ’° Birim Fiyatlar")
    
    def load_birim_fiyatlar(self) -> None:
        """Birim fiyatlarÄ± yÃ¼kle"""
        # Sekme henÃ¼z oluÅŸturulmamÄ±ÅŸsa (lazy loading) yÃ¼kleme yapma
        if not hasattr(self, 'fiyat_filter_combo') or not self._tabs_created.get('birim_fiyat', False):
            return
        
        from PyQt6.QtWidgets import QApplication
        QApplication.processEvents()  # UI'Ä± gÃ¼ncelle
        
        aktif_only = self.fiyat_filter_combo.currentText() == "Sadece Aktif"
        fiyatlar = self.db.get_all_birim_fiyatlar(aktif_only=aktif_only)
        
        self.birim_fiyat_table.setRowCount(len(fiyatlar))
        
        for row, fiyat in enumerate(fiyatlar):
            self.birim_fiyat_table.setItem(row, 0, QTableWidgetItem(fiyat.get('poz_no', '')))
            self.birim_fiyat_table.setItem(row, 1, QTableWidgetItem(fiyat.get('poz_tanim', '')))
            self.birim_fiyat_table.setItem(row, 2, QTableWidgetItem(f"{fiyat.get('birim_fiyat', 0):,.2f} â‚º"))
            tarih = fiyat.get('tarih', '')[:10] if fiyat.get('tarih') else ''
            self.birim_fiyat_table.setItem(row, 3, QTableWidgetItem(tarih))
            self.birim_fiyat_table.setItem(row, 4, QTableWidgetItem(fiyat.get('kaynak', '')))
            aktif_text = "Evet" if fiyat.get('aktif', 0) == 1 else "HayÄ±r"
            self.birim_fiyat_table.setItem(row, 5, QTableWidgetItem(aktif_text))
            
            # ID'yi sakla
            item = self.birim_fiyat_table.item(row, 0)
            if item:
                item.setData(Qt.ItemDataRole.UserRole, fiyat.get('poz_no', ''))
            
            # Her 50 satÄ±rda bir UI'Ä± gÃ¼ncelle
            if row % 50 == 0:
                QApplication.processEvents()
    
    def view_fiyat_gecmisi(self, item: QTableWidgetItem) -> None:
        """Fiyat geÃ§miÅŸini ve karÅŸÄ±laÅŸtÄ±rmayÄ± gÃ¶ster"""
        row = item.row()
        poz_no_item = self.birim_fiyat_table.item(row, 0)
        if not poz_no_item:
            return
        
        poz_no = poz_no_item.data(Qt.ItemDataRole.UserRole)
        if not poz_no:
            poz_no = poz_no_item.text()
        
        # Fiyat geÃ§miÅŸini yÃ¼kle
        gecmis = self.db.get_birim_fiyat_gecmisi(poz_no=poz_no)
        self.fiyat_gecmisi_table.setRowCount(len(gecmis))
        
        for row_idx, fiyat in enumerate(gecmis):
            tarih = fiyat.get('tarih', '')[:10] if fiyat.get('tarih') else ''
            self.fiyat_gecmisi_table.setItem(row_idx, 0, QTableWidgetItem(tarih))
            self.fiyat_gecmisi_table.setItem(row_idx, 1, QTableWidgetItem(f"{fiyat.get('birim_fiyat', 0):,.2f} â‚º"))
            self.fiyat_gecmisi_table.setItem(row_idx, 2, QTableWidgetItem(fiyat.get('kaynak', '')))
            self.fiyat_gecmisi_table.setItem(row_idx, 3, QTableWidgetItem(fiyat.get('aciklama', '')))
            aktif_text = "Evet" if fiyat.get('aktif', 0) == 1 else "HayÄ±r"
            self.fiyat_gecmisi_table.setItem(row_idx, 4, QTableWidgetItem(aktif_text))
        
        # KarÅŸÄ±laÅŸtÄ±rma yap
        karsilastirma = self.db.compare_birim_fiyatlar(poz_no)
    
    def edit_birim_fiyat(self) -> None:
        """SeÃ§ili pozun birim fiyatÄ±nÄ± dÃ¼zelt"""
        current_row = self.birim_fiyat_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen dÃ¼zeltmek istediÄŸiniz bir fiyat satÄ±rÄ±nÄ± seÃ§in")
            return
        
        poz_no_item = self.birim_fiyat_table.item(current_row, 0)
        if not poz_no_item:
            return
        
        poz_no = poz_no_item.data(Qt.ItemDataRole.UserRole)
        if not poz_no:
            poz_no = poz_no_item.text()
        
        # Mevcut fiyatÄ± al
        fiyat_item = self.birim_fiyat_table.item(current_row, 2)
        mevcut_fiyat = 0.0
        if fiyat_item:
            fiyat_text = fiyat_item.text().replace("â‚º", "").replace(",", ".").strip()
            try:
                mevcut_fiyat = float(fiyat_text)
            except:
                pass
        
        # Yeni fiyat gir
        from PyQt6.QtWidgets import QInputDialog
        yeni_fiyat, ok = QInputDialog.getDouble(
            self,
            "Fiyat DÃ¼zelt",
            f"Poz {poz_no} iÃ§in yeni birim fiyatÄ± girin:",
            mevcut_fiyat,
            0.0,
            999999999.99,
            2
        )
        
        if ok and yeni_fiyat > 0:
            # Poz bilgisini al
            poz_data = self.db.get_poz_by_no(poz_no)
            if not poz_data:
                QMessageBox.warning(self, "UyarÄ±", f"Poz {poz_no} bulunamadÄ±")
                return
            
            poz_id = poz_data.get('id')
            
            # Eski aktif fiyatlarÄ± pasif yap
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE birim_fiyatlar SET aktif = 0
                    WHERE poz_id = ? AND aktif = 1
                """, (poz_id,))
            
            # Yeni fiyatÄ± ekle
            fiyat_id = self.db.add_birim_fiyat(
                poz_id=poz_id,
                poz_no=poz_no,
                birim_fiyat=yeni_fiyat,
                kaynak='Manuel DÃ¼zeltme',
                aciklama=f'Eski fiyat: {mevcut_fiyat:,.2f} â‚º'
            )
            
            if fiyat_id:
                # Poz'un resmi_fiyat'Ä±nÄ± da gÃ¼ncelle
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        UPDATE pozlar SET resmi_fiyat = ?
                        WHERE poz_no = ?
                    """, (yeni_fiyat, poz_no))
                
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"Poz {poz_no} iÃ§in birim fiyat {yeni_fiyat:,.2f} â‚º olarak gÃ¼ncellendi")
                self.load_birim_fiyatlar()
                # Fiyat geÃ§miÅŸini de yenile
                poz_no_item = self.birim_fiyat_table.item(current_row, 0)
                if poz_no_item:
                    self.view_fiyat_gecmisi(poz_no_item)
            else:
                QMessageBox.warning(self, "Hata", "Fiyat gÃ¼ncellenirken bir hata oluÅŸtu")
        
        if karsilastirma['fiyat_sayisi'] > 0:
            text = f"ğŸ“Š Poz: {poz_no}\n\n"
            text += f"ğŸ’° Toplam Fiyat KaydÄ±: {karsilastirma['fiyat_sayisi']}\n"
            text += f"ğŸ“‰ En DÃ¼ÅŸÃ¼k: {karsilastirma['en_dusuk']:,.2f} â‚º\n"
            text += f"ğŸ“ˆ En YÃ¼ksek: {karsilastirma['en_yuksek']:,.2f} â‚º\n"
            text += f"ğŸ“Š Ortalama: {karsilastirma['ortalama']:,.2f} â‚º\n"
            text += f"ğŸ“ Fark: {karsilastirma['en_yuksek'] - karsilastirma['en_dusuk']:,.2f} â‚º\n\n"
            
            if karsilastirma['kaynaklar']:
                text += "ğŸ“‹ Kaynaklar:\n"
                for kaynak, fiyatlar in karsilastirma['kaynaklar'].items():
                    ortalama_kaynak = sum(fiyatlar) / len(fiyatlar)
                    text += f"  â€¢ {kaynak}: {ortalama_kaynak:,.2f} â‚º ({len(fiyatlar)} kayÄ±t)\n"
        else:
            text = f"Poz {poz_no} iÃ§in henÃ¼z fiyat kaydÄ± yok."
        
        self.fiyat_karsilastirma_label.setText(text)
    
    def add_birim_fiyat(self) -> None:
        """Birim fiyat ekle dialogu"""
        from PyQt6.QtWidgets import QDialog, QFormLayout, QDateEdit
        from PyQt6.QtCore import QDate
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Birim Fiyat Ekle")
        dialog.setMinimumWidth(400)
        
        layout = QFormLayout(dialog)
        
        poz_no_input = QLineEdit()
        poz_no_input.setPlaceholderText("Ã–rn: 03.001")
        layout.addRow("Poz No:", poz_no_input)
        
        birim_fiyat_spin = QDoubleSpinBox()
        birim_fiyat_spin.setMaximum(999999999)
        birim_fiyat_spin.setDecimals(2)
        birim_fiyat_spin.setPrefix("â‚º ")
        layout.addRow("Birim Fiyat:", birim_fiyat_spin)
        
        tarih_input = QDateEdit()
        tarih_input.setDate(QDate.currentDate())
        tarih_input.setCalendarPopup(True)
        layout.addRow("Tarih:", tarih_input)
        
        kaynak_input = QLineEdit()
        kaynak_input.setPlaceholderText("Ã–rn: TedarikÃ§i A, Resmi Fiyat")
        layout.addRow("Kaynak:", kaynak_input)
        
        aciklama_input = QTextEdit()
        aciklama_input.setMaximumHeight(80)
        layout.addRow("AÃ§Ä±klama:", aciklama_input)
        
        btn_layout = QHBoxLayout()
        btn_ok = QPushButton("Kaydet")
        btn_ok.clicked.connect(dialog.accept)
        btn_cancel = QPushButton("Ä°ptal")
        btn_cancel.clicked.connect(dialog.reject)
        btn_layout.addWidget(btn_ok)
        btn_layout.addWidget(btn_cancel)
        layout.addRow(btn_layout)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            poz_no = poz_no_input.text().strip()
            birim_fiyat = birim_fiyat_spin.value()
            tarih = tarih_input.date().toString("yyyy-MM-dd")
            kaynak = kaynak_input.text().strip()
            aciklama = aciklama_input.toPlainText().strip()
            
            if not poz_no:
                QMessageBox.warning(self, "UyarÄ±", "Poz numarasÄ± gereklidir")
                return
            
            if birim_fiyat <= 0:
                QMessageBox.warning(self, "UyarÄ±", "Birim fiyat 0'dan bÃ¼yÃ¼k olmalÄ±dÄ±r")
                return
            
            fiyat_id = self.db.add_birim_fiyat(
                poz_no=poz_no,
                birim_fiyat=birim_fiyat,
                tarih=tarih,
                kaynak=kaynak,
                aciklama=aciklama
            )
            
            if fiyat_id:
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", "Birim fiyat eklendi")
                self.load_birim_fiyatlar()
                self.statusBar().showMessage(f"Birim fiyat eklendi: {poz_no}")
            else:
                QMessageBox.critical(self, "Hata", "Birim fiyat eklenirken bir hata oluÅŸtu")
    
    def create_ihale_tab(self, add_to_tabs: bool = True) -> None:
        """Ä°hale DosyasÄ± HazÄ±rlama sekmesini oluÅŸtur"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Ãœst panel: Ä°hale seÃ§imi ve poz arama
        top_layout = QHBoxLayout()
        
        # Ä°hale seÃ§imi
        ihale_label = QLabel("Ä°hale:")
        top_layout.addWidget(ihale_label)
        
        self.ihale_combo = QComboBox()
        self.ihale_combo.setMinimumWidth(200)
        self.ihale_combo.currentIndexChanged.connect(self.on_ihale_changed)
        top_layout.addWidget(self.ihale_combo)
        
        btn_new_ihale = QPushButton("Yeni Ä°hale")
        btn_new_ihale.clicked.connect(self.new_ihale)
        top_layout.addWidget(btn_new_ihale)
        
        top_layout.addStretch()
        
        # Poz arama
        search_label = QLabel("Poz Ara:")
        top_layout.addWidget(search_label)
        
        self.ihale_poz_search = QLineEdit()
        self.ihale_poz_search.setPlaceholderText("Poz no veya tanÄ±m ara...")
        self.ihale_poz_search.setMinimumWidth(200)
        self.ihale_poz_search.textChanged.connect(self.on_ihale_poz_search)
        top_layout.addWidget(self.ihale_poz_search)
        
        btn_add_poz = QPushButton("Listeye Ekle")
        btn_add_poz.clicked.connect(self.add_poz_to_ihale)
        top_layout.addWidget(btn_add_poz)
        
        layout.addLayout(top_layout)
        
        # Splitter: Sol tarafta poz arama sonuÃ§larÄ±, saÄŸ tarafta ihale kalemleri
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Sol: Poz arama sonuÃ§larÄ±
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)
        
        poz_title = QLabel("ğŸ” Poz Arama SonuÃ§larÄ±")
        poz_title.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        left_layout.addWidget(poz_title)
        
        self.ihale_poz_results_table = QTableWidget()
        self.ihale_poz_results_table.setColumnCount(4)
        self.ihale_poz_results_table.setHorizontalHeaderLabels(["Poz No", "TanÄ±m", "Birim", "Birim Fiyat"])
        self.ihale_poz_results_table.setAlternatingRowColors(True)
        self.ihale_poz_results_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.ihale_poz_results_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.ihale_poz_results_table.horizontalHeader().setStretchLastSection(True)
        self.ihale_poz_results_table.setColumnWidth(0, 120)
        self.ihale_poz_results_table.setColumnWidth(1, 300)
        self.ihale_poz_results_table.setColumnWidth(2, 80)
        self.ihale_poz_results_table.setColumnWidth(3, 120)
        self.ihale_poz_results_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.ihale_poz_results_table.itemDoubleClicked.connect(self.add_selected_poz_to_ihale)
        # Tablo gÃ¶rÃ¼nÃ¼rlÃ¼ÄŸÃ¼nÃ¼ garanti et
        self.ihale_poz_results_table.setVisible(True)
        left_layout.addWidget(self.ihale_poz_results_table)
        
        splitter.addWidget(left_widget)
        
        # SaÄŸ: Ä°hale kalemleri
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)
        
        kalem_title = QLabel("ğŸ“‹ Ä°hale Kalem Listesi")
        kalem_title.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        right_layout.addWidget(kalem_title)
        
        # Butonlar
        kalem_btn_layout = QHBoxLayout()
        
        btn_delete_kalem = QPushButton("Kalem Sil")
        btn_delete_kalem.clicked.connect(self.delete_ihale_kalem)
        btn_delete_kalem.setStyleSheet("background-color: #c9184a;")
        
        btn_edit_tanim = QPushButton("âœï¸ TanÄ±mÄ± DÃ¼zelt")
        btn_edit_tanim.clicked.connect(self.edit_ihale_tanim)
        right_layout.addWidget(btn_edit_tanim)
        kalem_btn_layout.addWidget(btn_delete_kalem)
        
        btn_export = QPushButton("Ä°hale DosyasÄ± OluÅŸtur (PDF)")
        btn_export.clicked.connect(self.export_ihale_pdf)
        kalem_btn_layout.addWidget(btn_export)
        
        btn_export_excel = QPushButton("Ä°hale DosyasÄ± OluÅŸtur (Excel)")
        btn_export_excel.clicked.connect(self.export_ihale_excel)
        kalem_btn_layout.addWidget(btn_export_excel)
        
        kalem_btn_layout.addStretch()
        
        # KDV oranÄ± seÃ§imi
        kdv_label = QLabel("KDV:")
        kalem_btn_layout.addWidget(kdv_label)
        self.ihale_kdv_rate = QComboBox()
        self.ihale_kdv_rate.addItems(["%1", "%10", "%20"])
        self.ihale_kdv_rate.setCurrentText("%20")
        self.ihale_kdv_rate.currentTextChanged.connect(self.load_ihale_kalemleri)
        kalem_btn_layout.addWidget(self.ihale_kdv_rate)
        
        # Toplam etiketleri
        self.ihale_total_label = QLabel("Toplam (KDV HariÃ§): 0.00 â‚º")
        self.ihale_total_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        kalem_btn_layout.addWidget(self.ihale_total_label)
        
        self.ihale_total_kdv_label = QLabel("Toplam (KDV Dahil): 0.00 â‚º")
        self.ihale_total_kdv_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        self.ihale_total_kdv_label.setStyleSheet("color: #00BFFF;")
        kalem_btn_layout.addWidget(self.ihale_total_kdv_label)
        
        right_layout.addLayout(kalem_btn_layout)
        
        self.ihale_kalem_table = QTableWidget()
        self.ihale_kalem_table.setColumnCount(7)
        self.ihale_kalem_table.setHorizontalHeaderLabels([
            "SÄ±ra", "Poz No", "TanÄ±m", "Birim Miktar", "Birim", "Birim Fiyat", "Toplam"
        ])
        self.ihale_kalem_table.setAlternatingRowColors(True)
        self.ihale_kalem_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.ihale_kalem_table.horizontalHeader().setStretchLastSection(True)
        self.ihale_kalem_table.setColumnWidth(0, 50)
        self.ihale_kalem_table.setColumnWidth(1, 120)
        self.ihale_kalem_table.setColumnWidth(2, 400)  # TanÄ±m sÃ¼tunu geniÅŸletildi
        self.ihale_kalem_table.setColumnWidth(3, 120)
        self.ihale_kalem_table.setColumnWidth(4, 80)
        self.ihale_kalem_table.setColumnWidth(5, 120)
        # TanÄ±m sÃ¼tununu geniÅŸletilebilir yap
        self.ihale_kalem_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)
        # Birim Miktar ve Birim Fiyat sÃ¼tunlarÄ± dÃ¼zenlenebilir
        self.ihale_kalem_table.itemChanged.connect(self.on_ihale_kalem_changed)
        # TanÄ±m sÃ¼tununa Ã§ift tÄ±klayÄ±nca tam metni gÃ¶ster
        self.ihale_kalem_table.itemDoubleClicked.connect(self.show_full_tanim)
        right_layout.addWidget(self.ihale_kalem_table)
        
        splitter.addWidget(right_widget)
        
        splitter.setSizes([400, 600])
        layout.addWidget(splitter)
        
        # Mevcut ihale ID'si
        self.current_ihale_id: Optional[int] = None
        
        self.ihale_widget = tab
        if add_to_tabs:
            self.tabs.addTab(tab, "ğŸ“„ Ä°hale DosyasÄ±")
    
    def create_menu_bar(self) -> None:
        """MenÃ¼ Ã§ubuÄŸunu oluÅŸtur"""
        menubar = self.menuBar()
        
        # Dosya menÃ¼sÃ¼
        file_menu = menubar.addMenu("Dosya")
        
        # Yeni proje
        new_action = file_menu.addAction("Yeni Proje")
        new_action.setShortcut("Ctrl+N")
        new_action.triggered.connect(self.new_project)
        
        # Proje aÃ§
        open_action = file_menu.addAction("Proje AÃ§")
        open_action.setShortcut("Ctrl+O")
        open_action.triggered.connect(self.open_project)
        
        file_menu.addSeparator()
        
        # Yedekleme menÃ¼sÃ¼
        backup_menu = file_menu.addMenu("Yedekleme")
        
        # Proje yedekle
        backup_project_action = backup_menu.addAction("Projeyi Yedekle")
        backup_project_action.triggered.connect(self.backup_current_project)
        
        # TÃ¼m projeleri yedekle
        backup_all_action = backup_menu.addAction("TÃ¼m Projeleri Yedekle")
        backup_all_action.triggered.connect(self.backup_all_projects)
        
        backup_menu.addSeparator()
        
        # Geri yÃ¼kle
        restore_action = backup_menu.addAction("Yedekten Geri YÃ¼kle")
        restore_action.triggered.connect(self.restore_project)
        
        # Versiyonlama menÃ¼sÃ¼
        version_menu = file_menu.addMenu("Versiyonlama")
        
        # Versiyon oluÅŸtur
        create_version_action = version_menu.addAction("Versiyon OluÅŸtur")
        create_version_action.triggered.connect(self.create_project_version)
        
        # VersiyonlarÄ± gÃ¶rÃ¼ntÃ¼le
        view_versions_action = version_menu.addAction("VersiyonlarÄ± GÃ¶rÃ¼ntÃ¼le")
        view_versions_action.triggered.connect(self.view_project_versions)
        
        # Versiyondan geri yÃ¼kle
        restore_version_action = version_menu.addAction("Versiyondan Geri YÃ¼kle")
        restore_version_action.triggered.connect(self.restore_from_version)
        
        file_menu.addSeparator()
        
        # Ã‡Ä±kÄ±ÅŸ
        exit_action = file_menu.addAction("Ã‡Ä±kÄ±ÅŸ")
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        
        # Veri menÃ¼sÃ¼
        data_menu = menubar.addMenu("Veri")
        load_pozlar_action = data_menu.addAction("PozlarÄ± YÃ¼kle")
        load_pozlar_action.triggered.connect(self.load_pozlar)
        data_menu.addSeparator()
        
        # Excel Import
        excel_import_action = data_menu.addAction("Excel'den Kalem Ä°Ã§e Aktar")
        excel_import_action.triggered.connect(self.import_from_excel)
        
        # PDF Import
        pdf_import_action = data_menu.addAction("PDF'den Birim Fiyat Ä°Ã§e Aktar")
        pdf_import_action.triggered.connect(self.import_from_pdf)
        
        # PDF Import Temizle
        pdf_clear_action = data_menu.addAction("PDF'den Eklenen PozlarÄ± Temizle")
        pdf_clear_action.triggered.connect(self.clear_pdf_imported_data)
        
        data_menu.addSeparator()
        check_pozlar_action = data_menu.addAction("Poz Durumunu Kontrol Et")
        check_pozlar_action.triggered.connect(self.check_pozlar_status)
        
        # AraÃ§lar menÃ¼sÃ¼
        tools_menu = menubar.addMenu("AraÃ§lar")
        
        # Birim dÃ¶nÃ¼ÅŸtÃ¼rÃ¼cÃ¼
        unit_converter_action = tools_menu.addAction("Birim DÃ¶nÃ¼ÅŸtÃ¼rÃ¼cÃ¼")
        unit_converter_action.triggered.connect(self.show_unit_converter)
        
        tools_menu.addSeparator()
        
        # Otomatik fire oranÄ± hesaplama
        auto_fire_action = tools_menu.addAction("Otomatik Fire OranÄ± Hesapla")
        auto_fire_action.triggered.connect(self.calculate_auto_fire_rates)
        
        # YardÄ±m menÃ¼sÃ¼
        help_menu = menubar.addMenu("YardÄ±m")
        about_action = help_menu.addAction("HakkÄ±nda")
        about_action.triggered.connect(self.show_about)
        
    # Proje Ä°ÅŸlemleri
    def load_data_async(self) -> None:
        """VeritabanÄ± verilerini async yÃ¼kle"""
        # Projeleri async yÃ¼kle
        self.initial_data_thread = InitialDataLoaderThread(self.db)
        self.initial_data_thread.projects_loaded.connect(self.on_projects_loaded)
        self.initial_data_thread.start()
    
    @pyqtSlot(list)
    def on_projects_loaded(self, projects: List[Dict[str, Any]]) -> None:
        """Projeler yÃ¼klendiÄŸinde Ã§aÄŸrÄ±lÄ±r"""
        self.project_tree.clear()
        for project in projects:
            item = QTreeWidgetItem(self.project_tree)
            item.setText(0, project['ad'])
            item.setData(0, Qt.ItemDataRole.UserRole, project['id'])
        self.statusBar().showMessage(f"{len(projects)} proje yÃ¼klendi")
    
    def load_projects(self) -> None:
        """Projeleri yÃ¼kle (sync versiyon - eski)"""
        self.project_tree.clear()
        projects = self.db.get_all_projects()
        
        for project in projects:
            item = QTreeWidgetItem(self.project_tree)
            item.setText(0, project['ad'])
            item.setData(0, Qt.ItemDataRole.UserRole, project['id'])
            
    def on_project_selected(self, item: QTreeWidgetItem, column: int) -> None:
        """Proje seÃ§ildiÄŸinde"""
        if not item:
            return
            
        project_id = item.data(0, Qt.ItemDataRole.UserRole)
        if project_id:
            self.current_project_id = project_id
            # Projeyi seÃ§ili olarak iÅŸaretle
            self.project_tree.setCurrentItem(item)
            # Verileri yÃ¼kle (sadece sekmeler oluÅŸturulmuÅŸsa)
            if hasattr(self, 'metraj_table'):
                self.load_metraj_data()
            if hasattr(self, 'taseron_table'):
                self.load_taseron_data()
            if hasattr(self, 'ozet_kalem_label'):
                self.update_proje_ozet()
            self.load_project_notes()
            self.statusBar().showMessage(f"Proje seÃ§ildi: {item.text(0)}")
        else:
            self.statusBar().showMessage("GeÃ§ersiz proje seÃ§imi")
            
    def new_project(self) -> None:
        """Yeni proje oluÅŸtur"""
        from PyQt6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(
            self, "Yeni Proje", "Proje AdÄ±:"
        )
        if ok and name:
            project_id = self.db.create_project(name)
            if project_id:
                # Projeleri async yÃ¼kle
                self.load_data_async()
                # Yeni oluÅŸturulan projeyi otomatik seÃ§ (biraz bekle)
                from PyQt6.QtWidgets import QApplication
                QApplication.processEvents()
                # Proje listesi yÃ¼klendikten sonra seÃ§
                from PyQt6.QtCore import QTimer
                def select_new_project():
                    for i in range(self.project_tree.topLevelItemCount()):
                        item = self.project_tree.topLevelItem(i)
                        if item and item.data(0, Qt.ItemDataRole.UserRole) == project_id:
                            self.project_tree.setCurrentItem(item)
                            self.on_project_selected(item, 0)
                            break
                QTimer.singleShot(100, select_new_project)  # 100ms sonra seÃ§
                self.statusBar().showMessage(f"Yeni proje oluÅŸturuldu: {name}")
                
    def open_project(self) -> None:
        """Proje aÃ§ (ÅŸimdilik bilgi mesajÄ±)"""
        QMessageBox.information(
            self, "Bilgi", "Proje aÃ§ma Ã¶zelliÄŸi yakÄ±nda eklenecek"
        )
        
    def show_project_context_menu(self, position) -> None:
        """Proje aÄŸacÄ±nda saÄŸ tÄ±klama menÃ¼sÃ¼"""
        item = self.project_tree.itemAt(position)
        if not item:
            return
            
        menu = QMenu(self)
        
        # Proje seÃ§
        select_action = menu.addAction("Projeyi SeÃ§")
        select_action.triggered.connect(lambda: self.on_project_selected(item, 0))
        
        menu.addSeparator()
        
        # Proje sil
        delete_action = menu.addAction("Projeyi Sil")
        delete_action.triggered.connect(lambda: self.delete_project(item))
        delete_action.setStyleSheet("color: #c9184a;")
        
        menu.exec(self.project_tree.mapToGlobal(position))
        
    def delete_selected_project(self) -> None:
        """SeÃ§ili projeyi sil"""
        current_item = self.project_tree.currentItem()
        if not current_item:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen silmek iÃ§in bir proje seÃ§in")
            return
            
        self.delete_project(current_item)
        
    def delete_project(self, item: QTreeWidgetItem) -> None:
        """Projeyi sil"""
        if not item:
            return
            
        project_id = item.data(0, Qt.ItemDataRole.UserRole)
        project_name = item.text(0)
        
        if not project_id:
            QMessageBox.warning(self, "UyarÄ±", "GeÃ§ersiz proje seÃ§imi")
            return
            
        # Onay dialogu
        reply = QMessageBox.question(
            self, "Proje Silme OnayÄ±",
            f"'{project_name}' projesini silmek istediÄŸinize emin misiniz?\n\n"
            "âš ï¸ UYARI: Bu iÅŸlem geri alÄ±namaz!\n"
            "Projeye ait tÃ¼m metraj kalemleri ve taÅŸeron teklifleri de silinecektir.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                if self.db.delete_project(project_id):
                    # EÄŸer silinen proje seÃ§iliyse, seÃ§imi temizle
                    if self.current_project_id == project_id:
                        self.current_project_id = None
                        self.metraj_table.setRowCount(0)
                        self.taseron_table.setRowCount(0)
                        self.total_label.setText("Toplam: 0.00 â‚º")
                        
                    # Proje listesini yenile
                    self.load_projects()
                    self.statusBar().showMessage(f"Proje silindi: {project_name}")
                else:
                    QMessageBox.warning(self, "UyarÄ±", "Proje silinirken bir hata oluÅŸtu")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Proje silinirken hata oluÅŸtu:\n{str(e)}")
        
    # Metraj Ä°ÅŸlemleri
    def load_metraj_data(self) -> None:
        """Metraj verilerini yÃ¼kle"""
        # Sekme henÃ¼z oluÅŸturulmamÄ±ÅŸsa (lazy loading) yÃ¼kleme yapma
        if not hasattr(self, 'metraj_table') or not self._tabs_created.get('metraj', False):
            return
        
        if not self.current_project_id:
            return
        
        from PyQt6.QtWidgets import QApplication
        QApplication.processEvents()  # UI'Ä± gÃ¼ncelle
            
        items = self.db.get_project_metraj(self.current_project_id)
        QApplication.processEvents()  # UI'Ä± gÃ¼ncelle
        self.metraj_table.setRowCount(len(items))
        
        total = 0.0
        for row, item in enumerate(items):
            self.metraj_table.setItem(row, 0, QTableWidgetItem(str(item['id'])))
            self.metraj_table.setItem(row, 1, QTableWidgetItem(item.get('poz_no', '')))
            self.metraj_table.setItem(row, 2, QTableWidgetItem(item['tanim']))
            self.metraj_table.setItem(row, 3, QTableWidgetItem(str(item['miktar'])))
            self.metraj_table.setItem(row, 4, QTableWidgetItem(item['birim']))
            self.metraj_table.setItem(row, 5, QTableWidgetItem(f"{item['birim_fiyat']:.2f}"))
            self.metraj_table.setItem(row, 6, QTableWidgetItem(f"{item['toplam']:.2f}"))
            total += item['toplam']
            
            # Her 50 satÄ±rda bir UI'Ä± gÃ¼ncelle
            if row % 50 == 0:
                QApplication.processEvents()
            
        # KDV hesaplama
        kdv_rate_text = self.metraj_kdv_rate.currentText().replace("%", "")
        kdv_rate = float(kdv_rate_text)
        kdv_hesap = self.calculator.calculate_with_kdv(total, kdv_rate)
        
        self.total_label.setText(f"Toplam (KDV HariÃ§): {total:,.2f} â‚º")
        self.total_kdv_label.setText(f"Toplam (KDV %{kdv_rate_text} Dahil): {kdv_hesap['kdv_dahil']:,.2f} â‚º")
        
        # SeÃ§ili satÄ±r yoksa malzeme tablosunu temizle
        if self.metraj_table.currentRow() < 0:
            self.metraj_malzeme_table.setRowCount(0)
            self.metraj_malzeme_total.setText("Toplam: 0.00 â‚º")
            self.metraj_fire_info.setText("")
    
    def on_metraj_item_selected(self) -> None:
        """Metraj tablosunda bir satÄ±r seÃ§ildiÄŸinde malzeme detaylarÄ±nÄ± gÃ¶ster"""
        current_row = self.metraj_table.currentRow()
        
        if current_row < 0:
            self.metraj_malzeme_table.setRowCount(0)
            self.metraj_malzeme_total.setText("Toplam: 0.00 â‚º")
            self.metraj_fire_info.setText("")
            return
        
        # SeÃ§ili satÄ±rdan poz bilgilerini al
        poz_no_item = self.metraj_table.item(current_row, 1)
        miktar_item = self.metraj_table.item(current_row, 3)
        
        if not poz_no_item or not miktar_item:
            self.metraj_malzeme_table.setRowCount(0)
            return
        
        poz_no = poz_no_item.text()
        miktar_text = miktar_item.text()
        
        if not poz_no or not miktar_text:
            self.metraj_malzeme_table.setRowCount(0)
            return
        
        try:
            miktar = float(miktar_text)
            
            # Poz bazlÄ± fire oranÄ±nÄ± al
            poz = self.db.get_poz(poz_no)
            if not poz:
                self.metraj_malzeme_table.setRowCount(0)
                self.metraj_fire_info.setText("âš ï¸ Poz bulunamadÄ±")
                return
            
            fire_orani = poz.get('fire_orani', 0.05)
            
            # Malzemeleri hesapla
            materials = self.material_calculator.calculate_materials_for_poz_no(
                poz_no, miktar, fire_orani_override=None  # Poz bazlÄ± fire oranÄ± kullan
            )
            
            if not materials:
                self.metraj_malzeme_table.setRowCount(0)
                self.metraj_fire_info.setText(
                    f"â„¹ï¸ Bu poz iÃ§in malzeme formÃ¼lÃ¼ tanÄ±mlanmamÄ±ÅŸ. "
                    f"Fire oranÄ±: %{fire_orani*100:.2f}"
                )
                self.metraj_malzeme_total.setText("Toplam: 0.00 â‚º")
                return
            
            # Malzeme tablosunu doldur
            self.metraj_malzeme_table.setRowCount(len(materials))
            
            malzeme_total = 0.0
            
            for row, material in enumerate(materials):
                # Malzeme adÄ±
                item = QTableWidgetItem(material.get('malzeme_adi', ''))
                self.metraj_malzeme_table.setItem(row, 0, item)
                
                # Miktar (fire dahil)
                miktar_val = material.get('miktar', 0)
                item = QTableWidgetItem(f"{miktar_val:,.2f}")
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.metraj_malzeme_table.setItem(row, 1, item)
                
                # Birim
                item = QTableWidgetItem(material.get('birim', ''))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.metraj_malzeme_table.setItem(row, 2, item)
                
                # Birim fiyat (veritabanÄ±ndan Ã§ek)
                malzeme_id = material.get('malzeme_id')
                birim_fiyat = 0.0
                if malzeme_id:
                    malzeme_info = self.db.get_malzeme(malzeme_id)
                    if malzeme_info:
                        birim_fiyat = malzeme_info.get('birim_fiyat', 0.0)
                
                # Birim fiyat dÃ¼zenlenebilir olmalÄ±
                item = QTableWidgetItem(f"{birim_fiyat:,.2f}")
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                item.setData(Qt.ItemDataRole.UserRole, malzeme_id)  # Malzeme ID'sini sakla
                item.setData(Qt.ItemDataRole.UserRole + 1, miktar_val)  # MiktarÄ± sakla
                self.metraj_malzeme_table.setItem(row, 3, item)
                
                # Toplam (hesaplanmÄ±ÅŸ)
                toplam = miktar_val * birim_fiyat
                malzeme_total += toplam
                item = QTableWidgetItem(f"{toplam:,.2f} â‚º")
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Toplam dÃ¼zenlenemez
                self.metraj_malzeme_table.setItem(row, 4, item)
            
            # Toplam ve fire bilgisi
            self.metraj_malzeme_total.setText(f"Toplam: {malzeme_total:,.2f} â‚º")
            self.metraj_fire_info.setText(
                f"â„¹ï¸ Fire oranÄ±: %{fire_orani*100:.2f} (Poz bazlÄ± - LiteratÃ¼r deÄŸeri) | "
                f"Ä°ÅŸ miktarÄ±: {miktar:,.2f} {poz.get('birim', '')}"
            )
            
        except Exception as e:
            print(f"Malzeme hesaplama hatasÄ±: {e}")
            self.metraj_malzeme_table.setRowCount(0)
            self.metraj_fire_info.setText(f"âš ï¸ Hata: {str(e)}")
    
    def on_malzeme_fiyat_changed(self, row: int, column: int) -> None:
        """Malzeme birim fiyatÄ± deÄŸiÅŸtiÄŸinde toplamÄ± gÃ¼ncelle"""
        if column != 3:  # Sadece birim fiyat sÃ¼tunu (3. sÃ¼tun)
            return
        
        try:
            # Birim fiyatÄ± al
            fiyat_item = self.metraj_malzeme_table.item(row, 3)
            if not fiyat_item:
                return
            
            # Fiyat metnini temizle (â‚º iÅŸareti ve boÅŸluklarÄ± kaldÄ±r)
            fiyat_text = fiyat_item.text().replace("â‚º", "").replace(",", ".").strip()
            birim_fiyat = float(fiyat_text) if fiyat_text else 0.0
            
            # MiktarÄ± al (UserRole + 1'den)
            miktar = fiyat_item.data(Qt.ItemDataRole.UserRole + 1)
            if miktar is None:
                # Miktar sÃ¼tunundan al
                miktar_item = self.metraj_malzeme_table.item(row, 1)
                if miktar_item:
                    miktar_text = miktar_item.text().replace(",", ".").strip()
                    miktar = float(miktar_text) if miktar_text else 0.0
                else:
                    miktar = 0.0
            
            # ToplamÄ± hesapla
            toplam = miktar * birim_fiyat
            
            # Toplam sÃ¼tununu gÃ¼ncelle
            toplam_item = QTableWidgetItem(f"{toplam:,.2f} â‚º")
            toplam_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            toplam_item.setFlags(toplam_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.metraj_malzeme_table.setItem(row, 4, toplam_item)
            
            # Birim fiyatÄ± formatla (â‚º iÅŸareti olmadan)
            fiyat_item.setText(f"{birim_fiyat:,.2f}")
            
            # Genel toplamÄ± gÃ¼ncelle
            self.update_malzeme_total()
            
        except (ValueError, TypeError) as e:
            print(f"Fiyat gÃ¼ncelleme hatasÄ±: {e}")
    
    def update_malzeme_total(self) -> None:
        """Malzeme tablosundaki toplam maliyeti gÃ¼ncelle"""
        total = 0.0
        for row in range(self.metraj_malzeme_table.rowCount()):
            toplam_item = self.metraj_malzeme_table.item(row, 4)
            if toplam_item:
                toplam_text = toplam_item.text().replace("â‚º", "").replace(",", ".").strip()
                try:
                    total += float(toplam_text) if toplam_text else 0.0
                except ValueError:
                    pass
        
        self.metraj_malzeme_total.setText(f"Toplam: {total:,.2f} â‚º")
        
    def add_metraj_item(self) -> None:
        """Metraj kalemi ekle"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
            
        # Dialog penceresini aÃ§
        dialog = MetrajItemDialog(self.db, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_data()
            
            # VeritabanÄ±na ekle
            try:
                item_id = self.db.add_metraj_kalem(
                    proje_id=self.current_project_id,
                    tanim=data['tanim'],
                    miktar=data['miktar'],
                    birim=data['birim'],
                    birim_fiyat=data['birim_fiyat'],
                    poz_no=data['poz_no'] if data['poz_no'] else '',
                    kategori=data['kategori'] if data['kategori'] else ''
                )
                
                if item_id:
                    self.load_metraj_data()
                    # Yeni kalem eklendikten sonra seÃ§ili satÄ±rÄ± gÃ¼ncelle
                    if self.metraj_table.rowCount() > 0:
                        self.metraj_table.selectRow(self.metraj_table.rowCount() - 1)
                    self.statusBar().showMessage("Kalem baÅŸarÄ±yla eklendi")
                else:
                    QMessageBox.warning(self, "UyarÄ±", "Kalem eklenirken bir hata oluÅŸtu")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Kalem eklenirken hata oluÅŸtu:\n{str(e)}")
        
    def edit_metraj_item(self) -> None:
        """Metraj kalemi dÃ¼zenle"""
        current_row = self.metraj_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen dÃ¼zenlemek iÃ§in bir satÄ±r seÃ§in")
            return
            
        # SeÃ§ili kalemin ID'sini al
        item_id = int(self.metraj_table.item(current_row, 0).text())
        
        # Kalem verilerini getir
        try:
            items = self.db.get_project_metraj(self.current_project_id)
            item_data = next((item for item in items if item['id'] == item_id), None)
            
            if not item_data:
                QMessageBox.warning(self, "UyarÄ±", "Kalem bulunamadÄ±")
                return
                
            # Dialog penceresini aÃ§
            dialog = MetrajItemDialog(self.db, self, item_data)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                data = dialog.get_data()
                
                # VeritabanÄ±nÄ± gÃ¼ncelle
                if self.db.update_metraj_kalem(
                    item_id=item_id,
                    tanim=data['tanim'],
                    miktar=data['miktar'],
                    birim=data['birim'],
                    birim_fiyat=data['birim_fiyat'],
                    poz_no=data['poz_no'] if data['poz_no'] else '',
                    kategori=data['kategori'] if data['kategori'] else '',
                    notlar=data['notlar'] if data.get('notlar') else ''
                ):
                    self.load_metraj_data()
                    self.update_proje_ozet()  # Ã–zeti gÃ¼ncelle
                    self.statusBar().showMessage("Kalem baÅŸarÄ±yla gÃ¼ncellendi")
                else:
                    QMessageBox.warning(self, "UyarÄ±", "Kalem gÃ¼ncellenirken bir hata oluÅŸtu")
                    
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kalem dÃ¼zenlenirken hata oluÅŸtu:\n{str(e)}")
            
    def delete_metraj_item(self) -> None:
        """Metraj kalemi sil"""
        current_row = self.metraj_table.currentRow()
        if current_row >= 0:
            reply = QMessageBox.question(
                self, "Onay", "Bu kalemi silmek istediÄŸinize emin misiniz?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                item_id = int(self.metraj_table.item(current_row, 0).text())
                if self.db.delete_item(item_id):
                    self.load_metraj_data()
                    self.update_proje_ozet()  # Ã–zeti gÃ¼ncelle
                    self.statusBar().showMessage("Kalem silindi")
        else:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen silmek iÃ§in bir satÄ±r seÃ§in")
            
    # TaÅŸeron Ä°ÅŸlemleri
    def load_taseron_data(self) -> None:
        """TaÅŸeron verilerini yÃ¼kle"""
        # Sekme henÃ¼z oluÅŸturulmamÄ±ÅŸsa (lazy loading) yÃ¼kleme yapma
        if not hasattr(self, 'taseron_table') or not self._tabs_created.get('taseron', False):
            return
        
        if not self.current_project_id:
            return
        
        from PyQt6.QtWidgets import QApplication
        QApplication.processEvents()  # UI'Ä± gÃ¼ncelle
            
        offers = self.db.get_taseron_teklifleri(self.current_project_id)
        QApplication.processEvents()  # UI'Ä± gÃ¼ncelle
        self.taseron_table.setRowCount(len(offers))
        
        for row, offer in enumerate(offers):
            # ID (gizli)
            self.taseron_table.setItem(row, 0, QTableWidgetItem(str(offer['id'])))
            # Firma
            self.taseron_table.setItem(row, 1, QTableWidgetItem(offer['firma_adi']))
            # Kalem/TanÄ±m
            tanim = offer.get('tanim', '')
            if not tanim:
                tanim = f"Poz: {offer.get('poz_no', 'N/A')}"
            self.taseron_table.setItem(row, 2, QTableWidgetItem(tanim))
            # Miktar
            miktar = offer.get('miktar', 0)
            miktar_item = QTableWidgetItem(f"{miktar:.2f}" if miktar > 0 else "-")
            miktar_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.taseron_table.setItem(row, 3, miktar_item)
            # Birim
            self.taseron_table.setItem(row, 4, QTableWidgetItem(offer.get('birim', '')))
            # Fiyat
            fiyat_item = QTableWidgetItem(f"{offer['fiyat']:.2f} â‚º")
            fiyat_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.taseron_table.setItem(row, 5, fiyat_item)
            # Toplam
            toplam_item = QTableWidgetItem(f"{offer.get('toplam', 0):,.2f} â‚º")
            toplam_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.taseron_table.setItem(row, 6, toplam_item)
            
            # Her 50 satÄ±rda bir UI'Ä± gÃ¼ncelle
            if row % 50 == 0:
                QApplication.processEvents()
        
        # Toplam hesaplama (KDV ile)
        total = sum(offer.get('toplam', 0) for offer in offers)
        kdv_rate_text = self.taseron_kdv_rate.currentText().replace("%", "")
        kdv_rate = float(kdv_rate_text)
        kdv_hesap = self.calculator.calculate_with_kdv(total, kdv_rate)
        
        self.taseron_total_label.setText(f"Toplam (KDV HariÃ§): {total:,.2f} â‚º")
        self.taseron_total_kdv_label.setText(f"Toplam (KDV %{kdv_rate_text} Dahil): {kdv_hesap['kdv_dahil']:,.2f} â‚º")
            
    def add_taseron_offer(self) -> None:
        """TaÅŸeron teklifi ekle"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
            
        # Dialog penceresini aÃ§
        dialog = TaseronOfferDialog(self.db, self, proje_id=self.current_project_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_data()
            
            # VeritabanÄ±na ekle
            try:
                offer_id = self.db.add_taseron_teklif(
                    proje_id=self.current_project_id,
                    firma_adi=data['firma_adi'],
                    kalem_id=data['kalem_id'],
                    fiyat=data['fiyat'],
                    poz_no=data['poz_no'] if data['poz_no'] else '',
                    tanim=data['tanim'],
                    miktar=data['miktar'],
                    birim=data['birim']
                )
                
                if offer_id:
                    # Durum ve notlarÄ± gÃ¼ncelle
                    self.db.update_taseron_teklif(offer_id, durum=data['durum'], notlar=data['notlar'])
                    
                    self.load_taseron_data()
                    self.update_proje_ozet()  # Ã–zeti gÃ¼ncelle
                    self.statusBar().showMessage("Teklif baÅŸarÄ±yla eklendi")
                else:
                    QMessageBox.warning(self, "UyarÄ±", "Teklif eklenirken bir hata oluÅŸtu")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Teklif eklenirken hata oluÅŸtu:\n{str(e)}")
        
    def edit_taseron_offer(self) -> None:
        """TaÅŸeron teklifi dÃ¼zenle"""
        current_row = self.taseron_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen dÃ¼zenlemek iÃ§in bir satÄ±r seÃ§in")
            return
            
        # SeÃ§ili teklifin ID'sini al
        offer_id = int(self.taseron_table.item(current_row, 0).text())
        
        # Teklif verilerini getir
        try:
            offers = self.db.get_taseron_teklifleri(self.current_project_id)
            offer_data = next((offer for offer in offers if offer['id'] == offer_id), None)
            
            if not offer_data:
                QMessageBox.warning(self, "UyarÄ±", "Teklif bulunamadÄ±")
                return
                
            # Dialog penceresini aÃ§
            dialog = TaseronOfferDialog(self.db, self, offer_data, proje_id=self.current_project_id)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                data = dialog.get_data()
                
                # VeritabanÄ±nÄ± gÃ¼ncelle
                if self.db.update_taseron_teklif(
                    offer_id=offer_id,
                    firma_adi=data['firma_adi'],
                    kalem_id=data['kalem_id'],
                    fiyat=data['fiyat'],
                    poz_no=data['poz_no'] if data['poz_no'] else '',
                    tanim=data['tanim'],
                    miktar=data['miktar'],
                    birim=data['birim'],
                    durum=data['durum'],
                    notlar=data['notlar']
                ):
                    self.load_taseron_data()
                    self.update_proje_ozet()  # Ã–zeti gÃ¼ncelle
                    self.statusBar().showMessage("Teklif baÅŸarÄ±yla gÃ¼ncellendi")
                else:
                    QMessageBox.warning(self, "UyarÄ±", "Teklif gÃ¼ncellenirken bir hata oluÅŸtu")
                    
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Teklif dÃ¼zenlenirken hata oluÅŸtu:\n{str(e)}")
            
    def delete_taseron_offer(self) -> None:
        """TaÅŸeron teklifi sil"""
        current_row = self.taseron_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen silmek iÃ§in bir satÄ±r seÃ§in")
            return
            
        reply = QMessageBox.question(
            self, "Onay", "Bu teklifi silmek istediÄŸinize emin misiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            offer_id = int(self.taseron_table.item(current_row, 0).text())
            if self.db.delete_taseron_teklif(offer_id):
                self.load_taseron_data()
                self.update_proje_ozet()  # Ã–zeti gÃ¼ncelle
                self.statusBar().showMessage("Teklif silindi")
            else:
                QMessageBox.warning(self, "UyarÄ±", "Teklif silinirken bir hata oluÅŸtu")
    
    def compare_offers(self) -> None:
        """Teklifleri karÅŸÄ±laÅŸtÄ±r"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
            
        offers = self.db.get_taseron_teklifleri(self.current_project_id)
        if not offers:
            QMessageBox.information(self, "Bilgi", "KarÅŸÄ±laÅŸtÄ±rÄ±lacak teklif yok")
            self.comparison_table.setRowCount(0)
            self.comparison_summary_label.setText("")
            return
            
        comparison = self.calculator.compare_taseron_offers(offers)
        
        # Firma bazÄ±nda toplamlarÄ± hesapla
        firma_totals = {}
        for offer in offers:
            firma = offer['firma_adi']
            toplam = offer.get('toplam', 0)
            durum = offer.get('durum', 'beklemede')
            
            if firma not in firma_totals:
                firma_totals[firma] = {
                    'toplam': 0.0,
                    'durum': durum,
                    'teklif_sayisi': 0
                }
            
            firma_totals[firma]['toplam'] += toplam
            firma_totals[firma]['teklif_sayisi'] += 1
        
        # KarÅŸÄ±laÅŸtÄ±rma tablosunu doldur
        self.comparison_table.setRowCount(len(firma_totals))
        
        ortalama = comparison.get('ortalama', 0.0)
        row = 0
        for firma, data in sorted(firma_totals.items(), key=lambda x: x[1]['toplam']):
            # Firma
            self.comparison_table.setItem(row, 0, QTableWidgetItem(firma))
            
            # Toplam Tutar
            toplam_item = QTableWidgetItem(f"{data['toplam']:,.2f} â‚º")
            toplam_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.comparison_table.setItem(row, 1, toplam_item)
            
            # Durum
            durum_item = QTableWidgetItem(data['durum'].title())
            durum_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.comparison_table.setItem(row, 2, durum_item)
            
            # Fark (Ortalamadan)
            fark = data['toplam'] - ortalama
            fark_text = f"{fark:+,.2f} â‚º"
            fark_item = QTableWidgetItem(fark_text)
            fark_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            
            # Fark pozitifse yeÅŸil, negatifse kÄ±rmÄ±zÄ±
            if fark < 0:
                fark_item.setForeground(Qt.GlobalColor.darkGreen)
            elif fark > 0:
                fark_item.setForeground(Qt.GlobalColor.red)
            
            self.comparison_table.setItem(row, 3, fark_item)
            
            row += 1
        
        # Ã–zet bilgi
        summary = f"ğŸ“Š Toplam {len(firma_totals)} firma, {len(offers)} teklif | "
        if comparison['en_dusuk']:
            summary += f"En DÃ¼ÅŸÃ¼k: {comparison['en_dusuk']['firma']} ({comparison['en_dusuk']['tutar']:,.2f} â‚º) | "
        summary += f"Ortalama: {ortalama:,.2f} â‚º"
        
        self.comparison_summary_label.setText(summary)
    
    def export_taseron_excel(self) -> None:
        """TaÅŸeron tekliflerini Excel'e export et"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
            
        offers = self.db.get_taseron_teklifleri(self.current_project_id)
        if not offers:
            QMessageBox.warning(self, "UyarÄ±", "Export edilecek teklif bulunamadÄ±")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Excel'e Kaydet", "", "Excel DosyalarÄ± (*.xlsx)"
        )
        
        if file_path:
            proje = self.db.get_project(self.current_project_id) if self.current_project_id else None
            proje_adi = proje.get('ad', '') if proje else ''
            
            if self.export_manager.export_taseron_offers_to_excel(offers, Path(file_path), proje_adi):
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"TaÅŸeron teklifleri Excel'e aktarÄ±ldÄ±:\n{file_path}")
                self.statusBar().showMessage(f"Excel export tamamlandÄ±: {file_path}")
            else:
                QMessageBox.critical(self, "Hata", "Excel export sÄ±rasÄ±nda bir hata oluÅŸtu.")
    
    def export_taseron_pdf(self) -> None:
        """TaÅŸeron tekliflerini PDF'e export et"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
            
        offers = self.db.get_taseron_teklifleri(self.current_project_id)
        if not offers:
            QMessageBox.warning(self, "UyarÄ±", "Export edilecek teklif bulunamadÄ±")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "PDF'e Kaydet", "", "PDF DosyalarÄ± (*.pdf)"
        )
        
        if file_path:
            proje = self.db.get_project(self.current_project_id) if self.current_project_id else None
            proje_adi = proje.get('ad', '') if proje else ''
            
            if self.export_manager.export_taseron_offers_to_pdf(offers, Path(file_path), proje_adi):
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"TaÅŸeron teklifleri PDF'e aktarÄ±ldÄ±:\n{file_path}")
                self.statusBar().showMessage(f"PDF export tamamlandÄ±: {file_path}")
            else:
                QMessageBox.critical(self, "Hata", "PDF export sÄ±rasÄ±nda bir hata oluÅŸtu.")
        
    def check_and_load_pozlar_async(self) -> None:
        """Uygulama aÃ§Ä±ldÄ±ÄŸÄ±nda pozlarÄ± kontrol et ve gerekirse yÃ¼kle (async)"""
        # Arka planda yÃ¼kleme iÃ§in thread oluÅŸtur
        self.data_loader_thread = DataLoaderThread(self.db)
        self.data_loader_thread.data_loaded.connect(self.on_data_loaded)
        self.data_loader_thread.poz_question_needed.connect(self.show_poz_question)
        self.data_loader_thread.start()
        
        # Durum Ã§ubuÄŸunda bilgi gÃ¶ster
        self.statusBar().showMessage("Veriler kontrol ediliyor...")
    
    @pyqtSlot(dict)
    def on_data_loaded(self, result: Dict[str, Any]) -> None:
        """Veri yÃ¼kleme tamamlandÄ±ÄŸÄ±nda Ã§aÄŸrÄ±lÄ±r"""
        if result.get('malzemeler_loaded', False) or result.get('formuller_loaded', False):
            self.statusBar().showMessage(
                f"Veriler hazÄ±r: {result.get('malzeme_count', 0)} malzeme, "
                f"{result.get('formul_count', 0)} formÃ¼l"
            )
        else:
            self.statusBar().showMessage("HazÄ±r")
    
    @pyqtSlot()
    def show_poz_question(self) -> None:
        """Poz yÃ¼kleme sorusu gÃ¶ster"""
        reply = QMessageBox.question(
            self, "Veri YÃ¼kleme",
            "Pozlar henÃ¼z yÃ¼klenmemiÅŸ. Åimdi yÃ¼klemek ister misiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes
        )
        if reply == QMessageBox.StandardButton.Yes:
            # PozlarÄ± yÃ¼kle (yine async olabilir ama ÅŸimdilik sync)
            self.load_pozlar(silent=False)
    
    def check_and_load_pozlar(self) -> None:
        """Uygulama aÃ§Ä±ldÄ±ÄŸÄ±nda pozlarÄ± kontrol et ve gerekirse yÃ¼kle (sync versiyon - eski)"""
        if not check_pozlar_loaded(self.db):
            reply = QMessageBox.question(
                self, "Veri YÃ¼kleme",
                "Pozlar henÃ¼z yÃ¼klenmemiÅŸ. Åimdi yÃ¼klemek ister misiniz?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.load_pozlar(silent=False)
        
        # Malzeme ve formÃ¼lleri kontrol et ve yÃ¼kle
        if not check_malzemeler_loaded(self.db) or not check_formuller_loaded(self.db):
            # Sessizce yÃ¼kle (kullanÄ±cÄ±ya sorma)
            try:
                result = initialize_material_data(self.db, force_reload=False)
                if result['malzemeler']['success'] > 0 or result['formuller']['success'] > 0:
                    self.statusBar().showMessage(
                        f"Malzeme verileri yÃ¼klendi: "
                        f"{result['malzemeler']['success']} malzeme, "
                        f"{result['formuller']['success']} formÃ¼l"
                    )
            except Exception as e:
                print(f"Malzeme yÃ¼kleme hatasÄ±: {e}")
                
    def load_pozlar(self, silent: bool = False) -> None:
        """PozlarÄ± veritabanÄ±na yÃ¼kle (async - UI'Ä± bloklamaz)"""
        try:
            # Mevcut pozlar var mÄ± kontrol et
            if check_pozlar_loaded(self.db) and not silent:
                reply = QMessageBox.question(
                    self, "Onay",
                    "Pozlar zaten yÃ¼klÃ¼. Yeniden yÃ¼klemek istiyor musunuz?\n"
                    "(Mevcut pozlar silinmeyecek, sadece yeni olanlar eklenecek)",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply != QMessageBox.StandardButton.Yes:
                    return
            
            # Progress dialog gÃ¶ster
            from PyQt6.QtWidgets import QProgressDialog
            progress = QProgressDialog("Pozlar yÃ¼kleniyor...", "Ä°ptal", 0, 0, self)
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setCancelButton(None)  # Ä°ptal butonunu kaldÄ±r (uzun sÃ¼rmeyecek)
            progress.show()
            QApplication.processEvents()  # UI'Ä± gÃ¼ncelle
            
            # PozlarÄ± yÃ¼kle (kÄ±sa sÃ¼reli iÅŸlem, ama yine de progress gÃ¶ster)
            result = initialize_database_data(self.db, force_reload=False)
            
            progress.close()
            from PyQt6.QtWidgets import QApplication
            QApplication.processEvents()  # UI'Ä± gÃ¼ncelle
            
            if not silent:
                if result['pozlar']['success'] > 0:
                    QMessageBox.information(
                        self, "BaÅŸarÄ±lÄ±",
                        f"âœ… {result['pozlar']['success']} poz baÅŸarÄ±yla yÃ¼klendi!\n\n"
                        f"{result['message']}"
                    )
                    self.statusBar().showMessage(f"{result['pozlar']['success']} poz yÃ¼klendi")
                else:
                    QMessageBox.warning(
                        self, "UyarÄ±",
                        "Pozlar yÃ¼klenemedi veya zaten yÃ¼klÃ¼.\n\n"
                        f"{result['message']}"
                    )
            else:
                self.statusBar().showMessage(result['message'])
                
        except Exception as e:
            QMessageBox.critical(
                self, "Hata",
                f"Pozlar yÃ¼klenirken hata oluÅŸtu:\n{str(e)}"
            )
            
    def check_pozlar_status(self) -> None:
        """Poz durumunu kontrol et ve gÃ¶ster"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) as count FROM pozlar")
                result = cursor.fetchone()
                count = result['count'] if result else 0
                
                # Kategori bazlÄ± sayÄ±lar
                cursor.execute("""
                    SELECT kategori, COUNT(*) as count 
                    FROM pozlar 
                    GROUP BY kategori 
                    ORDER BY kategori
                """)
                categories = cursor.fetchall()
                
                message = f"ğŸ“Š Poz Durumu:\n\n"
                message += f"Toplam Poz SayÄ±sÄ±: {count}\n\n"
                message += "Kategori BazÄ±nda:\n"
                for cat in categories:
                    message += f"  â€¢ {cat['kategori']}: {cat['count']} poz\n"
                
                QMessageBox.information(self, "Poz Durumu", message)
                self.statusBar().showMessage(f"Toplam {count} poz mevcut")
                
        except Exception as e:
            QMessageBox.critical(
                self, "Hata",
                f"Poz durumu kontrol edilirken hata oluÅŸtu:\n{str(e)}"
            )
    
    def on_tab_changed(self, index: int) -> None:
        """Sekme deÄŸiÅŸtiÄŸinde Ã§aÄŸrÄ±lÄ±r (lazy loading ile)"""
        try:
            # Index 0 (Metraj Cetveli) iÃ§in bir ÅŸey yapma, zaten oluÅŸturulmuÅŸ
            if index == 0:
                return
            
            # Lazy loading: Sekmeyi ilk kez aÃ§Ä±ldÄ±ÄŸÄ±nda oluÅŸtur
            if index == 1 and not self._tabs_created['ozet']:
                try:
                    # Proje Ã–zeti sekmesi
                    placeholder = self.tabs.widget(1)
                    self.create_proje_ozet_tab(add_to_tabs=False)
                    # Signal'Ä± geÃ§ici olarak blokla (sonsuz dÃ¶ngÃ¼yÃ¼ Ã¶nlemek iÃ§in)
                    self.tabs.blockSignals(True)
                    self.tabs.removeTab(1)
                    self.tabs.insertTab(1, self.ozet_widget, "ğŸ“ˆ Proje Ã–zeti")
                    self.tabs.setCurrentIndex(1)
                    self.tabs.blockSignals(False)
                    self._tabs_created['ozet'] = True
                    if placeholder:
                        placeholder.deleteLater()
                except Exception as e:
                    self.tabs.blockSignals(False)  # Hata durumunda da bloklamayÄ± kaldÄ±r
                    print(f"Proje Ã–zeti sekmesi oluÅŸturma hatasÄ±: {e}")
                    import traceback
                    traceback.print_exc()
                    raise
            elif index == 2 and not self._tabs_created['taseron']:
                try:
                    # TaÅŸeron Analizi sekmesi
                    placeholder = self.tabs.widget(2)
                    self.create_taseron_tab(add_to_tabs=False)
                    self.tabs.blockSignals(True)
                    self.tabs.removeTab(2)
                    self.tabs.insertTab(2, self.taseron_widget, "ğŸ’¼ TaÅŸeron Analizi")
                    self.tabs.setCurrentIndex(2)
                    self.tabs.blockSignals(False)
                    self._tabs_created['taseron'] = True
                    if placeholder:
                        placeholder.deleteLater()
                except Exception as e:
                    self.tabs.blockSignals(False)
                    print(f"TaÅŸeron Analizi sekmesi oluÅŸturma hatasÄ±: {e}")
                    import traceback
                    traceback.print_exc()
                    raise
            elif index == 3 and not self._tabs_created['malzeme']:
                try:
                    # Malzeme Listesi sekmesi
                    placeholder = self.tabs.widget(3)
                    self.create_malzeme_tab(add_to_tabs=False)
                    self.tabs.blockSignals(True)
                    self.tabs.removeTab(3)
                    self.tabs.insertTab(3, self.malzeme_widget, "ğŸ“¦ Malzeme Listesi")
                    self.tabs.setCurrentIndex(3)
                    self.tabs.blockSignals(False)
                    self._tabs_created['malzeme'] = True
                    if placeholder:
                        placeholder.deleteLater()
                except Exception as e:
                    self.tabs.blockSignals(False)
                    print(f"Malzeme Listesi sekmesi oluÅŸturma hatasÄ±: {e}")
                    import traceback
                    traceback.print_exc()
                    raise
            elif index == 4 and not self._tabs_created['sablonlar']:
                try:
                    # Åablonlar sekmesi
                    placeholder = self.tabs.widget(4)
                    self.create_sablonlar_tab(add_to_tabs=False)
                    self.tabs.blockSignals(True)
                    self.tabs.removeTab(4)
                    self.tabs.insertTab(4, self.sablonlar_widget, "ğŸ“‹ Åablonlar")
                    self.tabs.setCurrentIndex(4)
                    self.tabs.blockSignals(False)
                    self._tabs_created['sablonlar'] = True
                    self.load_templates()  # Ä°lk aÃ§Ä±lÄ±ÅŸta yÃ¼kle
                    if placeholder:
                        placeholder.deleteLater()
                except Exception as e:
                    self.tabs.blockSignals(False)
                    print(f"Åablonlar sekmesi oluÅŸturma hatasÄ±: {e}")
                    import traceback
                    traceback.print_exc()
                    raise
            elif index == 5 and not self._tabs_created['birim_fiyat']:
                try:
                    # Birim Fiyat YÃ¶netimi sekmesi
                    placeholder = self.tabs.widget(5)
                    self.create_birim_fiyat_tab(add_to_tabs=False)
                    self.tabs.blockSignals(True)
                    self.tabs.removeTab(5)
                    self.tabs.insertTab(5, self.birim_fiyat_widget, "ğŸ’° Birim Fiyatlar")
                    self.tabs.setCurrentIndex(5)
                    self.tabs.blockSignals(False)
                    self._tabs_created['birim_fiyat'] = True
                    self.load_birim_fiyatlar()  # Ä°lk aÃ§Ä±lÄ±ÅŸta yÃ¼kle
                    if placeholder:
                        placeholder.deleteLater()
                except Exception as e:
                    self.tabs.blockSignals(False)
                    print(f"Birim Fiyat sekmesi oluÅŸturma hatasÄ±: {e}")
                    import traceback
                    traceback.print_exc()
                    raise
            elif index == 6 and not self._tabs_created['ihale']:
                try:
                    # Ä°hale DosyasÄ± HazÄ±rlama sekmesi
                    placeholder = self.tabs.widget(6)
                    self.create_ihale_tab(add_to_tabs=False)
                    self.tabs.blockSignals(True)
                    self.tabs.removeTab(6)
                    self.tabs.insertTab(6, self.ihale_widget, "ğŸ“„ Ä°hale DosyasÄ±")
                    self.tabs.setCurrentIndex(6)
                    self.tabs.blockSignals(False)
                    self._tabs_created['ihale'] = True
                    self.load_ihaleler()  # Ä°lk aÃ§Ä±lÄ±ÅŸta yÃ¼kle
                    if placeholder:
                        placeholder.deleteLater()
                except Exception as e:
                    self.tabs.blockSignals(False)
                    print(f"Ä°hale DosyasÄ± sekmesi oluÅŸturma hatasÄ±: {e}")
                    import traceback
                    traceback.print_exc()
                    raise
            
            # Proje Ã–zeti sekmesine geÃ§ildiÄŸinde gÃ¼ncelle (sadece sekme zaten oluÅŸturulmuÅŸsa)
            if index == 1 and self._tabs_created['ozet']:
                try:
                    self.update_proje_ozet()
                except Exception as e:
                    print(f"Proje Ã¶zeti gÃ¼ncelleme hatasÄ±: {e}")
                    import traceback
                    traceback.print_exc()
        except Exception as e:
            # Hata durumunda logla ve dosyaya yaz
            error_msg = f"Sekme deÄŸiÅŸtirme hatasÄ± (index: {index}): {e}"
            print(f"\nâŒ {error_msg}")
            import traceback
            error_trace = traceback.format_exc()
            print(error_trace)
            
            # HatayÄ± dosyaya yaz
            try:
                error_log_path = Path(__file__).parent.parent.parent / "error_log.txt"
                with open(error_log_path, 'a', encoding='utf-8') as f:
                    from datetime import datetime
                    f.write(f"\n{'='*60}\n")
                    f.write(f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"{error_msg}\n")
                    f.write(f"{error_trace}\n")
                    f.write(f"{'='*60}\n")
                print(f"âœ… Hata log dosyasÄ±na yazÄ±ldÄ±: {error_log_path}")
            except Exception as log_error:
                print(f"âŒ Log yazma hatasÄ±: {log_error}")
                import traceback
                traceback.print_exc()
            
            # KullanÄ±cÄ±ya bilgi ver (ama uygulamayÄ± kapatma)
            try:
                QMessageBox.critical(
                    self, "Hata",
                    f"Sekme deÄŸiÅŸtirilirken bir hata oluÅŸtu:\n{str(e)}\n\n"
                    f"Hata detaylarÄ± 'error_log.txt' dosyasÄ±na kaydedildi.\n\n"
                    f"LÃ¼tfen programÄ± yeniden baÅŸlatÄ±n."
                )
            except Exception as msg_error:
                print(f"QMessageBox hatasÄ±: {msg_error}")
                # UygulamayÄ± kapatma, sadece logla
            
    def update_proje_ozet(self) -> None:
        """Proje Ã¶zeti sekmesini gÃ¼ncelle"""
        # Sekme henÃ¼z oluÅŸturulmamÄ±ÅŸsa (lazy loading) gÃ¼ncelleme yapma
        if not hasattr(self, 'ozet_kalem_label') or not self._tabs_created.get('ozet', False):
            return
        
        if not self.current_project_id:
            # Proje seÃ§ili deÄŸilse temizle
            self.ozet_kalem_label.setText("0")
            self.ozet_maliyet_label.setText("0.00 â‚º")
            self.ozet_kdv_label.setText("0.00 â‚º")
            self.ozet_taseron_label.setText("0")
            self.ozet_kategori_table.setRowCount(0)
            self.ozet_pahali_table.setRowCount(0)
            self.ozet_malzeme_label.setText("Malzeme listesi hesaplanmadÄ±.\n'Malzeme Listesi' sekmesinden hesaplayÄ±nÄ±z.")
            self.ozet_taseron_detay_label.setText("TaÅŸeron teklif bilgisi yok.")
            self.stats_table.setRowCount(0)
            
            # Grafikleri temizle
            if hasattr(self, 'kategori_canvas') and self.kategori_canvas:
                try:
                    self.kategori_ax.clear()
                    self.kategori_ax.text(0.5, 0.5, 'Veri yok', ha='center', va='center', 
                                         transform=self.kategori_ax.transAxes, fontsize=12)
                    self.kategori_canvas.draw()
                except:
                    pass
            
            if hasattr(self, 'pahali_canvas') and self.pahali_canvas:
                try:
                    self.pahali_ax.clear()
                    self.pahali_ax.text(0.5, 0.5, 'Veri yok', ha='center', va='center',
                                       transform=self.pahali_ax.transAxes, fontsize=12)
                    self.pahali_canvas.draw()
                except:
                    pass
            return
        
        try:
            # Proje bilgilerini al
            proje = self.db.get_project(self.current_project_id)
            metraj_items = self.db.get_project_metraj(self.current_project_id)
            taseron_offers = self.db.get_taseron_teklifleri(self.current_project_id)
            
            # Toplam kalem sayÄ±sÄ±
            toplam_kalem = len(metraj_items)
            self.ozet_kalem_label.setText(str(toplam_kalem))
            
            # Toplam maliyet
            toplam_maliyet = sum(item.get('toplam', 0) for item in metraj_items)
            self.ozet_maliyet_label.setText(f"{toplam_maliyet:,.2f} â‚º")
            
            # KDV hesaplama
            kdv_rate_text = self.ozet_kdv_rate.currentText().replace("%", "")
            kdv_rate = float(kdv_rate_text)
            kdv_hesap = self.calculator.calculate_with_kdv(toplam_maliyet, kdv_rate)
            self.ozet_kdv_label.setText(f"{kdv_hesap['kdv_dahil']:,.2f} â‚º")
            
            # TaÅŸeron teklif sayÄ±sÄ±
            toplam_taseron = len(taseron_offers)
            self.ozet_taseron_label.setText(str(toplam_taseron))
            
            # Kategori bazÄ±nda daÄŸÄ±lÄ±m
            kategori_dict = {}
            for item in metraj_items:
                kategori = item.get('kategori', 'Kategori Yok')
                if kategori not in kategori_dict:
                    kategori_dict[kategori] = {'sayi': 0, 'toplam': 0.0}
                kategori_dict[kategori]['sayi'] += 1
                kategori_dict[kategori]['toplam'] += item.get('toplam', 0)
            
            self.ozet_kategori_table.setRowCount(len(kategori_dict))
            sorted_kategoriler = sorted(kategori_dict.items(), key=lambda x: x[1]['toplam'], reverse=True)
            for row, (kategori, data) in enumerate(sorted_kategoriler):
                self.ozet_kategori_table.setItem(row, 0, QTableWidgetItem(kategori))
                self.ozet_kategori_table.setItem(row, 1, QTableWidgetItem(str(data['sayi'])))
                toplam_item = QTableWidgetItem(f"{data['toplam']:,.2f} â‚º")
                toplam_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.ozet_kategori_table.setItem(row, 2, toplam_item)
            
            # Kategori Pie Chart
            if self.kategori_canvas and kategori_dict:
                try:
                    self.kategori_ax.clear()
                    kategoriler = [k for k, _ in sorted_kategoriler]
                    toplamlar = [d['toplam'] for _, d in sorted_kategoriler]
                    
                    # Renk paleti
                    colors = ['#c9184a', '#00BFFF', '#4CAF50', '#FF9800', '#9C27B0', '#F44336', '#2196F3', '#FFC107']
                    colors = colors[:len(kategoriler)] if len(kategoriler) <= len(colors) else colors * (len(kategoriler) // len(colors) + 1)
                    
                    wedges, texts, autotexts = self.kategori_ax.pie(
                        toplamlar, labels=kategoriler, autopct='%1.1f%%',
                        colors=colors[:len(kategoriler)], startangle=90
                    )
                    
                    # YÃ¼zde metinlerini daha okunabilir yap
                    for autotext in autotexts:
                        autotext.set_color('white')
                        autotext.set_fontweight('bold')
                        autotext.set_fontsize(9)
                    
                    self.kategori_ax.set_title('Kategori BazÄ±nda Maliyet DaÄŸÄ±lÄ±mÄ±', fontsize=11, fontweight='bold')
                    self.kategori_figure.tight_layout()
                    self.kategori_canvas.draw()
                except Exception as e:
                    print(f"Pie chart Ã§izme hatasÄ±: {e}")
            
            # En pahalÄ± 5 kalem
            sorted_items = sorted(metraj_items, key=lambda x: x.get('toplam', 0), reverse=True)[:5]
            self.ozet_pahali_table.setRowCount(len(sorted_items))
            for row, item in enumerate(sorted_items):
                tanim = item.get('tanim', '')[:40] + ('...' if len(item.get('tanim', '')) > 40 else '')
                self.ozet_pahali_table.setItem(row, 0, QTableWidgetItem(tanim))
                miktar_text = f"{item.get('miktar', 0):,.2f} {item.get('birim', '')}"
                miktar_item = QTableWidgetItem(miktar_text)
                miktar_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.ozet_pahali_table.setItem(row, 1, miktar_item)
                toplam_item = QTableWidgetItem(f"{item.get('toplam', 0):,.2f} â‚º")
                toplam_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.ozet_pahali_table.setItem(row, 2, toplam_item)
            
            # En PahalÄ± Kalemler Bar Chart
            if self.pahali_canvas and sorted_items:
                try:
                    self.pahali_ax.clear()
                    kalem_isimleri = [item.get('tanim', '')[:30] + ('...' if len(item.get('tanim', '')) > 30 else '') 
                                     for item in sorted_items]
                    toplamlar = [item.get('toplam', 0) for item in sorted_items]
                    
                    bars = self.pahali_ax.barh(kalem_isimleri, toplamlar, color='#c9184a', alpha=0.8)
                    
                    # DeÄŸerleri Ã§ubuklarÄ±n Ã¼zerine yaz
                    for i, (bar, toplam) in enumerate(zip(bars, toplamlar)):
                        width = bar.get_width()
                        self.pahali_ax.text(width, bar.get_y() + bar.get_height()/2, 
                                          f'{toplam:,.0f} â‚º',
                                          ha='left', va='center', fontweight='bold', fontsize=9)
                    
                    self.pahali_ax.set_xlabel('Toplam Maliyet (â‚º)', fontsize=10)
                    self.pahali_ax.set_title('En PahalÄ± 5 Kalem', fontsize=11, fontweight='bold')
                    self.pahali_ax.grid(axis='x', alpha=0.3)
                    self.pahali_figure.tight_layout()
                    self.pahali_canvas.draw()
                except Exception as e:
                    print(f"Bar chart Ã§izme hatasÄ±: {e}")
            
            # Malzeme Ã¶zeti
            if self.current_materials:
                toplam_malzeme_cesit = len(self.current_materials)
                toplam_malzeme_miktar = sum(m.get('miktar', 0) for m in self.current_materials)
                self.ozet_malzeme_label.setText(
                    f"ğŸ“¦ Toplam {toplam_malzeme_cesit} farklÄ± malzeme tÃ¼rÃ¼\n"
                    f"ğŸ“Š Toplam malzeme miktarÄ±: {toplam_malzeme_miktar:,.2f}"
                )
            else:
                self.ozet_malzeme_label.setText(
                    "Malzeme listesi hesaplanmadÄ±.\n"
                    "'Malzeme Listesi' sekmesinden 'Malzemeleri Hesapla' butonuna tÄ±klayÄ±nÄ±z."
                )
            
            # TaÅŸeron Ã¶zeti
            if taseron_offers:
                firma_dict = {}
                for offer in taseron_offers:
                    firma = offer.get('firma_adi', '')
                    toplam = offer.get('toplam', 0)
                    if firma not in firma_dict:
                        firma_dict[firma] = 0.0
                    firma_dict[firma] += toplam
                
                if firma_dict:
                    en_dusuk = min(firma_dict.items(), key=lambda x: x[1])
                    en_yuksek = max(firma_dict.items(), key=lambda x: x[1])
                    ortalama = sum(firma_dict.values()) / len(firma_dict)
                    
                    self.ozet_taseron_detay_label.setText(
                        f"ğŸ“Š Toplam {len(firma_dict)} firma\n"
                        f"ğŸ’° En DÃ¼ÅŸÃ¼k: {en_dusuk[0]} ({en_dusuk[1]:,.2f} â‚º)\n"
                        f"ğŸ’° En YÃ¼ksek: {en_yuksek[0]} ({en_yuksek[1]:,.2f} â‚º)\n"
                        f"ğŸ“ˆ Ortalama: {ortalama:,.2f} â‚º"
                    )
                else:
                    self.ozet_taseron_detay_label.setText("TaÅŸeron teklif bilgisi yok.")
            else:
                self.ozet_taseron_detay_label.setText("TaÅŸeron teklif bilgisi yok.")
                
        except Exception as e:
            print(f"Proje Ã¶zeti gÃ¼ncelleme hatasÄ±: {e}")
            import traceback
            traceback.print_exc()
    
    def export_proje_ozet_pdf(self) -> None:
        """Proje Ã¶zetini PDF olarak export et"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "PDF Rapor OluÅŸtur", "", "PDF DosyalarÄ± (*.pdf)"
        )
        
        if file_path:
            try:
                proje = self.db.get_project(self.current_project_id)
                metraj_items = self.db.get_project_metraj(self.current_project_id)
                taseron_offers = self.db.get_taseron_teklifleri(self.current_project_id)
                
                # KDV hesaplama
                kdv_rate_text = self.ozet_kdv_rate.currentText().replace("%", "")
                kdv_rate = float(kdv_rate_text)
                toplam_maliyet = sum(item.get('toplam', 0) for item in metraj_items)
                kdv_hesap = self.calculator.calculate_with_kdv(toplam_maliyet, kdv_rate)
                
                # PDF oluÅŸtur
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import cm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
                from datetime import datetime
                
                # Logo yolu kontrolÃ¼
                logo_path = Path(__file__).parent.parent.parent / "assets" / "logo.png"
                has_logo = logo_path.exists()
                
                doc = SimpleDocTemplate(str(file_path), pagesize=A4)
                story = []
                styles = getSampleStyleSheet()
                
                # Logo ekle
                if has_logo:
                    try:
                        logo = Image(str(logo_path), width=2*inch, height=0.8*inch)
                        logo.hAlign = 'CENTER'
                        story.append(logo)
                        story.append(Spacer(1, 0.2*inch))
                    except Exception as e:
                        print(f"Logo yÃ¼kleme hatasÄ±: {e}")
                
                # BaÅŸlÄ±k
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=18,
                    textColor=colors.HexColor('#1a1a2e'),
                    spaceAfter=30,
                    alignment=1
                )
                story.append(Paragraph(f"Proje Ã–zet Raporu - {proje.get('ad', '')}", title_style))
                story.append(Spacer(1, 0.5*cm))
                
                # Ã–zet bilgiler
                info_data = [
                    ['Proje AdÄ±', proje.get('ad', '')],
                    ['OluÅŸturulma Tarihi', proje.get('olusturma_tarihi', '')[:10] if proje.get('olusturma_tarihi') else ''],
                    ['Toplam Kalem SayÄ±sÄ±', str(len(metraj_items))],
                    ['Toplam Maliyet (KDV HariÃ§)', f"{toplam_maliyet:,.2f} TL"],
                    ['KDV (%' + kdv_rate_text + ')', f"{kdv_hesap['kdv']:,.2f} TL"],
                    ['Toplam Maliyet (KDV Dahil)', f"{kdv_hesap['kdv_dahil']:,.2f} TL"],
                    ['TaÅŸeron Teklif SayÄ±sÄ±', str(len(taseron_offers))],
                ]
                
                info_table = Table(info_data, colWidths=[6*cm, 6*cm])
                info_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#16213e')),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
                    ('BACKGROUND', (1, 0), (1, -1), colors.white),
                    ('TEXTCOLOR', (1, 0), (1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ]))
                story.append(info_table)
                story.append(Spacer(1, 0.5*cm))
                
                # Kategori daÄŸÄ±lÄ±mÄ±
                kategori_dict = {}
                for item in metraj_items:
                    kategori = item.get('kategori', 'Kategori Yok')
                    if kategori not in kategori_dict:
                        kategori_dict[kategori] = {'sayi': 0, 'toplam': 0.0}
                    kategori_dict[kategori]['sayi'] += 1
                    kategori_dict[kategori]['toplam'] += item.get('toplam', 0)
                
                if kategori_dict:
                    story.append(Paragraph("Kategori BazÄ±nda DaÄŸÄ±lÄ±m", styles['Heading2']))
                    kategori_data = [['Kategori', 'Kalem SayÄ±sÄ±', 'Toplam Maliyet']]
                    for kategori, data in sorted(kategori_dict.items(), key=lambda x: x[1]['toplam'], reverse=True):
                        kategori_data.append([
                            kategori,
                            str(data['sayi']),
                            f"{data['toplam']:,.2f} TL"
                        ])
                    
                    kategori_table = Table(kategori_data, colWidths=[6*cm, 3*cm, 3*cm])
                    kategori_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16213e')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
                    ]))
                    story.append(kategori_table)
                
                # PDF oluÅŸtur
                doc.build(story)
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"Proje Ã¶zet raporu PDF'e aktarÄ±ldÄ±:\n{file_path}")
                self.statusBar().showMessage(f"PDF rapor oluÅŸturuldu: {file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"PDF oluÅŸturulurken hata oluÅŸtu:\n{str(e)}")
                print(f"PDF export hatasÄ±: {e}")
    
    def export_proje_ozet_excel(self) -> None:
        """Proje Ã¶zetini Excel olarak export et"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Excel Rapor OluÅŸtur", "", "Excel DosyalarÄ± (*.xlsx)"
        )
        
        if file_path:
            try:
                import pandas as pd
                from openpyxl import load_workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                
                proje = self.db.get_project(self.current_project_id)
                metraj_items = self.db.get_project_metraj(self.current_project_id)
                taseron_offers = self.db.get_taseron_teklifleri(self.current_project_id)
                
                # KDV hesaplama
                kdv_rate_text = self.ozet_kdv_rate.currentText().replace("%", "")
                kdv_rate = float(kdv_rate_text)
                toplam_maliyet = sum(item.get('toplam', 0) for item in metraj_items)
                kdv_hesap = self.calculator.calculate_with_kdv(toplam_maliyet, kdv_rate)
                
                # Excel writer
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # Ã–zet bilgiler
                    ozet_data = {
                        'Bilgi': [
                            'Proje AdÄ±',
                            'OluÅŸturulma Tarihi',
                            'Toplam Kalem SayÄ±sÄ±',
                            'Toplam Maliyet (KDV HariÃ§)',
                            f'KDV (%{kdv_rate_text})',
                            'Toplam Maliyet (KDV Dahil)',
                            'TaÅŸeron Teklif SayÄ±sÄ±'
                        ],
                        'DeÄŸer': [
                            proje.get('ad', ''),
                            proje.get('olusturma_tarihi', '')[:10] if proje.get('olusturma_tarihi') else '',
                            str(len(metraj_items)),
                            f"{toplam_maliyet:,.2f} TL",
                            f"{kdv_hesap['kdv']:,.2f} TL",
                            f"{kdv_hesap['kdv_dahil']:,.2f} TL",
                            str(len(taseron_offers))
                        ]
                    }
                    df_ozet = pd.DataFrame(ozet_data)
                    df_ozet.to_excel(writer, sheet_name='Proje Ã–zeti', index=False)
                    
                    # Kategori daÄŸÄ±lÄ±mÄ±
                    kategori_dict = {}
                    for item in metraj_items:
                        kategori = item.get('kategori', 'Kategori Yok')
                        if kategori not in kategori_dict:
                            kategori_dict[kategori] = {'sayi': 0, 'toplam': 0.0}
                        kategori_dict[kategori]['sayi'] += 1
                        kategori_dict[kategori]['toplam'] += item.get('toplam', 0)
                    
                    if kategori_dict:
                        kategori_data = {
                            'Kategori': [],
                            'Kalem SayÄ±sÄ±': [],
                            'Toplam Maliyet': []
                        }
                        for kategori, data in sorted(kategori_dict.items(), key=lambda x: x[1]['toplam'], reverse=True):
                            kategori_data['Kategori'].append(kategori)
                            kategori_data['Kalem SayÄ±sÄ±'].append(data['sayi'])
                            kategori_data['Toplam Maliyet'].append(f"{data['toplam']:,.2f} TL")
                        
                        df_kategori = pd.DataFrame(kategori_data)
                        df_kategori.to_excel(writer, sheet_name='Kategori DaÄŸÄ±lÄ±mÄ±', index=False)
                    
                    # En pahalÄ± kalemler
                    sorted_items = sorted(metraj_items, key=lambda x: x.get('toplam', 0), reverse=True)[:10]
                    pahali_data = {
                        'Kalem': [item.get('tanim', '') for item in sorted_items],
                        'Miktar': [f"{item.get('miktar', 0):,.2f} {item.get('birim', '')}" for item in sorted_items],
                        'Toplam': [f"{item.get('toplam', 0):,.2f} TL" for item in sorted_items]
                    }
                    df_pahali = pd.DataFrame(pahali_data)
                    df_pahali.to_excel(writer, sheet_name='En PahalÄ± Kalemler', index=False)
                
                # Stil ayarlarÄ±
                wb = load_workbook(file_path)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # BaÅŸlÄ±k satÄ±rÄ±nÄ± kalÄ±n yap
                    header_fill = PatternFill(start_color='16213e', end_color='16213e', fill_type='solid')
                    for cell in ws[1]:
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                wb.save(file_path)
                
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"Proje Ã¶zet raporu Excel'e aktarÄ±ldÄ±:\n{file_path}")
                self.statusBar().showMessage(f"Excel rapor oluÅŸturuldu: {file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Excel oluÅŸturulurken hata oluÅŸtu:\n{str(e)}")
                print(f"Excel export hatasÄ±: {e}")
    
    def backup_current_project(self) -> None:
        """SeÃ§ili projeyi yedekle"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
        
        # Proje adÄ±nÄ± al
        project = self.db.get_project(self.current_project_id)
        if not project:
            QMessageBox.warning(self, "UyarÄ±", "Proje bulunamadÄ±")
            return
        
        # Yedek dosyasÄ± seÃ§
        default_name = f"{project['ad']}_yedek_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Projeyi Yedekle", default_name, "JSON DosyalarÄ± (*.json)"
        )
        
        if file_path:
            if self.db.backup_project(self.current_project_id, Path(file_path)):
                QMessageBox.information(
                    self, "BaÅŸarÄ±lÄ±",
                    f"Proje baÅŸarÄ±yla yedeklendi:\n{file_path}"
                )
                self.statusBar().showMessage(f"Proje yedeklendi: {file_path}")
            else:
                QMessageBox.critical(
                    self, "Hata",
                    "Yedekleme sÄ±rasÄ±nda bir hata oluÅŸtu."
                )
    
    def backup_all_projects(self) -> None:
        """TÃ¼m projeleri yedekle"""
        # Yedek dosyasÄ± seÃ§
        default_name = f"tum_projeler_yedek_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "TÃ¼m Projeleri Yedekle", default_name, "JSON DosyalarÄ± (*.json)"
        )
        
        if file_path:
            if self.db.backup_all_projects(Path(file_path)):
                QMessageBox.information(
                    self, "BaÅŸarÄ±lÄ±",
                    f"TÃ¼m projeler baÅŸarÄ±yla yedeklendi:\n{file_path}"
                )
                self.statusBar().showMessage(f"TÃ¼m projeler yedeklendi: {file_path}")
            else:
                QMessageBox.critical(
                    self, "Hata",
                    "Yedekleme sÄ±rasÄ±nda bir hata oluÅŸtu."
                )
    
    def restore_project(self) -> None:
        """Yedekten proje geri yÃ¼kle"""
        # Yedek dosyasÄ± seÃ§
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Yedekten Geri YÃ¼kle", "", "JSON DosyalarÄ± (*.json)"
        )
        
        if file_path:
            # Proje adÄ± sor
            from PyQt6.QtWidgets import QInputDialog
            project_name, ok = QInputDialog.getText(
                self, "Proje AdÄ±",
                "Yeni proje adÄ± (boÅŸ bÄ±rakÄ±rsanÄ±z yedekteki ad kullanÄ±lÄ±r):"
            )
            
            if ok:
                new_name = project_name.strip() if project_name.strip() else None
                project_id = self.db.restore_project(Path(file_path), new_name)
                
                if project_id:
                    QMessageBox.information(
                        self, "BaÅŸarÄ±lÄ±",
                        f"Proje baÅŸarÄ±yla geri yÃ¼klendi!"
                    )
                    # Proje listesini yenile
                    self.load_projects()
                    # Yeni projeyi seÃ§
                    self.current_project_id = project_id
                    self.load_metraj_data()
                    self.load_taseron_data()
                    self.update_proje_ozet()
                    self.statusBar().showMessage("Proje geri yÃ¼klendi")
                else:
                    QMessageBox.critical(
                        self, "Hata",
                        "Geri yÃ¼kleme sÄ±rasÄ±nda bir hata oluÅŸtu."
                    )
    
    def create_project_version(self) -> None:
        """Proje versiyonu oluÅŸtur"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
        
        from PyQt6.QtWidgets import QInputDialog
        
        # Versiyon adÄ± al
        version_name, ok = QInputDialog.getText(
            self, "Versiyon OluÅŸtur",
            "Versiyon adÄ±:",
            text=f"Versiyon {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        
        if not ok or not version_name.strip():
            return
        
        # AÃ§Ä±klama al
        description, ok = QInputDialog.getText(
            self, "Versiyon AÃ§Ä±klamasÄ±",
            "Versiyon aÃ§Ä±klamasÄ± (isteÄŸe baÄŸlÄ±):"
        )
        
        if not ok:
            return
        
        # Versiyon oluÅŸtur
        try:
            version_id = self.db.create_project_version(
                project_id=self.current_project_id,
                version_name=version_name.strip(),
                description=description.strip(),
                created_by="KullanÄ±cÄ±"
            )
            
            QMessageBox.information(
                self, "BaÅŸarÄ±lÄ±",
                f"Versiyon baÅŸarÄ±yla oluÅŸturuldu!\nVersiyon ID: {version_id}"
            )
            self.statusBar().showMessage(f"Versiyon oluÅŸturuldu: {version_name}")
        except Exception as e:
            QMessageBox.critical(
                self, "Hata",
                f"Versiyon oluÅŸturulurken hata oluÅŸtu:\n{str(e)}"
            )
    
    def view_project_versions(self) -> None:
        """Proje versiyonlarÄ±nÄ± gÃ¶rÃ¼ntÃ¼le"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
        
        versions = self.db.get_project_versions(self.current_project_id)
        
        if not versions:
            QMessageBox.information(
                self, "Bilgi",
                "Bu proje iÃ§in henÃ¼z versiyon oluÅŸturulmamÄ±ÅŸ."
            )
            return
        
        # Versiyon listesi dialogu
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, QPushButton, QHeaderView
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Proje VersiyonlarÄ±")
        dialog.setGeometry(200, 200, 800, 500)
        
        layout = QVBoxLayout(dialog)
        
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels([
            "Versiyon No", "Versiyon AdÄ±", "OluÅŸturulma Tarihi", "AÃ§Ä±klama", "OluÅŸturan"
        ])
        table.setRowCount(len(versions))
        table.horizontalHeader().setStretchLastSection(True)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        for row, version in enumerate(versions):
            table.setItem(row, 0, QTableWidgetItem(str(version.get('version_number', ''))))
            table.setItem(row, 1, QTableWidgetItem(version.get('version_name', '')))
            table.setItem(row, 2, QTableWidgetItem(version.get('created_at', '')))
            table.setItem(row, 3, QTableWidgetItem(version.get('description', '')))
            table.setItem(row, 4, QTableWidgetItem(version.get('created_by', '')))
        
        layout.addWidget(table)
        
        btn_layout = QHBoxLayout()
        btn_restore = QPushButton("SeÃ§ili Versiyondan Geri YÃ¼kle")
        btn_restore.clicked.connect(lambda: self.restore_selected_version(dialog, table))
        btn_layout.addWidget(btn_restore)
        
        btn_close = QPushButton("Kapat")
        btn_close.clicked.connect(dialog.close)
        btn_layout.addWidget(btn_close)
        
        layout.addLayout(btn_layout)
        
        dialog.exec()
    
    def restore_selected_version(self, dialog: QDialog, table: QTableWidget) -> None:
        """SeÃ§ili versiyondan geri yÃ¼kle"""
        current_row = table.currentRow()
        if current_row < 0:
            QMessageBox.warning(dialog, "UyarÄ±", "LÃ¼tfen bir versiyon seÃ§in")
            return
        
        version_id_item = table.item(current_row, 0)
        if not version_id_item:
            return
        
        # Versiyon ID'yi bul
        version_number = int(version_id_item.text())
        versions = self.db.get_project_versions(self.current_project_id)
        selected_version = next((v for v in versions if v['version_number'] == version_number), None)
        
        if not selected_version:
            QMessageBox.warning(dialog, "UyarÄ±", "Versiyon bulunamadÄ±")
            return
        
        from PyQt6.QtWidgets import QInputDialog
        
        # Yeni proje adÄ± al
        project_name, ok = QInputDialog.getText(
            dialog, "Yeni Proje AdÄ±",
            "Yeni proje adÄ± (boÅŸ bÄ±rakÄ±rsanÄ±z versiyon adÄ± kullanÄ±lÄ±r):",
            text=f"{selected_version['version_name']} (Geri YÃ¼klenen)"
        )
        
        if not ok:
            return
        
        new_name = project_name.strip() if project_name.strip() else None
        
        # Geri yÃ¼kle
        try:
            new_project_id = self.db.restore_project_version(selected_version['id'], new_name)
            
            if new_project_id:
                QMessageBox.information(
                    dialog, "BaÅŸarÄ±lÄ±",
                    f"Versiyon baÅŸarÄ±yla geri yÃ¼klendi!\nYeni proje ID: {new_project_id}"
                )
                dialog.close()
                self.load_projects()
                self.statusBar().showMessage("Versiyon geri yÃ¼klendi")
            else:
                QMessageBox.critical(
                    dialog, "Hata",
                    "Geri yÃ¼kleme sÄ±rasÄ±nda bir hata oluÅŸtu."
                )
        except Exception as e:
            QMessageBox.critical(
                dialog, "Hata",
                f"Geri yÃ¼kleme sÄ±rasÄ±nda hata oluÅŸtu:\n{str(e)}"
            )
    
    def restore_from_version(self) -> None:
        """Versiyondan geri yÃ¼kle (kÄ±sayol)"""
        self.view_project_versions()
    
    def load_project_notes(self) -> None:
        """Proje notlarÄ±nÄ± yÃ¼kle"""
        if not self.current_project_id:
            self.project_notes_text.clear()
            self.project_notes_text.setEnabled(False)
            return
        
        project = self.db.get_project(self.current_project_id)
        if project:
            notes = project.get('notlar', '') or ''
            self.project_notes_text.setPlainText(notes)
            self.project_notes_text.setEnabled(True)
        else:
            self.project_notes_text.clear()
            self.project_notes_text.setEnabled(False)
    
    def save_project_notes(self) -> None:
        """Proje notlarÄ±nÄ± kaydet"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
        
        notes = self.project_notes_text.toPlainText()
        if self.db.update_project(self.current_project_id, notlar=notes):
            QMessageBox.information(self, "BaÅŸarÄ±lÄ±", "Notlar kaydedildi")
            self.statusBar().showMessage("Proje notlarÄ± kaydedildi")
        else:
            QMessageBox.critical(self, "Hata", "Notlar kaydedilirken bir hata oluÅŸtu")
    
    def import_from_excel(self) -> None:
        """Excel dosyasÄ±ndan kalem iÃ§e aktar"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
        
        # Excel dosyasÄ± seÃ§
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Excel DosyasÄ± SeÃ§", "", "Excel DosyalarÄ± (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
        
        try:
            import pandas as pd
            
            # Excel dosyasÄ±nÄ± oku
            df = pd.read_excel(file_path)
            
            # SÃ¼tun adlarÄ±nÄ± normalize et (boÅŸluklarÄ± temizle)
            original_columns = df.columns.tolist()
            df.columns = [str(col).strip() for col in df.columns]
            
            # SÃ¼tun adlarÄ±nÄ± eÅŸleÅŸtir (TÃ¼rkÃ§e ve Ä°ngilizce desteÄŸi)
            # Hem tam eÅŸleÅŸme hem de case-insensitive eÅŸleÅŸme
            column_mapping_dict = {
                # TÃ¼rkÃ§e -> Ä°ngilizce
                'Poz No': 'poz_no',
                'poz no': 'poz_no',
                'POZ NO': 'poz_no',
                'TanÄ±m': 'tanim',
                'tanÄ±m': 'tanim',
                'TANIM': 'tanim',
                'Tanim': 'tanim',
                'Miktar': 'miktar',
                'miktar': 'miktar',
                'MIKTAR': 'miktar',
                'Birim': 'birim',
                'birim': 'birim',
                'BÄ°RÄ°M': 'birim',
                'Birim Fiyat': 'birim_fiyat',
                'birim fiyat': 'birim_fiyat',
                'BÄ°RÄ°M FÄ°YAT': 'birim_fiyat',
                'BirimFiyat': 'birim_fiyat',
                'Kategori': 'kategori',
                'kategori': 'kategori',
                'KATEGORÄ°': 'kategori',
                'Kaynak': 'kaynak',
                'kaynak': 'kaynak',
                'KAYNAK': 'kaynak',
            }
            
            # Ä°ngilizce sÃ¼tun adlarÄ± zaten doÄŸruysa ekle
            for eng_col in ['poz_no', 'tanim', 'miktar', 'birim', 'birim_fiyat', 'kategori', 'kaynak']:
                if eng_col not in column_mapping_dict:
                    column_mapping_dict[eng_col] = eng_col
            
            # SÃ¼tun adlarÄ±nÄ± normalize et
            normalized_columns = {}
            for col in df.columns:
                col_clean = str(col).strip()
                # Ã–nce tam eÅŸleÅŸme
                if col_clean in column_mapping_dict:
                    normalized_columns[col] = column_mapping_dict[col_clean]
                # Sonra case-insensitive eÅŸleÅŸme
                else:
                    col_lower = col_clean.lower()
                    found = False
                    for key, value in column_mapping_dict.items():
                        if key.lower() == col_lower:
                            normalized_columns[col] = value
                            found = True
                            break
                    if not found:
                        # EÅŸleÅŸme bulunamadÄ±, olduÄŸu gibi bÄ±rak
                        normalized_columns[col] = col_clean
            
            # SÃ¼tun adlarÄ±nÄ± deÄŸiÅŸtir
            df = df.rename(columns=normalized_columns)
            
            # Debug: SÃ¼tun adlarÄ±nÄ± kontrol et
            print(f"Original columns: {original_columns}")
            print(f"Normalized columns: {df.columns.tolist()}")
            
            # Gerekli sÃ¼tunlarÄ± kontrol et (miktar ve birim opsiyonel)
            required_columns = ['poz_no', 'tanim']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                # Mevcut sÃ¼tunlarÄ± gÃ¶ster
                available_cols = ', '.join(df.columns.tolist())
                QMessageBox.warning(
                    self, "Hata",
                    f"Excel dosyasÄ±nda gerekli sÃ¼tunlar eksik:\n{', '.join(missing_columns)}\n\n"
                    f"Gerekli sÃ¼tunlar:\n"
                    f"  - poz_no (veya 'Poz No')\n"
                    f"  - tanim (veya 'TanÄ±m')\n\n"
                    f"Opsiyonel sÃ¼tunlar:\n"
                    f"  - birim (veya 'Birim') - Yoksa varsayÄ±lan 'adet' kullanÄ±lÄ±r\n"
                    f"  - miktar (veya 'Miktar') - Yoksa varsayÄ±lan 1.0 kullanÄ±lÄ±r\n"
                    f"  - birim_fiyat (veya 'Birim Fiyat') - Yoksa 0 kullanÄ±lÄ±r\n\n"
                    f"Mevcut sÃ¼tunlar:\n{available_cols}"
                )
                return
            
            # Miktar sÃ¼tunu yoksa ekle (varsayÄ±lan 1.0)
            if 'miktar' not in df.columns:
                df['miktar'] = 1.0
                print(f"Added 'miktar' column with default value 1.0")
            
            # Birim sÃ¼tunu yoksa ekle (varsayÄ±lan 'adet')
            if 'birim' not in df.columns:
                df['birim'] = 'adet'
                print(f"Added 'birim' column with default value 'adet'")
            
            # Birim fiyat sÃ¼tunu yoksa ekle (varsayÄ±lan 0)
            if 'birim_fiyat' not in df.columns:
                df['birim_fiyat'] = 0.0
                print(f"Added 'birim_fiyat' column with default value 0.0")
            
            print(f"Final columns before processing: {df.columns.tolist()}")
            
            # Veri kontrolÃ¼: BoÅŸ satÄ±rlarÄ± temizle
            df = df.dropna(subset=['poz_no', 'tanim'], how='all')  # Her iki sÃ¼tun da boÅŸsa sil
            
            if df.empty:
                QMessageBox.warning(
                    self, "UyarÄ±",
                    "Excel dosyasÄ±nda iÅŸlenecek veri bulunamadÄ±.\n\n"
                    "LÃ¼tfen 'Poz No' ve 'TanÄ±m' sÃ¼tunlarÄ±nÄ±n dolu olduÄŸundan emin olun."
                )
                return
            
            print(f"Processing {len(df)} rows...")
            
            # Progress dialog ekle (Ã§ok satÄ±r varsa)
            from PyQt6.QtWidgets import QProgressDialog
            from PyQt6.QtCore import Qt
            if len(df) > 100:
                progress = QProgressDialog("Excel verileri iÅŸleniyor...", "Ä°ptal", 0, len(df), self)
                progress.setWindowModality(Qt.WindowModality.WindowModal)
                progress.setMinimumDuration(0)
                progress.show()
            
            # Kalemleri ekle
            success_count = 0
            error_count = 0
            errors = []
            skipped_empty = 0
            
            for idx, (index, row) in enumerate(df.iterrows()):
                try:
                    # Progress gÃ¼ncelle
                    if len(df) > 100 and idx % 100 == 0:
                        progress.setValue(idx)
                        progress.setLabelText(f"Excel verileri iÅŸleniyor... {idx}/{len(df)}")
                        from PyQt6.QtWidgets import QApplication
                        QApplication.processEvents()
                        if progress.wasCanceled():
                            break
                    
                    # SÃ¼tun deÄŸerlerini al (normalize edilmiÅŸ sÃ¼tun adlarÄ± ile)
                    poz_no_raw = row.get('poz_no', '')
                    tanim_raw = row.get('tanim', '')
                    
                    # NaN kontrolÃ¼ ve string'e Ã§evirme
                    if pd.isna(poz_no_raw):
                        poz_no = ''
                    else:
                        poz_no = str(poz_no_raw).strip()
                        if poz_no.lower() == 'nan' or poz_no == '':
                            poz_no = ''
                    
                    if pd.isna(tanim_raw):
                        tanim = ''
                    else:
                        tanim = str(tanim_raw).strip()
                        if tanim.lower() == 'nan' or tanim == '':
                            tanim = ''
                    
                    # BoÅŸ satÄ±rlarÄ± atla
                    if not poz_no and not tanim:
                        skipped_empty += 1
                        continue
                    
                    # Poz no veya tanÄ±m boÅŸsa hata
                    if not poz_no:
                        error_count += 1
                        if len(errors) < 20:  # Ä°lk 20 hatayÄ± gÃ¶ster
                            errors.append(f"SatÄ±r {index + 2}: Poz no boÅŸ (TanÄ±m: '{tanim[:50]}...' if len(tanim) > 50 else tanim)")
                        continue
                    
                    if not tanim:
                        error_count += 1
                        if len(errors) < 20:
                            errors.append(f"SatÄ±r {index + 2}: TanÄ±m boÅŸ (Poz: '{poz_no}')")
                        continue
                    
                    # Kategori (opsiyonel)
                    kategori = ''
                    if 'kategori' in df.columns:
                        kategori_raw = row.get('kategori', '')
                        if pd.notna(kategori_raw):
                            kategori_str = str(kategori_raw).strip()
                            if kategori_str.lower() != 'nan' and kategori_str:
                                kategori = kategori_str
                    
                    # Miktar - varsa kullan, yoksa 1.0
                    miktar_val = row.get('miktar', 1.0)
                    if pd.isna(miktar_val):
                        miktar = 1.0
                    else:
                        try:
                            miktar_str = str(miktar_val).strip()
                            if miktar_str.lower() == 'nan' or miktar_str == '':
                                miktar = 1.0
                            else:
                                miktar = float(miktar_val)
                                if miktar < 0:
                                    miktar = 1.0
                        except (ValueError, TypeError) as e:
                            print(f"SatÄ±r {index + 2}: Miktar dÃ¶nÃ¼ÅŸÃ¼m hatasÄ±: {miktar_val} -> 1.0")
                            miktar = 1.0
                    
                    # Birim - varsa kullan, yoksa 'adet'
                    birim_val = row.get('birim', 'adet')
                    if pd.isna(birim_val):
                        birim = 'adet'
                    else:
                        birim_str = str(birim_val).strip()
                        if birim_str.lower() == 'nan' or not birim_str:
                            birim = 'adet'
                        else:
                            birim = birim_str
                    
                    # Birim fiyat - varsa kullan, yoksa 0
                    birim_fiyat_val = row.get('birim_fiyat', 0)
                    if pd.isna(birim_fiyat_val):
                        birim_fiyat = 0.0
                    else:
                        try:
                            birim_fiyat_str = str(birim_fiyat_val).strip()
                            if birim_fiyat_str.lower() == 'nan' or birim_fiyat_str == '':
                                birim_fiyat = 0.0
                            else:
                                # VirgÃ¼lÃ¼ noktaya Ã§evir (TÃ¼rkÃ§e format)
                                birim_fiyat_str = birim_fiyat_str.replace(',', '.').replace(' ', '')
                                # Binlik ayÄ±rÄ±cÄ±larÄ± temizle (1.234,56 -> 1234.56)
                                if '.' in birim_fiyat_str and ',' in birim_fiyat_str:
                                    # TÃ¼rkÃ§e format: 1.234,56
                                    birim_fiyat_str = birim_fiyat_str.replace('.', '').replace(',', '.')
                                birim_fiyat = float(birim_fiyat_str)
                                if birim_fiyat < 0:
                                    birim_fiyat = 0.0
                                print(f"DEBUG: SatÄ±r {index + 2}: Birim fiyat okundu: {birim_fiyat_val} -> {birim_fiyat}")
                        except (ValueError, TypeError) as e:
                            print(f"SatÄ±r {index + 2}: Birim fiyat dÃ¶nÃ¼ÅŸÃ¼m hatasÄ±: {birim_fiyat_val} -> 0.0")
                            birim_fiyat = 0.0
                    
                    # Ã–nce poz'u ekle/gÃ¼ncelle (add_poz zaten varsa gÃ¼nceller)
                    poz_id = self.db.add_poz(
                        poz_no=poz_no,
                        tanim=tanim,
                        birim=birim,
                        kategori=kategori if kategori else "",
                        resmi_fiyat=birim_fiyat if birim_fiyat > 0 else 0
                    )
                    
                    # Birim fiyat ekle (eÄŸer birim fiyat > 0 ise)
                    if birim_fiyat > 0:
                        fiyat_id = self.db.add_birim_fiyat(
                            poz_id=poz_id,
                            poz_no=poz_no,
                        birim_fiyat=birim_fiyat,
                            kaynak='Excel Import'
                    )
                        print(f"DEBUG: Poz {poz_no} iÃ§in birim fiyat eklendi: {birim_fiyat} (ID: {fiyat_id})")
                    else:
                        print(f"DEBUG: Poz {poz_no} iÃ§in birim fiyat 0, eklenmedi")
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    error_msg = str(e)
                    if len(errors) < 20:  # Ä°lk 20 hatayÄ± gÃ¶ster
                        errors.append(f"SatÄ±r {index + 2}: {error_msg}")
                    print(f"SatÄ±r {index + 2} hatasÄ±: {error_msg}")
                    import traceback
                    if error_count <= 5:  # Ä°lk 5 hatanÄ±n detayÄ±nÄ± gÃ¶ster
                        traceback.print_exc()
                    continue
            
            # Progress dialog'u kapat
            if len(df) > 100:
                progress.setValue(len(df))
                progress.close()
            
            # SonuÃ§ mesajÄ±
            message = f"Ä°Ã§e aktarma tamamlandÄ±!\n\n"
            message += f"âœ… BaÅŸarÄ±lÄ±: {success_count}\n"
            message += f"âŒ HatalÄ±: {error_count}\n"
            if skipped_empty > 0:
                message += f"â­ï¸ BoÅŸ satÄ±rlar atlandÄ±: {skipped_empty}\n"
            
            if errors:
                if error_count <= 20:
                    message += f"\n\nHatalar:\n" + "\n".join(errors)
                else:
                    message += f"\n\n(Ä°lk 20 hata gÃ¶steriliyor, toplam {error_count} hata var)\n\nHatalar:\n" + "\n".join(errors[:20])
            
            if success_count > 0:
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", message)
                # Verileri yenile
                self.load_metraj_data()
                self.update_proje_ozet()
                self.statusBar().showMessage(f"{success_count} kalem iÃ§e aktarÄ±ldÄ±")
            else:
                QMessageBox.warning(
                    self, "UyarÄ±", 
                    message + "\n\nHiÃ§bir kalem eklenemedi. LÃ¼tfen Excel dosyasÄ±nÄ± kontrol edin."
                )
                
        except Exception as e:
            QMessageBox.critical(
                self, "Hata",
                f"Excel dosyasÄ± iÅŸlenirken hata oluÅŸtu:\n{str(e)}\n\n"
                f"LÃ¼tfen Excel dosyasÄ±nÄ±n formatÄ±nÄ± kontrol edin:\n"
                f"- 'Poz No' veya 'poz_no' sÃ¼tunu olmalÄ±\n"
                f"- 'TanÄ±m' veya 'tanim' sÃ¼tunu olmalÄ±\n"
                f"- DiÄŸer sÃ¼tunlar (Miktar, Birim, Birim Fiyat) opsiyoneldir"
            )
            import traceback
            traceback.print_exc()
    
    def import_from_pdf(self) -> None:
        """PDF'den birim fiyat iÃ§e aktar"""
        # PDF dosyasÄ± seÃ§
        file_path, _ = QFileDialog.getOpenFileName(
            self, "PDF DosyasÄ± SeÃ§", "", "PDF DosyalarÄ± (*.pdf)"
        )
        
        if not file_path:
            return
        
        try:
            from PyQt6.QtWidgets import QProgressDialog, QDialog, QVBoxLayout, QLabel, QTableWidget, QPushButton, QHBoxLayout
            from PyQt6.QtCore import Qt
            
            # Progress dialog
            progress = QProgressDialog("PDF iÅŸleniyor...", "Ä°ptal", 0, 100, self)
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            
            def progress_callback(current, total):
                percent = int((current / total) * 100)
                progress.setValue(percent)
                progress.setLabelText(f"PDF iÅŸleniyor... Sayfa {current}/{total}")
                if progress.wasCanceled():
                    return False
                return True
            
            # PDF'i iÅŸle
            importer = PDFBirimFiyatImporter()
            extracted_data = importer.extract_from_pdf(Path(file_path), progress_callback)
            
            progress.setValue(100)
            
            if not extracted_data:
                QMessageBox.warning(
                    self, "UyarÄ±",
                    "PDF'den poz ve fiyat bilgisi Ã§Ä±karÄ±lamadÄ±.\n\n"
                    "PDF formatÄ±nÄ± kontrol edin veya manuel olarak ekleyin."
                )
                return
            
            # Ã–nizleme ve onay dialogu
            preview_dialog = QDialog(self)
            preview_dialog.setWindowTitle("PDF Ä°Ã§e Aktarma Ã–nizleme")
            preview_dialog.setMinimumSize(900, 650)
            
            layout = QVBoxLayout(preview_dialog)
            
            info_label = QLabel(
                f"ğŸ“„ {len(extracted_data)} adet poz ve fiyat bulundu.\n\n"
                f"LÃ¼tfen Ã¶nizlemeyi kontrol edin ve onaylayÄ±n:"
            )
            layout.addWidget(info_label)
            
            preview_table = QTableWidget()
            preview_table.setColumnCount(4)
            preview_table.setHorizontalHeaderLabels(["Poz No", "TanÄ±m", "Birim Fiyat", "Kaynak"])
            preview_table.setRowCount(min(len(extracted_data), 100))  # Ä°lk 100 kayÄ±t
            
            for row, item in enumerate(extracted_data[:100]):
                preview_table.setItem(row, 0, QTableWidgetItem(item.get('poz_no', '')))
                preview_table.setItem(row, 1, QTableWidgetItem(item.get('tanim', '')[:50]))
                fiyat = item.get('birim_fiyat', 0)
                preview_table.setItem(row, 2, QTableWidgetItem(f"{fiyat:,.2f} â‚º" if fiyat else "BulunamadÄ±"))
                preview_table.setItem(row, 3, QTableWidgetItem(item.get('kaynak', '')))
            
            preview_table.horizontalHeader().setStretchLastSection(True)
            layout.addWidget(preview_table)
            
            if len(extracted_data) > 100:
                more_label = QLabel(f"... ve {len(extracted_data) - 100} kayÄ±t daha")
                layout.addWidget(more_label)
            
            btn_layout = QHBoxLayout()
            
            # Excel'e Aktar butonu
            btn_export_excel = QPushButton("ğŸ“Š Excel'e Aktar")
            btn_export_excel.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 8px;")
            
            def export_to_excel():
                """PDF verilerini Excel'e aktar"""
                excel_path, _ = QFileDialog.getSaveFileName(
                    preview_dialog, 
                    "Excel DosyasÄ± OluÅŸtur", 
                    f"PDF_Pozlar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    "Excel DosyalarÄ± (*.xlsx)"
                )
                
                if excel_path:
                    try:
                        import pandas as pd
                        from openpyxl.styles import Font, Alignment, PatternFill
                        
                        # DataFrame oluÅŸtur
                        data = []
                        for item in extracted_data:
                            data.append({
                                'Poz No': item.get('poz_no', ''),
                                'TanÄ±m': item.get('tanim', ''),
                                'Miktar': 1.0,  # VarsayÄ±lan miktar (kullanÄ±cÄ± dÃ¼zenleyebilir)
                                'Birim': '',  # KullanÄ±cÄ± dolduracak
                                'Birim Fiyat': item.get('birim_fiyat', 0) if item.get('birim_fiyat') else '',
                                'Kategori': '',  # KullanÄ±cÄ± dolduracak
                                'Kaynak': item.get('kaynak', 'PDF Import')
                            })
                        
                        df = pd.DataFrame(data)
                        
                        # Excel'e yaz
                        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                            df.to_excel(writer, sheet_name='Pozlar', index=False)
                            
                            # Stil ayarlarÄ±
                            worksheet = writer.sheets['Pozlar']
                            
                            # SÃ¼tun geniÅŸlikleri
                            worksheet.column_dimensions['A'].width = 20  # Poz No
                            worksheet.column_dimensions['B'].width = 60  # TanÄ±m
                            worksheet.column_dimensions['C'].width = 12  # Miktar
                            worksheet.column_dimensions['D'].width = 10  # Birim
                            worksheet.column_dimensions['E'].width = 15  # Birim Fiyat
                            worksheet.column_dimensions['F'].width = 20  # Kategori
                            worksheet.column_dimensions['G'].width = 15  # Kaynak
                            
                            # BaÅŸlÄ±k satÄ±rÄ±nÄ± stilize et
                            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            header_font = Font(bold=True, color="FFFFFF")
                            header_alignment = Alignment(horizontal='center', vertical='center')
                            
                            for cell in worksheet[1]:
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_alignment
                            
                            # SayÄ± formatlarÄ±
                            from openpyxl.styles import numbers
                            # Miktar sÃ¼tunu (C)
                            for row in range(2, len(df) + 2):
                                cell = worksheet[f'C{row}']
                                if cell.value:
                                    cell.number_format = '#,##0.00'
                            # Birim Fiyat sÃ¼tunu (E)
                            for row in range(2, len(df) + 2):
                                cell = worksheet[f'E{row}']
                                if cell.value:
                                    cell.number_format = '#,##0.00'
                        
                        QMessageBox.information(
                            preview_dialog,
                            "BaÅŸarÄ±lÄ±",
                            f"âœ… Excel dosyasÄ± oluÅŸturuldu!\n\n"
                            f"ğŸ“ Konum: {excel_path}\n\n"
                            f"ğŸ“ {len(extracted_data)} adet poz Excel'e aktarÄ±ldÄ±.\n\n"
                            f"ğŸ’¡ Excel'de verileri kontrol edip dÃ¼zenleyebilir,\n"
                            f"sonra 'Excel'den Kalem Ä°Ã§e Aktar' ile programa yÃ¼kleyebilirsiniz."
                        )
                        
                        # Dialog'u kapat
                        preview_dialog.accept()
                        
                    except Exception as e:
                        QMessageBox.critical(
                            preview_dialog,
                            "Hata",
                            f"Excel dosyasÄ± oluÅŸturulurken hata oluÅŸtu:\n{str(e)}"
                        )
                        import traceback
                        traceback.print_exc()
            
            btn_export_excel.clicked.connect(export_to_excel)
            
            btn_ok = QPushButton("âœ… DoÄŸrudan Ä°Ã§e Aktar")
            btn_ok.clicked.connect(preview_dialog.accept)
            btn_cancel = QPushButton("âŒ Ä°ptal")
            btn_cancel.clicked.connect(preview_dialog.reject)
            
            btn_layout.addWidget(btn_export_excel)
            btn_layout.addStretch()
            btn_layout.addWidget(btn_ok)
            btn_layout.addWidget(btn_cancel)
            layout.addLayout(btn_layout)
            
            # Bilgi mesajÄ± ekle
            info_text = QLabel(
                "ğŸ’¡ Ä°pucu: Excel'e aktarÄ±p kontrol etmek daha gÃ¼venilirdir!\n"
                "Excel'de verileri dÃ¼zenleyebilir, sonra 'Excel'den Kalem Ä°Ã§e Aktar' ile yÃ¼kleyebilirsiniz."
            )
            info_text.setStyleSheet("color: #666; font-style: italic; padding: 5px;")
            layout.insertWidget(1, info_text)
            
            if preview_dialog.exec() != QDialog.DialogCode.Accepted:
                return
            
            # VeritabanÄ±na kaydet
            success_count = 0
            error_count = 0
            poz_added_count = 0
            fiyat_added_count = 0
            errors = []
            
            progress = QProgressDialog("VeritabanÄ±na kaydediliyor...", "Ä°ptal", 0, len(extracted_data), self)
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            
            for idx, item in enumerate(extracted_data):
                progress.setValue(idx)
                progress.setLabelText(f"Kaydediliyor... {idx + 1}/{len(extracted_data)}")
                
                if progress.wasCanceled():
                    break
                
                try:
                    poz_no = item.get('poz_no', '').strip()
                    birim_fiyat = item.get('birim_fiyat', 0)
                    tanim = item.get('tanim', '') or "PDF'den iÃ§e aktarÄ±ldÄ±"
                    
                    if not poz_no:
                        error_count += 1
                        continue
                    
                    # Ã–NCE POZU POZLAR TABLOSUNA EKLE (yoksa)
                    poz = self.db.get_poz(poz_no)
                    if not poz:
                        # Poz yoksa ekle
                        try:
                            # Birim bilgisini tahmin et (tanÄ±mdan veya varsayÄ±lan)
                            birim = "mÂ²"  # VarsayÄ±lan birim
                            if "mÂ³" in tanim.lower() or "metrekÃ¼p" in tanim.lower():
                                birim = "mÂ³"
                            elif "mÂ²" in tanim.lower() or "metrekare" in tanim.lower():
                                birim = "mÂ²"
                            elif "kg" in tanim.lower() or "kilogram" in tanim.lower():
                                birim = "kg"
                            elif "adet" in tanim.lower() or "ad." in tanim.lower():
                                birim = "adet"
                            elif "m" in tanim.lower() and "mÂ²" not in tanim.lower() and "mÂ³" not in tanim.lower():
                                birim = "m"
                            
                            # Kategoriyi poz numarasÄ±ndan tahmin et
                            kategori = ""
                            if poz_no.startswith("03.") or poz_no.startswith("03-"):
                                kategori = "Toprak Ä°ÅŸleri"
                            elif poz_no.startswith("04.") or poz_no.startswith("04-"):
                                kategori = "Beton Ä°ÅŸleri"
                            elif poz_no.startswith("05.") or poz_no.startswith("05-"):
                                kategori = "Demir Ä°ÅŸleri"
                            elif poz_no.startswith("15.") or poz_no.startswith("15-"):
                                kategori = "YalÄ±tÄ±m Ä°ÅŸleri"
                            else:
                                kategori = "Genel"
                            
                            self.db.add_poz(
                                poz_no=poz_no,
                                tanim=tanim[:200],  # Ä°lk 200 karakter
                                birim=birim,
                                resmi_fiyat=birim_fiyat if birim_fiyat > 0 else 0,
                                kategori=kategori,
                                fire_orani=0.05  # VarsayÄ±lan fire oranÄ±
                            )
                            poz_added_count += 1
                        except Exception as e:
                            errors.append(f"Poz {poz_no} eklenirken hata: {str(e)}")
                    
                    # SONRA BÄ°RÄ°M FÄ°YATI EKLE (varsa)
                    # Fiyat varsa hem poz tablosundaki resmi_fiyat hem de birim_fiyatlar tablosuna ekle
                    if birim_fiyat and birim_fiyat > 0:
                        try:
                            # Birim fiyatlar tablosuna ekle
                            fiyat_id = self.db.add_birim_fiyat(
                                poz_no=poz_no,
                                birim_fiyat=birim_fiyat,
                                kaynak=item.get('kaynak', 'PDF Import'),
                                aciklama=tanim[:100]
                            )
                            if fiyat_id:
                                fiyat_added_count += 1
                            
                            # Poz tablosundaki resmi_fiyat'Ä± da gÃ¼ncelle (eÄŸer poz eklendiyse)
                            if poz_added_count > 0 or poz:
                                try:
                                    with self.db.get_connection() as conn:
                                        cursor = conn.cursor()
                                        cursor.execute("""
                                            UPDATE pozlar 
                                            SET resmi_fiyat = ? 
                                            WHERE poz_no = ? AND (resmi_fiyat = 0 OR resmi_fiyat IS NULL)
                                        """, (birim_fiyat, poz_no))
                                except:
                                    pass  # GÃ¼ncelleme baÅŸarÄ±sÄ±z olsa bile devam et
                        except Exception as e:
                            errors.append(f"Poz {poz_no} fiyat eklenirken hata: {str(e)}")
                    else:
                        # Fiyat yoksa da poz eklendi, bu baÅŸarÄ±lÄ± sayÄ±lÄ±r
                        pass
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f"Poz {item.get('poz_no', '')}: {str(e)}")
                    continue
            
            progress.setValue(len(extracted_data))
            
            # SonuÃ§ mesajÄ±
            message = f"PDF iÃ§e aktarma tamamlandÄ±!\n\n"
            message += f"âœ… Toplam iÅŸlenen: {success_count}\n"
            if poz_added_count > 0:
                message += f"ğŸ“ Yeni poz eklendi: {poz_added_count}\n"
            if fiyat_added_count > 0:
                message += f"ğŸ’° Birim fiyat eklendi: {fiyat_added_count}\n"
            message += f"âŒ HatalÄ±: {error_count}"
            
            if errors and error_count <= 20:
                message += f"\n\nHatalar:\n" + "\n".join(errors[:20])
            elif errors:
                message += f"\n\n(Ä°lk 20 hata gÃ¶steriliyor, toplam {error_count} hata var)"
            
            if success_count > 0:
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", message)
                # Birim fiyat sekmesi aÃ§Ä±ksa gÃ¼ncelle
                if hasattr(self, 'fiyat_filter_combo') and self._tabs_created.get('birim_fiyat', False):
                    self.load_birim_fiyatlar()
                self.statusBar().showMessage(f"{success_count} birim fiyat iÃ§e aktarÄ±ldÄ±")
            else:
                QMessageBox.warning(self, "UyarÄ±", message)
                
        except Exception as e:
            QMessageBox.critical(
                self, "Hata",
                f"PDF dosyasÄ± iÅŸlenirken hata oluÅŸtu:\n{str(e)}"
            )
            import traceback
            traceback.print_exc()
    
    def clear_pdf_imported_data(self) -> None:
        """PDF'den eklenen pozlarÄ± ve birim fiyatlarÄ± temizle"""
        # Onay mesajÄ±
        reply = QMessageBox.question(
            self, 
            "PDF PozlarÄ± Temizle",
            "PDF'den eklenen tÃ¼m pozlarÄ± ve birim fiyatlarÄ± silmek istediÄŸinizden emin misiniz?\n\n"
            "Bu iÅŸlem geri alÄ±namaz!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        try:
            # PDF'den eklenen verileri sil
            result = self.db.delete_pdf_imported_data()
            
            poz_count = result.get('pozlar', 0)
            fiyat_count = result.get('birim_fiyatlar', 0)
            
            message = f"PDF'den eklenen veriler temizlendi!\n\n"
            message += f"âœ… Silinen poz sayÄ±sÄ±: {poz_count}\n"
            message += f"âœ… Silinen birim fiyat sayÄ±sÄ±: {fiyat_count}\n\n"
            message += "ArtÄ±k PDF'yi yeniden yÃ¼kleyebilirsiniz."
            
            QMessageBox.information(self, "BaÅŸarÄ±lÄ±", message)
            
            # Ä°lgili sekmeleri gÃ¼ncelle
            if hasattr(self, 'fiyat_filter_combo') and self._tabs_created.get('birim_fiyat', False):
                self.load_birim_fiyatlar()
            
            self.statusBar().showMessage(f"{poz_count} poz ve {fiyat_count} birim fiyat silindi")
            
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Hata",
                f"PDF verileri temizlenirken hata oluÅŸtu:\n{str(e)}"
            )
            import traceback
            traceback.print_exc()
    
    def on_search_text_changed(self) -> None:
        """GeliÅŸmiÅŸ arama ve filtreleme - TÃ¼m modÃ¼llerde arama"""
        search_text = self.search_input.text().strip()
        search_type = self.search_type_combo.currentText()
        
        if not search_text:
            # Arama boÅŸsa normal listeyi gÃ¶ster
            self.load_projects()
            if self.current_project_id:
                if hasattr(self, 'metraj_table') and self._tabs_created.get('metraj', False):
                    self.load_metraj_data()
                if hasattr(self, 'taseron_table') and self._tabs_created.get('taseron', False):
                    self.load_taseron_data()
            if hasattr(self, 'ihale_kalem_table') and self._tabs_created.get('ihale', False):
                if hasattr(self, 'current_ihale_id') and self.current_ihale_id:
                    self.load_ihale_kalemleri()
            return
        
        search_lower = search_text.lower()
        
        # Proje aramasÄ±
        if search_type in ["TÃ¼mÃ¼", "Projeler"]:
            projects = self.db.get_all_projects()
            self.project_tree.clear()
            for project in projects:
                project_name = project['ad'].lower()
                project_desc = (project.get('aciklama', '') or '').lower()
                project_notes = (project.get('notlar', '') or '').lower()
                
                if (search_lower in project_name or 
                    search_lower in project_desc or 
                    search_lower in project_notes):
                    item = QTreeWidgetItem(self.project_tree)
                    item.setText(0, project['ad'])
                    item.setData(0, Qt.ItemDataRole.UserRole, project['id'])
        
        # Kalem aramasÄ± (seÃ§ili projede - Metraj)
        if search_type in ["TÃ¼mÃ¼", "Kalemler"] and self.current_project_id:
            if hasattr(self, 'metraj_table') and self._tabs_created.get('metraj', False):
                metraj_items = self.db.get_project_metraj(self.current_project_id)
                filtered_items = []
                for item in metraj_items:
                    tanim = (item.get('tanim', '') or '').lower()
                    poz_no = (item.get('poz_no', '') or '').lower()
                    kategori = (item.get('kategori', '') or '').lower()
                    notlar = (item.get('notlar', '') or '').lower()
                    
                    if (search_lower in tanim or
                        search_lower in poz_no or
                        search_lower in kategori or
                        search_lower in notlar):
                        filtered_items.append(item)
            
                # Metraj tablosunu filtrele
                self.metraj_table.setRowCount(len(filtered_items))
                for row, item in enumerate(filtered_items):
                    self.metraj_table.setItem(row, 0, QTableWidgetItem(str(item.get('id', ''))))
                    self.metraj_table.setItem(row, 1, QTableWidgetItem(item.get('poz_no', '')))
                    self.metraj_table.setItem(row, 2, QTableWidgetItem(item.get('tanim', '')))
                    self.metraj_table.setItem(row, 3, QTableWidgetItem(f"{item.get('miktar', 0):,.2f}"))
                    self.metraj_table.setItem(row, 4, QTableWidgetItem(item.get('birim', '')))
                    self.metraj_table.setItem(row, 5, QTableWidgetItem(f"{item.get('birim_fiyat', 0):,.2f}"))
                    self.metraj_table.setItem(row, 6, QTableWidgetItem(f"{item.get('toplam', 0):,.2f}"))
                
                # ToplamÄ± gÃ¼ncelle (KDV ile)
                toplam = sum(item.get('toplam', 0) for item in filtered_items)
                kdv_rate_text = self.metraj_kdv_rate.currentText().replace("%", "")
                kdv_rate = float(kdv_rate_text)
                kdv_hesap = self.calculator.calculate_with_kdv(toplam, kdv_rate)
                self.total_label.setText(f"Toplam (KDV HariÃ§): {toplam:,.2f} â‚º (FiltrelenmiÅŸ: {len(filtered_items)} kalem)")
                self.total_kdv_label.setText(f"Toplam (KDV %{kdv_rate_text} Dahil): {kdv_hesap['kdv_dahil']:,.2f} â‚º")
        
        # Poz aramasÄ± (tÃ¼m pozlar)
        if search_type in ["TÃ¼mÃ¼", "Pozlar"]:
            if hasattr(self, 'birim_fiyat_table') and self._tabs_created.get('birim_fiyat', False):
                pozlar = self.db.search_pozlar(search_text, limit=100)
                self.birim_fiyat_table.setRowCount(len(pozlar))
                for row, poz in enumerate(pozlar):
                    self.birim_fiyat_table.setItem(row, 0, QTableWidgetItem(poz.get('poz_no', '')))
                    self.birim_fiyat_table.setItem(row, 1, QTableWidgetItem(poz.get('tanim', '')))
                    self.birim_fiyat_table.setItem(row, 2, QTableWidgetItem(poz.get('birim', '')))
                    self.birim_fiyat_table.setItem(row, 3, QTableWidgetItem(f"{poz.get('resmi_fiyat', 0):,.2f}"))
                    self.birim_fiyat_table.setItem(row, 4, QTableWidgetItem(poz.get('kategori', '')))
        
        # Ä°hale kalemleri aramasÄ±
        if search_type in ["TÃ¼mÃ¼", "Kalemler"]:
            if hasattr(self, 'ihale_kalem_table') and self._tabs_created.get('ihale', False):
                if hasattr(self, 'current_ihale_id') and self.current_ihale_id:
                    kalemler = self.db.get_ihale_kalemleri(self.current_ihale_id)
                    filtered_kalemler = []
                    for kalem in kalemler:
                        poz_no = (kalem.get('poz_no', '') or '').lower()
                        tanim = (kalem.get('poz_tanim', '') or '').lower()
                        kategori = (kalem.get('kategori', '') or '').lower()
                        
                        if (search_lower in poz_no or
                            search_lower in tanim or
                            search_lower in kategori):
                            filtered_kalemler.append(kalem)
                    
                    # Ä°hale tablosunu filtrele
                    self.ihale_kalem_table.setRowCount(len(filtered_kalemler))
                    for row, kalem in enumerate(filtered_kalemler):
                        self.ihale_kalem_table.setItem(row, 0, QTableWidgetItem(str(kalem.get('sira_no', ''))))
                        self.ihale_kalem_table.setItem(row, 1, QTableWidgetItem(kalem.get('poz_no', '')))
                        self.ihale_kalem_table.setItem(row, 2, QTableWidgetItem(kalem.get('poz_tanim', '')))
                        self.ihale_kalem_table.setItem(row, 3, QTableWidgetItem(f"{kalem.get('birim_miktar', 0):,.2f}"))
                        self.ihale_kalem_table.setItem(row, 4, QTableWidgetItem(kalem.get('birim', '')))
                        self.ihale_kalem_table.setItem(row, 5, QTableWidgetItem(f"{kalem.get('birim_fiyat', 0):,.2f}"))
                        self.ihale_kalem_table.setItem(row, 6, QTableWidgetItem(f"{kalem.get('toplam', 0):,.2f}"))
                    
                    # ToplamÄ± gÃ¼ncelle
                    toplam = sum(kalem.get('toplam', 0) for kalem in filtered_kalemler)
                    if hasattr(self, 'ihale_total_label'):
                        # KDV hesaplama
                        kdv_rate_text = self.ihale_kdv_rate.currentText().replace("%", "")
                        kdv_rate = float(kdv_rate_text)
                        kdv_hesap = self.calculator.calculate_with_kdv(toplam, kdv_rate)
                        
                        self.ihale_total_label.setText(f"Toplam (KDV HariÃ§): {toplam:,.2f} â‚º (FiltrelenmiÅŸ: {len(filtered_kalemler)} kalem)")
                        self.ihale_total_kdv_label.setText(f"Toplam (KDV %{kdv_rate_text} Dahil): {kdv_hesap['kdv_dahil']:,.2f} â‚º")
        
        # TaÅŸeron aramasÄ±
        if search_type in ["TÃ¼mÃ¼", "Kalemler"]:
            if hasattr(self, 'taseron_table') and self._tabs_created.get('taseron', False):
                if self.current_project_id:
                    teklifler = self.db.get_taseron_teklifleri(self.current_project_id)
                    filtered_teklifler = []
                    for teklif in teklifler:
                        firma = (teklif.get('firma_adi', '') or '').lower()
                        poz_no = (teklif.get('poz_no', '') or '').lower()
                        tanim = (teklif.get('tanim', '') or '').lower()
                        notlar = (teklif.get('notlar', '') or '').lower()
                        
                        if (search_lower in firma or
                            search_lower in poz_no or
                            search_lower in tanim or
                            search_lower in notlar):
                            filtered_teklifler.append(teklif)
                    
                    # TaÅŸeron tablosunu filtrele
                    self.taseron_table.setRowCount(len(filtered_teklifler))
                    for row, teklif in enumerate(filtered_teklifler):
                        self.taseron_table.setItem(row, 0, QTableWidgetItem(teklif.get('firma_adi', '')))
                        self.taseron_table.setItem(row, 1, QTableWidgetItem(teklif.get('poz_no', '')))
                        self.taseron_table.setItem(row, 2, QTableWidgetItem(teklif.get('tanim', '')))
                        self.taseron_table.setItem(row, 3, QTableWidgetItem(f"{teklif.get('miktar', 0):,.2f}"))
                        self.taseron_table.setItem(row, 4, QTableWidgetItem(teklif.get('birim', '')))
                        self.taseron_table.setItem(row, 5, QTableWidgetItem(f"{teklif.get('fiyat', 0):,.2f}"))
                        self.taseron_table.setItem(row, 6, QTableWidgetItem(f"{teklif.get('toplam', 0):,.2f}"))
    
    def load_templates(self) -> None:
        """ÅablonlarÄ± yÃ¼kle"""
        from PyQt6.QtWidgets import QApplication
        QApplication.processEvents()  # UI'Ä± gÃ¼ncelle
        
        templates = self.db.get_all_templates()
        self.template_table.setRowCount(len(templates))
        
        for row, template in enumerate(templates):
            self.template_table.setItem(row, 0, QTableWidgetItem(template.get('ad', '')))
            self.template_table.setItem(row, 1, QTableWidgetItem(template.get('aciklama', '')))
            tarih = template.get('olusturma_tarihi', '')[:10] if template.get('olusturma_tarihi') else ''
            self.template_table.setItem(row, 2, QTableWidgetItem(tarih))
            
            # Kalem sayÄ±sÄ±nÄ± al
            items = self.db.get_template_items(template['id'])
            self.template_table.setItem(row, 3, QTableWidgetItem(str(len(items))))
            
            # ID'yi sakla
            item = self.template_table.item(row, 0)
            if item:
                item.setData(Qt.ItemDataRole.UserRole, template['id'])
            
            # Her 10 ÅŸablonda bir UI'Ä± gÃ¼ncelle
            if row % 10 == 0:
                QApplication.processEvents()
    
    def view_template_items(self, item: QTableWidgetItem) -> None:
        """Åablon kalemlerini gÃ¶ster"""
        row = item.row()
        template_item = self.template_table.item(row, 0)
        if not template_item:
            return
        
        template_id = template_item.data(Qt.ItemDataRole.UserRole)
        if not template_id:
            return
        
        self.current_template_id = template_id
        items = self.db.get_template_items(template_id)
        self.template_items_table.setRowCount(len(items))
        
        for row_idx, item_data in enumerate(items):
            # ID (gizli)
            id_item = QTableWidgetItem(str(item_data.get('id', '')))
            id_item.setData(Qt.ItemDataRole.UserRole, item_data.get('id'))
            self.template_items_table.setItem(row_idx, 0, id_item)
            
            self.template_items_table.setItem(row_idx, 1, QTableWidgetItem(item_data.get('poz_no', '')))
            self.template_items_table.setItem(row_idx, 2, QTableWidgetItem(item_data.get('tanim', '')))
            self.template_items_table.setItem(row_idx, 3, QTableWidgetItem(item_data.get('kategori', '')))
            self.template_items_table.setItem(row_idx, 4, QTableWidgetItem(f"{item_data.get('miktar', 0):,.2f}"))
            self.template_items_table.setItem(row_idx, 5, QTableWidgetItem(item_data.get('birim', '')))
            self.template_items_table.setItem(row_idx, 6, QTableWidgetItem(f"{item_data.get('birim_fiyat', 0):,.2f}"))
            self.template_items_table.setItem(row_idx, 7, QTableWidgetItem(f"{item_data.get('toplam', 0):,.2f}"))
    
    def create_template_from_project(self) -> None:
        """Mevcut projeden ÅŸablon oluÅŸtur"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
        
        project = self.db.get_project(self.current_project_id)
        if not project:
            QMessageBox.warning(self, "UyarÄ±", "Proje bulunamadÄ±")
            return
        
        # Åablon adÄ± ve aÃ§Ä±klamasÄ± sor
        from PyQt6.QtWidgets import QInputDialog
        
        template_name, ok1 = QInputDialog.getText(
            self, "Åablon OluÅŸtur",
            f"Åablon adÄ±:\n(Proje: {project['ad']})"
        )
        
        if not ok1 or not template_name.strip():
            return
        
        template_description, ok2 = QInputDialog.getText(
            self, "Åablon AÃ§Ä±klamasÄ±",
            "Åablon aÃ§Ä±klamasÄ± (isteÄŸe baÄŸlÄ±):"
        )
        
        if not ok2:
            return
        
        # Åablon oluÅŸtur
        template_id = self.db.create_template_from_project(
            self.current_project_id,
            template_name.strip(),
            template_description.strip()
        )
        
        if template_id:
            QMessageBox.information(
                self, "BaÅŸarÄ±lÄ±",
                f"Åablon baÅŸarÄ±yla oluÅŸturuldu!\n\n"
                f"Åablon adÄ±: {template_name}\n"
                f"Kalem sayÄ±sÄ±: {len(self.db.get_project_metraj(self.current_project_id))}"
            )
            self.load_templates()
            self.statusBar().showMessage(f"Åablon oluÅŸturuldu: {template_name}")
        else:
            QMessageBox.critical(self, "Hata", "Åablon oluÅŸturulurken bir hata oluÅŸtu")
    
    def create_project_from_template(self) -> None:
        """Åablondan proje oluÅŸtur"""
        current_row = self.template_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen bir ÅŸablon seÃ§in")
            return
        
        template_item = self.template_table.item(current_row, 0)
        if not template_item:
            return
        
        template_id = template_item.data(Qt.ItemDataRole.UserRole)
        if not template_id:
            return
        
        template = self.db.get_template(template_id)
        if not template:
            QMessageBox.warning(self, "UyarÄ±", "Åablon bulunamadÄ±")
            return
        
        # Proje adÄ± ve aÃ§Ä±klamasÄ± sor
        from PyQt6.QtWidgets import QInputDialog
        
        project_name, ok1 = QInputDialog.getText(
            self, "Proje OluÅŸtur",
            f"Yeni proje adÄ±:\n(Åablon: {template['ad']})"
        )
        
        if not ok1 or not project_name.strip():
            return
        
        project_description, ok2 = QInputDialog.getText(
            self, "Proje AÃ§Ä±klamasÄ±",
            "Proje aÃ§Ä±klamasÄ± (isteÄŸe baÄŸlÄ±):"
        )
        
        if not ok2:
            return
        
        # Proje oluÅŸtur
        project_id = self.db.create_project_from_template(
            template_id,
            project_name.strip(),
            project_description.strip()
        )
        
        if project_id:
            QMessageBox.information(
                self, "BaÅŸarÄ±lÄ±",
                f"Proje baÅŸarÄ±yla oluÅŸturuldu!\n\n"
                f"Proje adÄ±: {project_name}\n"
                f"Kalem sayÄ±sÄ±: {len(self.db.get_template_items(template_id))}"
            )
            # Proje listesini yenile
            self.load_projects()
            # Yeni projeyi seÃ§
            self.current_project_id = project_id
            self.load_metraj_data()
            self.load_taseron_data()
            self.update_proje_ozet()
            self.load_project_notes()
            self.statusBar().showMessage(f"Proje oluÅŸturuldu: {project_name}")
        else:
            QMessageBox.critical(self, "Hata", "Proje oluÅŸturulurken bir hata oluÅŸtu")
    
    def delete_template(self) -> None:
        """Åablonu sil"""
        current_row = self.template_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen silmek istediÄŸiniz ÅŸablonu seÃ§in")
            return
        
        template_item = self.template_table.item(current_row, 0)
        if not template_item:
            return
        
        template_id = template_item.data(Qt.ItemDataRole.UserRole)
        template_name = template_item.text()
        
        if not template_id:
            return
        
        # Onay al
        reply = QMessageBox.question(
            self, "Åablon Sil",
            f"'{template_name}' ÅŸablonunu silmek istediÄŸinize emin misiniz?\n\n"
            "Bu iÅŸlem geri alÄ±namaz!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            if self.db.delete_template(template_id):
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", "Åablon silindi")
                self.load_templates()
                self.template_items_table.setRowCount(0)
                self.statusBar().showMessage("Åablon silindi")
            else:
                QMessageBox.critical(self, "Hata", "Åablon silinirken bir hata oluÅŸtu")
    
    def create_empty_template(self) -> None:
        """BoÅŸ ÅŸablon oluÅŸtur"""
        from PyQt6.QtWidgets import QInputDialog
        
        template_name, ok1 = QInputDialog.getText(
            self, "BoÅŸ Åablon OluÅŸtur",
            "Åablon adÄ±:"
        )
        
        if not ok1 or not template_name.strip():
            return
        
        template_description, ok2 = QInputDialog.getText(
            self, "Åablon AÃ§Ä±klamasÄ±",
            "Åablon aÃ§Ä±klamasÄ± (isteÄŸe baÄŸlÄ±):"
        )
        
        if not ok2:
            return
        
        template_id = self.db.create_template(template_name.strip(), template_description.strip())
        
        if template_id:
            QMessageBox.information(
                self, "BaÅŸarÄ±lÄ±",
                f"BoÅŸ ÅŸablon baÅŸarÄ±yla oluÅŸturuldu!\n\n"
                f"Åablon adÄ±: {template_name}\n"
                f"Åimdi ÅŸablona kalem ekleyebilirsiniz."
            )
            self.load_templates()
            # Yeni oluÅŸturulan ÅŸablonu seÃ§
            for row in range(self.template_table.rowCount()):
                item = self.template_table.item(row, 0)
                if item and item.data(Qt.ItemDataRole.UserRole) == template_id:
                    self.template_table.selectRow(row)
                    self.current_template_id = template_id
                    self.view_template_items(item)
                    break
            self.statusBar().showMessage(f"BoÅŸ ÅŸablon oluÅŸturuldu: {template_name}")
        else:
            QMessageBox.critical(self, "Hata", "Åablon oluÅŸturulurken bir hata oluÅŸtu")
    
    def create_template_from_categories(self) -> None:
        """Kategori bazlÄ± ÅŸablon oluÅŸtur"""
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QCheckBox, QScrollArea, QWidget, QPushButton, QHBoxLayout, QMessageBox
        
        # Kategorileri al
        kategoriler = self.db.get_all_kategoriler()
        
        if not kategoriler:
            QMessageBox.warning(self, "UyarÄ±", "Sistemde kategori bulunamadÄ±. Ã–nce projeler oluÅŸturun.")
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Kategori BazlÄ± Åablon OluÅŸtur")
        dialog.setMinimumWidth(500)
        dialog.setMinimumHeight(400)
        
        layout = QVBoxLayout(dialog)
        
        # Åablon adÄ± ve aÃ§Ä±klama
        name_layout = QHBoxLayout()
        name_layout.addWidget(QLabel("Åablon AdÄ±:"))
        name_input = QLineEdit()
        name_layout.addWidget(name_input)
        layout.addLayout(name_layout)
        
        desc_layout = QHBoxLayout()
        desc_layout.addWidget(QLabel("AÃ§Ä±klama:"))
        desc_input = QLineEdit()
        desc_layout.addWidget(desc_input)
        layout.addLayout(desc_layout)
        
        layout.addWidget(QLabel("Kategorileri seÃ§in:"))
        
        # Kategori checkbox'larÄ±
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        
        category_checkboxes = {}
        for kategori in kategoriler:
            checkbox = QCheckBox(kategori)
            category_checkboxes[kategori] = checkbox
            scroll_layout.addWidget(checkbox)
        
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        layout.addWidget(scroll)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_ok = QPushButton("OluÅŸtur")
        btn_cancel = QPushButton("Ä°ptal")
        btn_ok.clicked.connect(dialog.accept)
        btn_cancel.clicked.connect(dialog.reject)
        btn_layout.addWidget(btn_ok)
        btn_layout.addWidget(btn_cancel)
        layout.addLayout(btn_layout)
        
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        
        template_name = name_input.text().strip()
        if not template_name:
            QMessageBox.warning(self, "UyarÄ±", "Åablon adÄ± boÅŸ olamaz")
            return
        
        # SeÃ§ili kategorileri al
        selected_categories = [k for k, cb in category_checkboxes.items() if cb.isChecked()]
        
        if not selected_categories:
            QMessageBox.warning(self, "UyarÄ±", "En az bir kategori seÃ§melisiniz")
            return
        
        # Åablon oluÅŸtur
        template_id = self.db.create_template(template_name, desc_input.text().strip())
        
        if not template_id:
            QMessageBox.critical(self, "Hata", "Åablon oluÅŸturulurken bir hata oluÅŸtu")
            return
        
        # Kategorilerden kalemleri al ve ekle
        items = self.db.get_items_by_kategoriler(selected_categories)
        
        if not items:
            QMessageBox.warning(self, "UyarÄ±", "SeÃ§ili kategorilerde kalem bulunamadÄ±")
            self.db.delete_template(template_id)
            return
        
        added_count = 0
        for item in items:
            birim_fiyat = item.get('birim_fiyat', 0) or 0
            self.db.add_template_item(
                sablon_id=template_id,
                poz_no=item.get('poz_no', ''),
                tanim=item.get('tanim', ''),
                kategori=item.get('kategori', ''),
                miktar=0,  # VarsayÄ±lan miktar 0
                birim=item.get('birim', ''),
                birim_fiyat=birim_fiyat,
                toplam=0
            )
            added_count += 1
        
        QMessageBox.information(
            self, "BaÅŸarÄ±lÄ±",
            f"Kategori bazlÄ± ÅŸablon oluÅŸturuldu!\n\n"
            f"Åablon adÄ±: {template_name}\n"
            f"Eklenen kalem sayÄ±sÄ±: {added_count}"
        )
        self.load_templates()
        self.statusBar().showMessage(f"Kategori bazlÄ± ÅŸablon oluÅŸturuldu: {template_name}")
    
    def create_template_from_preset(self) -> None:
        """HazÄ±r ÅŸablon ÅŸablonlarÄ±ndan oluÅŸtur"""
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QComboBox, QPushButton, QHBoxLayout
        
        # HazÄ±r ÅŸablon ÅŸablonlarÄ±
        presets = {
            "100mÂ² Konut Standart": {
                "kategoriler": ["Beton Ä°ÅŸleri", "Demir Ä°ÅŸleri", "SÄ±va Ä°ÅŸleri", "Boyama Ä°ÅŸleri", "Elektrik Ä°ÅŸleri", "Tesisat Ä°ÅŸleri"],
                "aciklama": "100mÂ² standart konut iÃ§in temel iÅŸ kalemleri"
            },
            "200mÂ² Villa Standart": {
                "kategoriler": ["Beton Ä°ÅŸleri", "Demir Ä°ÅŸleri", "SÄ±va Ä°ÅŸleri", "Boyama Ä°ÅŸleri", "Elektrik Ä°ÅŸleri", "Tesisat Ä°ÅŸleri", "YalÄ±tÄ±m Ä°ÅŸleri"],
                "aciklama": "200mÂ² villa iÃ§in standart iÅŸ kalemleri"
            },
            "Ticari Bina Standart": {
                "kategoriler": ["Beton Ä°ÅŸleri", "Demir Ä°ÅŸleri", "SÄ±va Ä°ÅŸleri", "Boyama Ä°ÅŸleri", "Elektrik Ä°ÅŸleri", "Tesisat Ä°ÅŸleri", "Asma Tavan Ä°ÅŸleri"],
                "aciklama": "Ticari bina iÃ§in standart iÅŸ kalemleri"
            },
            "AltyapÄ± Ä°ÅŸleri": {
                "kategoriler": ["Toprak Ä°ÅŸleri", "Beton Ä°ÅŸleri", "Demir Ä°ÅŸleri", "Yol Ä°ÅŸleri"],
                "aciklama": "AltyapÄ± projeleri iÃ§in temel iÅŸ kalemleri"
            }
        }
        
        dialog = QDialog(self)
        dialog.setWindowTitle("HazÄ±r Åablonlardan OluÅŸtur")
        dialog.setMinimumWidth(400)
        
        layout = QVBoxLayout(dialog)
        
        layout.addWidget(QLabel("HazÄ±r ÅŸablon seÃ§in:"))
        
        preset_combo = QComboBox()
        preset_combo.addItems(list(presets.keys()))
        layout.addWidget(preset_combo)
        
        desc_label = QLabel()
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)
        
        def update_description():
            selected = preset_combo.currentText()
            desc_label.setText(f"AÃ§Ä±klama: {presets[selected]['aciklama']}")
        
        preset_combo.currentTextChanged.connect(update_description)
        update_description()
        
        layout.addWidget(QLabel("Åablon AdÄ±:"))
        name_input = QLineEdit()
        layout.addWidget(name_input)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_ok = QPushButton("OluÅŸtur")
        btn_cancel = QPushButton("Ä°ptal")
        btn_ok.clicked.connect(dialog.accept)
        btn_cancel.clicked.connect(dialog.reject)
        btn_layout.addWidget(btn_ok)
        btn_layout.addWidget(btn_cancel)
        layout.addLayout(btn_layout)
        
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        
        selected_preset = preset_combo.currentText()
        template_name = name_input.text().strip() or selected_preset
        
        preset_data = presets[selected_preset]
        
        # Åablon oluÅŸtur
        template_id = self.db.create_template(template_name, preset_data['aciklama'])
        
        if not template_id:
            QMessageBox.critical(self, "Hata", "Åablon oluÅŸturulurken bir hata oluÅŸtu")
            return
        
        # Kategorilerden kalemleri al ve ekle
        items = self.db.get_items_by_kategoriler(preset_data['kategoriler'])
        
        added_count = 0
        for item in items:
            birim_fiyat = item.get('birim_fiyat', 0) or 0
            self.db.add_template_item(
                sablon_id=template_id,
                poz_no=item.get('poz_no', ''),
                tanim=item.get('tanim', ''),
                kategori=item.get('kategori', ''),
                miktar=0,
                birim=item.get('birim', ''),
                birim_fiyat=birim_fiyat,
                toplam=0
            )
            added_count += 1
        
        QMessageBox.information(
            self, "BaÅŸarÄ±lÄ±",
            f"HazÄ±r ÅŸablondan ÅŸablon oluÅŸturuldu!\n\n"
            f"Åablon adÄ±: {template_name}\n"
            f"Eklenen kalem sayÄ±sÄ±: {added_count}"
        )
        self.load_templates()
        self.statusBar().showMessage(f"HazÄ±r ÅŸablondan ÅŸablon oluÅŸturuldu: {template_name}")
    
    def edit_template(self) -> None:
        """Åablon bilgilerini dÃ¼zenle"""
        current_row = self.template_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen dÃ¼zenlemek istediÄŸiniz ÅŸablonu seÃ§in")
            return
        
        template_item = self.template_table.item(current_row, 0)
        if not template_item:
            return
        
        template_id = template_item.data(Qt.ItemDataRole.UserRole)
        if not template_id:
            return
        
        template = self.db.get_template(template_id)
        if not template:
            QMessageBox.warning(self, "UyarÄ±", "Åablon bulunamadÄ±")
            return
        
        from PyQt6.QtWidgets import QInputDialog
        
        new_name, ok1 = QInputDialog.getText(
            self, "Åablon AdÄ±nÄ± DÃ¼zenle",
            "Yeni ÅŸablon adÄ±:",
            text=template.get('ad', '')
        )
        
        if not ok1:
            return
        
        new_description, ok2 = QInputDialog.getText(
            self, "Åablon AÃ§Ä±klamasÄ±nÄ± DÃ¼zenle",
            "Yeni aÃ§Ä±klama:",
            text=template.get('aciklama', '')
        )
        
        if not ok2:
            return
        
        if self.db.update_template(template_id, ad=new_name.strip(), aciklama=new_description.strip()):
            QMessageBox.information(self, "BaÅŸarÄ±lÄ±", "Åablon gÃ¼ncellendi")
            self.load_templates()
            self.statusBar().showMessage("Åablon gÃ¼ncellendi")
        else:
            QMessageBox.critical(self, "Hata", "Åablon gÃ¼ncellenirken bir hata oluÅŸtu")
    
    def copy_template(self) -> None:
        """Åablonu kopyala"""
        current_row = self.template_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen kopyalamak istediÄŸiniz ÅŸablonu seÃ§in")
            return
        
        template_item = self.template_table.item(current_row, 0)
        if not template_item:
            return
        
        template_id = template_item.data(Qt.ItemDataRole.UserRole)
        template_name = template_item.text()
        
        if not template_id:
            return
        
        from PyQt6.QtWidgets import QInputDialog
        
        new_name, ok1 = QInputDialog.getText(
            self, "Åablon Kopyala",
            f"Yeni ÅŸablon adÄ±:\n(Orjinal: {template_name})",
            text=f"{template_name} (Kopya)"
        )
        
        if not ok1 or not new_name.strip():
            return
        
        new_description, ok2 = QInputDialog.getText(
            self, "Åablon AÃ§Ä±klamasÄ±",
            "Yeni aÃ§Ä±klama (isteÄŸe baÄŸlÄ±):"
        )
        
        if not ok2:
            return
        
        new_template_id = self.db.copy_template(template_id, new_name.strip(), new_description.strip())
        
        if new_template_id:
            items_count = len(self.db.get_template_items(new_template_id))
            QMessageBox.information(
                self, "BaÅŸarÄ±lÄ±",
                f"Åablon baÅŸarÄ±yla kopyalandÄ±!\n\n"
                f"Yeni ÅŸablon adÄ±: {new_name}\n"
                f"Kalem sayÄ±sÄ±: {items_count}"
            )
            self.load_templates()
            self.statusBar().showMessage(f"Åablon kopyalandÄ±: {new_name}")
        else:
            QMessageBox.critical(self, "Hata", "Åablon kopyalanÄ±rken bir hata oluÅŸtu")
    
    def add_template_item(self) -> None:
        """Åablona kalem ekle"""
        if not self.current_template_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir ÅŸablon seÃ§in")
            return
        
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QLineEdit, QPushButton, QHBoxLayout, QDoubleSpinBox, QComboBox
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Åablona Kalem Ekle")
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout(dialog)
        
        # Poz No
        layout.addWidget(QLabel("Poz No:"))
        poz_no_input = QLineEdit()
        layout.addWidget(poz_no_input)
        
        # TanÄ±m
        layout.addWidget(QLabel("TanÄ±m:"))
        tanim_input = QLineEdit()
        layout.addWidget(tanim_input)
        
        # Kategori
        layout.addWidget(QLabel("Kategori:"))
        kategori_input = QLineEdit()
        layout.addWidget(kategori_input)
        
        # Miktar
        layout.addWidget(QLabel("Miktar:"))
        miktar_input = QDoubleSpinBox()
        miktar_input.setMinimum(0)
        miktar_input.setMaximum(999999)
        miktar_input.setDecimals(2)
        layout.addWidget(miktar_input)
        
        # Birim
        layout.addWidget(QLabel("Birim:"))
        birim_input = QComboBox()
        birim_input.setEditable(True)
        birim_input.addItems(["mÂ²", "mÂ³", "m", "kg", "adet", "ton", "lt"])
        layout.addWidget(birim_input)
        
        # Birim Fiyat
        layout.addWidget(QLabel("Birim Fiyat:"))
        birim_fiyat_input = QDoubleSpinBox()
        birim_fiyat_input.setMinimum(0)
        birim_fiyat_input.setMaximum(999999)
        birim_fiyat_input.setDecimals(2)
        layout.addWidget(birim_fiyat_input)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_ok = QPushButton("Ekle")
        btn_cancel = QPushButton("Ä°ptal")
        
        def calculate_and_add():
            miktar = miktar_input.value()
            birim_fiyat = birim_fiyat_input.value()
            toplam = miktar * birim_fiyat
            
            if not tanim_input.text().strip():
                QMessageBox.warning(dialog, "UyarÄ±", "TanÄ±m boÅŸ olamaz")
                return
            
            item_id = self.db.add_template_item(
                sablon_id=self.current_template_id,
                poz_no=poz_no_input.text().strip(),
                tanim=tanim_input.text().strip(),
                kategori=kategori_input.text().strip(),
                miktar=miktar,
                birim=birim_input.currentText().strip(),
                birim_fiyat=birim_fiyat,
                toplam=toplam
            )
            
            if item_id:
                QMessageBox.information(dialog, "BaÅŸarÄ±lÄ±", "Kalem eklendi")
                dialog.accept()
                # Åablon kalemlerini yenile
                template_item = self.template_table.item(self.template_table.currentRow(), 0)
                if template_item:
                    self.view_template_items(template_item)
                self.load_templates()
            else:
                QMessageBox.critical(dialog, "Hata", "Kalem eklenirken bir hata oluÅŸtu")
        
        btn_ok.clicked.connect(calculate_and_add)
        btn_cancel.clicked.connect(dialog.reject)
        btn_layout.addWidget(btn_ok)
        btn_layout.addWidget(btn_cancel)
        layout.addLayout(btn_layout)
        
        dialog.exec()
    
    def edit_template_item(self) -> None:
        """Åablon kalemini dÃ¼zenle"""
        if not self.current_template_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir ÅŸablon seÃ§in")
            return
        
        current_row = self.template_items_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen dÃ¼zenlemek istediÄŸiniz kalemi seÃ§in")
            return
        
        id_item = self.template_items_table.item(current_row, 0)
        if not id_item:
            return
        
        item_id = id_item.data(Qt.ItemDataRole.UserRole)
        if not item_id:
            return
        
        item = self.db.get_template_item(item_id)
        if not item:
            QMessageBox.warning(self, "UyarÄ±", "Kalem bulunamadÄ±")
            return
        
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QLineEdit, QPushButton, QHBoxLayout, QDoubleSpinBox, QComboBox
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Åablon Kalemini DÃ¼zenle")
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout(dialog)
        
        # Poz No
        layout.addWidget(QLabel("Poz No:"))
        poz_no_input = QLineEdit(item.get('poz_no', ''))
        layout.addWidget(poz_no_input)
        
        # TanÄ±m
        layout.addWidget(QLabel("TanÄ±m:"))
        tanim_input = QLineEdit(item.get('tanim', ''))
        layout.addWidget(tanim_input)
        
        # Kategori
        layout.addWidget(QLabel("Kategori:"))
        kategori_input = QLineEdit(item.get('kategori', ''))
        layout.addWidget(kategori_input)
        
        # Miktar
        layout.addWidget(QLabel("Miktar:"))
        miktar_input = QDoubleSpinBox()
        miktar_input.setMinimum(0)
        miktar_input.setMaximum(999999)
        miktar_input.setDecimals(2)
        miktar_input.setValue(item.get('miktar', 0))
        layout.addWidget(miktar_input)
        
        # Birim
        layout.addWidget(QLabel("Birim:"))
        birim_input = QComboBox()
        birim_input.setEditable(True)
        birim_input.addItems(["mÂ²", "mÂ³", "m", "kg", "adet", "ton", "lt"])
        birim_input.setCurrentText(item.get('birim', ''))
        layout.addWidget(birim_input)
        
        # Birim Fiyat
        layout.addWidget(QLabel("Birim Fiyat:"))
        birim_fiyat_input = QDoubleSpinBox()
        birim_fiyat_input.setMinimum(0)
        birim_fiyat_input.setMaximum(999999)
        birim_fiyat_input.setDecimals(2)
        birim_fiyat_input.setValue(item.get('birim_fiyat', 0))
        layout.addWidget(birim_fiyat_input)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_ok = QPushButton("Kaydet")
        btn_cancel = QPushButton("Ä°ptal")
        
        def calculate_and_save():
            miktar = miktar_input.value()
            birim_fiyat = birim_fiyat_input.value()
            toplam = miktar * birim_fiyat
            
            if not tanim_input.text().strip():
                QMessageBox.warning(dialog, "UyarÄ±", "TanÄ±m boÅŸ olamaz")
                return
            
            if self.db.update_template_item(
                item_id,
                poz_no=poz_no_input.text().strip(),
                tanim=tanim_input.text().strip(),
                kategori=kategori_input.text().strip(),
                miktar=miktar,
                birim=birim_input.currentText().strip(),
                birim_fiyat=birim_fiyat,
                toplam=toplam
            ):
                QMessageBox.information(dialog, "BaÅŸarÄ±lÄ±", "Kalem gÃ¼ncellendi")
                dialog.accept()
                # Åablon kalemlerini yenile
                template_item = self.template_table.item(self.template_table.currentRow(), 0)
                if template_item:
                    self.view_template_items(template_item)
            else:
                QMessageBox.critical(dialog, "Hata", "Kalem gÃ¼ncellenirken bir hata oluÅŸtu")
        
        btn_ok.clicked.connect(calculate_and_save)
        btn_cancel.clicked.connect(dialog.reject)
        btn_layout.addWidget(btn_ok)
        btn_layout.addWidget(btn_cancel)
        layout.addLayout(btn_layout)
        
        dialog.exec()
    
    def delete_template_item(self) -> None:
        """Åablon kalemini sil"""
        if not self.current_template_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir ÅŸablon seÃ§in")
            return
        
        current_row = self.template_items_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen silmek istediÄŸiniz kalemi seÃ§in")
            return
        
        id_item = self.template_items_table.item(current_row, 0)
        if not id_item:
            return
        
        item_id = id_item.data(Qt.ItemDataRole.UserRole)
        tanim = self.template_items_table.item(current_row, 2).text()
        
        if not item_id:
            return
        
        reply = QMessageBox.question(
            self, "Kalem Sil",
            f"'{tanim}' kalemini silmek istediÄŸinize emin misiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            if self.db.delete_template_item(item_id):
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", "Kalem silindi")
                # Åablon kalemlerini yenile
                template_item = self.template_table.item(self.template_table.currentRow(), 0)
                if template_item:
                    self.view_template_items(template_item)
                self.load_templates()
            else:
                QMessageBox.critical(self, "Hata", "Kalem silinirken bir hata oluÅŸtu")
    
    def load_ihaleler(self) -> None:
        """Ä°haleleri yÃ¼kle"""
        from PyQt6.QtWidgets import QApplication
        QApplication.processEvents()  # UI'Ä± gÃ¼ncelle
        
        ihaleler = self.db.get_all_ihaleler()
        self.ihale_combo.clear()
        self.ihale_combo.addItem("-- Ä°hale SeÃ§in --", None)
        for ihale in ihaleler:
            self.ihale_combo.addItem(ihale['ad'], ihale['id'])
            QApplication.processEvents()  # UI'Ä± gÃ¼ncelle
    
    def on_ihale_changed(self) -> None:
        """Ä°hale seÃ§ildiÄŸinde"""
        ihale_id = self.ihale_combo.currentData()
        self.current_ihale_id = ihale_id
        if ihale_id:
            if hasattr(self, 'ihale_total_label'):
                self.ihale_total_label.setText("Toplam (KDV HariÃ§): 0.00 â‚º")
                if hasattr(self, 'ihale_total_kdv_label'):
                    self.ihale_total_kdv_label.setText("Toplam (KDV Dahil): 0.00 â‚º")
            self.load_ihale_kalemleri()
        else:
            self.ihale_kalem_table.setRowCount(0)
            self.ihale_total_label.setText("Toplam: 0.00 â‚º")
    
    def new_ihale(self) -> None:
        """Yeni ihale oluÅŸtur"""
        from PyQt6.QtWidgets import QInputDialog
        
        ad, ok1 = QInputDialog.getText(self, "Yeni Ä°hale", "Ä°hale adÄ±:")
        if not ok1 or not ad.strip():
            return
        
        aciklama, ok2 = QInputDialog.getText(self, "Ä°hale AÃ§Ä±klamasÄ±", "AÃ§Ä±klama (isteÄŸe baÄŸlÄ±):")
        if not ok2:
            return
        
        ihale_id = self.db.create_ihale(ad.strip(), aciklama.strip())
        if ihale_id:
            self.load_ihaleler()
            # Yeni oluÅŸturulan ihale seÃ§ili olsun
            index = self.ihale_combo.findData(ihale_id)
            if index >= 0:
                self.ihale_combo.setCurrentIndex(index)
            QMessageBox.information(self, "BaÅŸarÄ±lÄ±", "Ä°hale oluÅŸturuldu")
            self.statusBar().showMessage(f"Ä°hale oluÅŸturuldu: {ad}")
    
    def on_ihale_poz_search(self) -> None:
        """Poz arama metni deÄŸiÅŸtiÄŸinde"""
        # Tablo widget'Ä± henÃ¼z oluÅŸturulmamÄ±ÅŸsa (lazy loading) iÅŸlem yapma
        if not hasattr(self, 'ihale_poz_results_table'):
            print("DEBUG: ihale_poz_results_table henÃ¼z oluÅŸturulmamÄ±ÅŸ")
            return
        
        if not self._tabs_created.get('ihale', False):
            print("DEBUG: Ä°hale sekmesi henÃ¼z oluÅŸturulmamÄ±ÅŸ")
            return
        
        search_text = self.ihale_poz_search.text().strip()
        
        # Minimum 1 karakter yeterli olsun (poz numarasÄ± tek karakter olabilir)
        if len(search_text) < 1:
            self.ihale_poz_results_table.setRowCount(0)
            return
        
        try:
            # Ã–nce pozlarÄ± ara
            print(f"DEBUG: Arama yapÄ±lÄ±yor: '{search_text}'")
            pozlar = self.db.search_pozlar(search_text, limit=50)
            print(f"DEBUG: {len(pozlar)} poz bulundu")
            
            # SonuÃ§larÄ± gÃ¶ster
            self.ihale_poz_results_table.setRowCount(len(pozlar))
            
            if len(pozlar) == 0:
                # SonuÃ§ yoksa kullanÄ±cÄ±ya bilgi ver ve manuel ekleme seÃ§eneÄŸi sun
                self.statusBar().showMessage(f"'{search_text}' iÃ§in poz bulunamadÄ±. Manuel eklemek iÃ§in 'Listeye Ekle' butonuna tÄ±klayÄ±n.", 5000)
                
                # EÄŸer arama metni poz numarasÄ± formatÄ±ndaysa (nokta iÃ§eriyorsa), 
                # manuel olarak eklenebilir ÅŸekilde tabloya tek satÄ±r ekle
                if '.' in search_text and len(search_text) > 3:
                    # Poz numarasÄ± formatÄ±nda gÃ¶rÃ¼nÃ¼yor, manuel ekleme iÃ§in gÃ¶ster
                    self.ihale_poz_results_table.setRowCount(1)
                    poz_no_item = QTableWidgetItem(search_text)
                    self.ihale_poz_results_table.setItem(0, 0, poz_no_item)
                    self.ihale_poz_results_table.setItem(0, 1, QTableWidgetItem("(Manuel ekleme - Poz bulunamadÄ±)"))
                    self.ihale_poz_results_table.setItem(0, 2, QTableWidgetItem(""))
                    self.ihale_poz_results_table.setItem(0, 3, QTableWidgetItem("Fiyat yok"))
                    
                    # Poz bilgisini sakla (sadece poz_no ile)
                    poz_data = {
                        'poz_no': search_text,
                        'tanim': '',
                        'birim': '',
                        'kategori': ''
                    }
                    poz_no_item.setData(Qt.ItemDataRole.UserRole, poz_data)
            else:
                self.statusBar().showMessage(f"{len(pozlar)} poz bulundu", 2000)
            
            for row, poz in enumerate(pozlar):
                poz_no = poz.get('poz_no', '')
                poz_tanim = poz.get('tanim', '')
                birim = poz.get('birim', '')
                kategori = poz.get('kategori', '')
                
                # Poz no
                poz_no_item = QTableWidgetItem(poz_no)
                self.ihale_poz_results_table.setItem(row, 0, poz_no_item)
                
                # TanÄ±m
                self.ihale_poz_results_table.setItem(row, 1, QTableWidgetItem(poz_tanim))
                
                # Birim
                self.ihale_poz_results_table.setItem(row, 2, QTableWidgetItem(birim))
                
                # Birim fiyatÄ± getir
                fiyat_data = self.db.get_birim_fiyat(poz_no=poz_no)
                birim_fiyat = fiyat_data.get('birim_fiyat', 0) if fiyat_data else 0
                self.ihale_poz_results_table.setItem(row, 3, QTableWidgetItem(f"{birim_fiyat:,.2f} â‚º" if birim_fiyat else "Fiyat yok"))
                
                # Poz bilgisini sakla (tÃ¼m bilgileri iÃ§eren dict)
                poz_data = {
                    'poz_no': poz_no,
                    'tanim': poz_tanim,
                    'birim': birim,
                    'kategori': kategori
                }
                poz_no_item.setData(Qt.ItemDataRole.UserRole, poz_data)
            
            # Tabloyu gÃ¼ncelle ve gÃ¶rÃ¼nÃ¼r yap
            self.ihale_poz_results_table.resizeColumnsToContents()
            self.ihale_poz_results_table.setVisible(True)
            self.ihale_poz_results_table.update()  # Tabloyu yeniden Ã§iz
            
        except Exception as e:
            error_msg = f"Poz arama sÄ±rasÄ±nda hata oluÅŸtu:\n{str(e)}"
            QMessageBox.critical(self, "Hata", error_msg)
            self.statusBar().showMessage(f"Hata: {str(e)}", 5000)
            import traceback
            traceback.print_exc()
    
    def add_selected_poz_to_ihale(self, item: QTableWidgetItem) -> None:
        """SeÃ§ili pozu ihale listesine ekle (Ã§ift tÄ±klama)"""
        if not self.current_ihale_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir ihale seÃ§in")
            return
        
        row = item.row()
        poz_item = self.ihale_poz_results_table.item(row, 0)
        if not poz_item:
            QMessageBox.warning(self, "UyarÄ±", "Poz bilgisi bulunamadÄ±")
            return
        
        poz_data = poz_item.data(Qt.ItemDataRole.UserRole)
        if not poz_data:
            # Poz data yoksa, tablodan manuel olarak al
            poz_no = poz_item.text()
            poz_tanim_item = self.ihale_poz_results_table.item(row, 1)
            poz_tanim = poz_tanim_item.text() if poz_tanim_item else ""
            birim_item = self.ihale_poz_results_table.item(row, 2)
            birim = birim_item.text() if birim_item else ""
            
            # Poz bilgilerini veritabanÄ±ndan getir
            poz = self.db.get_poz_by_no(poz_no)
            if not poz:
                QMessageBox.warning(self, "UyarÄ±", f"Poz bulunamadÄ±: {poz_no}")
                return
            
            poz_data = {
                'poz_no': poz_no,
                'tanim': poz_tanim or poz.get('tanim', ''),
                'kategori': poz.get('kategori', ''),
                'birim': birim or poz.get('birim', '')
            }
        
        try:
            self._add_poz_to_ihale_list(poz_data)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Poz eklenirken hata oluÅŸtu:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    def add_poz_to_ihale(self) -> None:
        """Arama sonuÃ§larÄ±ndan seÃ§ili pozu ekle"""
        if not self.current_ihale_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir ihale seÃ§in")
            return
        
        current_row = self.ihale_poz_results_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen bir poz seÃ§in")
            return
        
        poz_item = self.ihale_poz_results_table.item(current_row, 0)
        if not poz_item:
            QMessageBox.warning(self, "UyarÄ±", "Poz bilgisi bulunamadÄ±")
            return
        
        poz_data = poz_item.data(Qt.ItemDataRole.UserRole)
        if not poz_data:
            # Poz data yoksa, tablodan manuel olarak al
            poz_no = poz_item.text()
            poz_tanim_item = self.ihale_poz_results_table.item(current_row, 1)
            poz_tanim = poz_tanim_item.text() if poz_tanim_item else ""
            birim_item = self.ihale_poz_results_table.item(current_row, 2)
            birim = birim_item.text() if birim_item else ""
            
            # Poz bilgilerini veritabanÄ±ndan getir (yoksa manuel ekleme yapÄ±lacak)
            poz = self.db.get_poz_by_no(poz_no)
            if poz:
                # Poz bulundu, bilgileri kullan
                poz_data = {
                    'poz_no': poz_no,
                    'tanim': poz_tanim or poz.get('tanim', ''),
                    'kategori': poz.get('kategori', ''),
                    'birim': birim or poz.get('birim', '')
                }
            else:
                # Poz bulunamadÄ±, manuel ekleme iÃ§in sadece poz_no ile devam et
                poz_data = {
                    'poz_no': poz_no,
                    'tanim': poz_tanim,
                    'kategori': '',
                    'birim': birim
                }
        
        try:
            self._add_poz_to_ihale_list(poz_data)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Poz eklenirken hata oluÅŸtu:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    def _add_poz_to_ihale_list(self, poz_data: Dict[str, Any]) -> None:
        """Pozu ihale listesine ekle (iÃ§ fonksiyon)"""
        try:
            poz_no = poz_data.get('poz_no', '')
            if not poz_no:
                QMessageBox.warning(self, "UyarÄ±", "Poz numarasÄ± bulunamadÄ±")
                return
            
            poz_tanim = poz_data.get('tanim', '')
            kategori = poz_data.get('kategori', '')
            birim = poz_data.get('birim', '')
            
            # EÄŸer poz veritabanÄ±nda yoksa, veritabanÄ±ndan tekrar kontrol et
            if not poz_tanim or poz_tanim == "(Manuel ekleme - Poz bulunamadÄ±)":
                poz = self.db.get_poz_by_no(poz_no)
                if poz:
                    poz_tanim = poz.get('tanim', '')
                    birim = poz.get('birim', '') if not birim else birim
                    kategori = poz.get('kategori', '') if not kategori else kategori
                else:
                    # Poz veritabanÄ±nda yok, kullanÄ±cÄ±dan bilgi al
                    from PyQt6.QtWidgets import QInputDialog
                    tanim, ok = QInputDialog.getText(
                        self, "Poz Bilgisi",
                        f"Poz '{poz_no}' veritabanÄ±nda bulunamadÄ±.\n\nLÃ¼tfen poz tanÄ±mÄ±nÄ± girin:",
                        text=""
                    )
                    if not ok or not tanim.strip():
                        return
                    poz_tanim = tanim.strip()
                    
                    # Birim seÃ§imi
                    birim_text, ok = QInputDialog.getText(
                        self, "Birim",
                        "Birim (mÂ², mÂ³, kg, adet, vb.):",
                        text="mÂ²"
                    )
                    if not ok:
                        birim_text = "mÂ²"
                    birim = birim_text.strip() if birim_text.strip() else "mÂ²"
            
            # Birim fiyatÄ± getir (otomatik) - Ã¶nce aktif, sonra herhangi bir fiyat
            fiyat_data = self.db.get_birim_fiyat(poz_no=poz_no, aktif_only=True)
            if not fiyat_data or not fiyat_data.get('birim_fiyat'):
                # Aktif fiyat yoksa, aktif olmayan fiyatlarÄ± da kontrol et
                fiyat_data = self.db.get_birim_fiyat(poz_no=poz_no, aktif_only=False)
            birim_fiyat = fiyat_data.get('birim_fiyat', 0) if fiyat_data else 0
            
            # EÄŸer hala 0 ise, poz'un resmi_fiyat'Ä±nÄ± kontrol et
            if birim_fiyat == 0:
                poz_data = self.db.get_poz_by_no(poz_no)
                if poz_data and poz_data.get('resmi_fiyat'):
                    birim_fiyat = poz_data.get('resmi_fiyat', 0)
            
            # Ä°hale kalemine ekle (birim miktar 0, kullanÄ±cÄ± girecek)
            kalem_id = self.db.add_ihale_kalem(
                ihale_id=self.current_ihale_id,
                poz_no=poz_no,
                poz_tanim=poz_tanim,
                kategori=kategori,
                birim_miktar=0,  # KullanÄ±cÄ± girecek
                birim=birim,
                birim_fiyat=birim_fiyat,
                toplam=0
            )
            
            if kalem_id:
                # Tabloyu yeniden yÃ¼kleme - kullanÄ±cÄ±nÄ±n dÃ¼zenlemelerini kaybetmemek iÃ§in
                # Sadece yeni eklenen satÄ±rÄ± ekle, tÃ¼m tabloyu yeniden yÃ¼kleme
                self.statusBar().showMessage(f"Poz eklendi: {poz_no}")
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"Poz baÅŸarÄ±yla eklendi:\n{poz_no} - {poz_tanim}")
                # Tabloyu yeniden yÃ¼kle (sadece yeni ekleme sonrasÄ±)
                self.load_ihale_kalemleri()
            else:
                QMessageBox.warning(self, "UyarÄ±", "Poz eklenirken bir hata oluÅŸtu")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Poz eklenirken hata oluÅŸtu:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    def load_ihale_kalemleri(self) -> None:
        """Ä°hale kalemlerini yÃ¼kle"""
        # Lazy loading kontrolÃ¼ - sekme henÃ¼z oluÅŸturulmamÄ±ÅŸsa Ã§Ä±k
        if not hasattr(self, 'ihale_kalem_table') or not self._tabs_created.get('ihale', False):
            return
        
        if not self.current_ihale_id:
            try:
                self.ihale_kalem_table.setRowCount(0)
            except:
                pass
            return
        
        try:
            import re
            # itemChanged sinyalini blokla (tablo yÃ¼klenirken sinyal tetiklenmesin)
            self.ihale_kalem_table.blockSignals(True)
            try:
                kalemler = self.db.get_ihale_kalemleri(self.current_ihale_id)
                self.ihale_kalem_table.setRowCount(len(kalemler))
                
                toplam = 0.0
                
                for row, kalem in enumerate(kalemler):
                    # SÄ±ra (dÃ¼zenlenemez)
                    sira_item = QTableWidgetItem(str(kalem.get('sira_no', row + 1)))
                    sira_item.setFlags(sira_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.ihale_kalem_table.setItem(row, 0, sira_item)
                    
                    # Poz No (dÃ¼zenlenemez)
                    poz_no_item = QTableWidgetItem(kalem.get('poz_no', ''))
                    poz_no_item.setFlags(poz_no_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.ihale_kalem_table.setItem(row, 1, poz_no_item)
                    
                    # TanÄ±m (temizle - fiyat bilgisi varsa Ã§Ä±kar, dÃ¼zenlenebilir)
                    poz_tanim = str(kalem.get('poz_tanim', '')).strip()
                    # "Sa 250,00" veya "Sa 250.00" gibi pattern'leri temizle
                    poz_tanim = re.sub(r'\s*Sa\s*\d+[.,]\d+', '', poz_tanim).strip()
                    tanim_item = QTableWidgetItem(poz_tanim)
                    # TanÄ±m dÃ¼zenlenebilir yapÄ±ldÄ±
                    # Tam metni tooltip olarak ekle (tÃ¼m metin gÃ¶rÃ¼nsÃ¼n)
                    tanim_item.setToolTip(poz_tanim)
                    # Word wrap Ã¶zelliÄŸi iÃ§in hizalama
                    tanim_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
                    self.ihale_kalem_table.setItem(row, 2, tanim_item)
                    # SatÄ±r yÃ¼ksekliÄŸini otomatik ayarla (uzun metinler iÃ§in)
                    if len(poz_tanim) > 80:
                        # Uzun metinler iÃ§in satÄ±r yÃ¼ksekliÄŸini artÄ±r
                        self.ihale_kalem_table.setRowHeight(row, max(40, min(100, len(poz_tanim) // 40 * 20)))
                    # SatÄ±r yÃ¼ksekliÄŸini otomatik ayarla (uzun metinler iÃ§in)
                    if len(poz_tanim) > 80:
                        # Uzun metinler iÃ§in satÄ±r yÃ¼ksekliÄŸini artÄ±r
                        self.ihale_kalem_table.setRowHeight(row, max(40, min(100, len(poz_tanim) // 40 * 20)))
                    
                    # Birim Miktar (dÃ¼zenlenebilir) - 0 ise boÅŸ gÃ¶ster
                    birim_miktar = kalem.get('birim_miktar', 0) or 0
                    # EÄŸer birim_miktar None veya 0 ise, tablodan oku (kullanÄ±cÄ± yazmÄ±ÅŸ olabilir)
                    if birim_miktar == 0:
                        # Tablodan oku (eÄŸer kullanÄ±cÄ± yazdÄ±ysa)
                        existing_item = self.ihale_kalem_table.item(row, 3)
                        if existing_item and existing_item.text().strip():
                            try:
                                miktar_text_existing = existing_item.text().replace(",", ".").strip()
                                birim_miktar = float(miktar_text_existing) if miktar_text_existing else 0.0
                            except:
                                pass
                    miktar_text = f"{birim_miktar:,.2f}" if birim_miktar > 0 else ""
                    miktar_item = QTableWidgetItem(miktar_text)
                    miktar_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    # Font'u bÃ¼yÃ¼t ve kalÄ±n yap
                    font = miktar_item.font()
                    font.setPointSize(font.pointSize() + 2)  # 2 punto bÃ¼yÃ¼t
                    font.setBold(True)  # KalÄ±n yap
                    miktar_item.setFont(font)
                    self.ihale_kalem_table.setItem(row, 3, miktar_item)
                    
                    # Birim (dÃ¼zenlenebilir)
                    birim_item = QTableWidgetItem(kalem.get('birim', ''))
                    birim_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                    self.ihale_kalem_table.setItem(row, 4, birim_item)
                    
                    # Birim Fiyat (dÃ¼zenlenebilir)
                    birim_fiyat = kalem.get('birim_fiyat', 0) or 0
                    # EÄŸer ihale_kalemleri tablosunda birim_fiyat 0 ise, birim_fiyatlar tablosundan al
                    if birim_fiyat == 0:
                        poz_no = kalem.get('poz_no', '')
                        if poz_no:
                            fiyat_data = self.db.get_birim_fiyat(poz_no=poz_no, aktif_only=False)
                            if fiyat_data and fiyat_data.get('birim_fiyat'):
                                birim_fiyat = fiyat_data.get('birim_fiyat', 0)
                                # Ä°hale kalemindeki birim fiyatÄ± gÃ¼ncelle
                                kalem_id = kalem.get('id')
                                if kalem_id:
                                    self.db.update_ihale_kalem(kalem_id, birim_fiyat=birim_fiyat)
                            else:
                                # Poz'un resmi_fiyat'Ä±nÄ± kontrol et
                                poz_data = self.db.get_poz_by_no(poz_no)
                                if poz_data and poz_data.get('resmi_fiyat'):
                                    birim_fiyat = poz_data.get('resmi_fiyat', 0)
                                    kalem_id = kalem.get('id')
                                    if kalem_id:
                                        self.db.update_ihale_kalem(kalem_id, birim_fiyat=birim_fiyat)
                    
                    fiyat_item = QTableWidgetItem(f"{birim_fiyat:,.2f}")
                    fiyat_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    self.ihale_kalem_table.setItem(row, 5, fiyat_item)
                    
                    # Toplam (hesaplanÄ±r, dÃ¼zenlenemez) - HER ZAMAN birim miktar ve birim fiyattan hesapla
                    # Ã–NEMLÄ°: Tabloda gÃ¶rÃ¼nen deÄŸerleri kullan (kullanÄ±cÄ± yazmÄ±ÅŸ olabilir)
                    # Tablodan birim miktar ve birim fiyatÄ± oku
                    miktar_item_tablo = self.ihale_kalem_table.item(row, 3)
                    fiyat_item_tablo = self.ihale_kalem_table.item(row, 5)
                    
                    birim_miktar_hesap = birim_miktar
                    birim_fiyat_hesap = birim_fiyat
                    
                    # EÄŸer tabloda deÄŸerler varsa onlarÄ± kullan
                    if miktar_item_tablo and miktar_item_tablo.text().strip():
                        try:
                            miktar_text_tablo = miktar_item_tablo.text().strip()
                            # TÃ¼rkÃ§e ve Ä°ngilizce format desteÄŸi
                            miktar_text_tablo = miktar_text_tablo.replace(" ", "")
                            if ',' in miktar_text_tablo and '.' in miktar_text_tablo:
                                last_dot = miktar_text_tablo.rfind('.')
                                last_comma = miktar_text_tablo.rfind(',')
                                if last_dot > last_comma:
                                    birim_miktar_hesap = float(miktar_text_tablo.replace(',', ''))
                                else:
                                    birim_miktar_hesap = float(miktar_text_tablo.replace('.', '').replace(',', '.'))
                            elif ',' in miktar_text_tablo:
                                birim_miktar_hesap = float(miktar_text_tablo.replace(',', '.'))
                            else:
                                birim_miktar_hesap = float(miktar_text_tablo.replace(',', '.'))
                        except:
                            pass
                    
                    if fiyat_item_tablo and fiyat_item_tablo.text().strip():
                        try:
                            fiyat_text_tablo = fiyat_item_tablo.text().replace("â‚º", "").strip()
                            # TÃ¼rkÃ§e ve Ä°ngilizce format desteÄŸi
                            fiyat_text_tablo = fiyat_text_tablo.replace(" ", "")
                            if ',' in fiyat_text_tablo and '.' in fiyat_text_tablo:
                                last_dot = fiyat_text_tablo.rfind('.')
                                last_comma = fiyat_text_tablo.rfind(',')
                                if last_dot > last_comma:
                                    birim_fiyat_hesap = float(fiyat_text_tablo.replace(',', ''))
                                else:
                                    birim_fiyat_hesap = float(fiyat_text_tablo.replace('.', '').replace(',', '.'))
                            elif ',' in fiyat_text_tablo:
                                birim_fiyat_hesap = float(fiyat_text_tablo.replace(',', '.'))
                            else:
                                birim_fiyat_hesap = float(fiyat_text_tablo.replace(',', '.'))
                        except:
                            pass
                    
                    # ToplamÄ± hesapla
                    toplam_deger = birim_miktar_hesap * birim_fiyat_hesap
                    
                    # VeritabanÄ±nÄ± gÃ¼ncelle
                    kalem_id = kalem.get('id')
                    if kalem_id:
                        # VeritabanÄ±ndaki toplam ile hesaplanan toplam farklÄ±ysa gÃ¼ncelle
                        db_toplam = kalem.get('toplam', 0) or 0
                        if abs(db_toplam - toplam_deger) > 0.01:
                            self.db.update_ihale_kalem(kalem_id, birim_miktar=birim_miktar_hesap, birim_fiyat=birim_fiyat_hesap, toplam=toplam_deger)
                    
                    toplam += toplam_deger
                    toplam_item = QTableWidgetItem(f"{toplam_deger:,.2f} â‚º")
                    toplam_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    toplam_item.setFlags(toplam_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.ihale_kalem_table.setItem(row, 6, toplam_item)
                    
                    # ID'yi sakla
                    item = self.ihale_kalem_table.item(row, 0)
                    if item:
                        item.setData(Qt.ItemDataRole.UserRole, kalem.get('id'))
                
                if hasattr(self, 'ihale_total_label'):
                    # KDV hesaplama
                    kdv_rate_text = self.ihale_kdv_rate.currentText().replace("%", "")
                    kdv_rate = float(kdv_rate_text)
                    kdv_hesap = self.calculator.calculate_with_kdv(toplam, kdv_rate)
                    
                    self.ihale_total_label.setText(f"Toplam (KDV HariÃ§): {toplam:,.2f} â‚º")
                    self.ihale_total_kdv_label.setText(f"Toplam (KDV %{kdv_rate_text} Dahil): {kdv_hesap['kdv_dahil']:,.2f} â‚º")
            finally:
                # Sinyali tekrar aÃ§
                self.ihale_kalem_table.blockSignals(False)
        except Exception as e:
            print(f"Ä°hale kalemleri yÃ¼kleme hatasÄ±: {e}")
            import traceback
            traceback.print_exc()
            # Hata olsa bile tabloyu temizle
            try:
                self.ihale_kalem_table.blockSignals(True)
                self.ihale_kalem_table.setRowCount(0)
                self.ihale_kalem_table.blockSignals(False)
            except:
                pass
    
    def show_full_tanim(self, item: QTableWidgetItem) -> None:
        """TanÄ±m sÃ¼tununa Ã§ift tÄ±klayÄ±nca tam metni gÃ¶ster"""
        # Sadece tanÄ±m sÃ¼tunu (2) iÃ§in iÅŸlem yap
        if item.column() != 2:
            return
        
        tanim_text = item.text()
        if not tanim_text:
            return
        
        # Dialog oluÅŸtur
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QTextEdit, QPushButton
        dialog = QDialog(self)
        dialog.setWindowTitle("Tam TanÄ±m")
        dialog.setMinimumWidth(700)
        dialog.setMinimumHeight(400)
        
        layout = QVBoxLayout(dialog)
        
        # Metin alanÄ±
        text_edit = QTextEdit()
        text_edit.setPlainText(tanim_text)
        text_edit.setReadOnly(True)
        text_edit.setFont(QFont("Arial", 10))
        layout.addWidget(text_edit)
        
        # Kapat butonu
        btn_close = QPushButton("Kapat")
        btn_close.clicked.connect(dialog.close)
        layout.addWidget(btn_close)
        
        dialog.exec()
    
    def on_ihale_kalem_changed(self, item: QTableWidgetItem) -> None:
        """Ä°hale kalemi deÄŸiÅŸtiÄŸinde (tanÄ±m, birim miktar, birim veya birim fiyat)"""
        # DÃ¼zenlenebilir sÃ¼tunlar (2: TanÄ±m, 3: Birim Miktar, 4: Birim, 5: Birim Fiyat)
        if item.column() not in [2, 3, 4, 5]:
            return
        
        # TanÄ±m deÄŸiÅŸtiyse sadece tanÄ±mÄ± gÃ¼ncelle
        if item.column() == 2:
            row = item.row()
            kalem_id_item = self.ihale_kalem_table.item(row, 0)
            if not kalem_id_item:
                return
            
            kalem_id = kalem_id_item.data(Qt.ItemDataRole.UserRole)
            if not kalem_id:
                return
            
            yeni_tanim = item.text().strip()
            if yeni_tanim:
                self.db.update_ihale_kalem(kalem_id, poz_tanim=yeni_tanim)
                # Tooltip'i de gÃ¼ncelle
                item.setToolTip(yeni_tanim)
                # SatÄ±r yÃ¼ksekliÄŸini gÃ¼ncelle
                if len(yeni_tanim) > 80:
                    self.ihale_kalem_table.setRowHeight(row, max(40, min(100, len(yeni_tanim) // 40 * 20)))
            return
        
        row = item.row()
        kalem_id_item = self.ihale_kalem_table.item(row, 0)
        if not kalem_id_item:
            return
        
        kalem_id = kalem_id_item.data(Qt.ItemDataRole.UserRole)
        if not kalem_id:
            return
        
        # Birim miktar, birim ve birim fiyatÄ± al
        miktar_item = self.ihale_kalem_table.item(row, 3)
        birim_item = self.ihale_kalem_table.item(row, 4)
        fiyat_item = self.ihale_kalem_table.item(row, 5)
        
        if not miktar_item or not birim_item or not fiyat_item:
            return
        
        # Birim miktar iÃ§in font ayarlarÄ±nÄ± koru (bÃ¼yÃ¼k ve kalÄ±n)
        if item.column() == 3:
            # Mevcut font'u al ve ayarlarÄ± koru
            font = miktar_item.font()
            if not font.bold() or font.pointSize() <= 10:
                font.setPointSize(font.pointSize() + 2)
                font.setBold(True)
                miktar_item.setFont(font)
        
        try:
            miktar_text = miktar_item.text().strip()
            birim_text = birim_item.text().strip()
            fiyat_text = fiyat_item.text().replace("â‚º", "").strip()
            
            # Birim miktar parse - TÃ¼rkÃ§e ve Ä°ngilizce format desteÄŸi
            birim_miktar = 0.0
            if miktar_text:
                try:
                    # Ã–nce boÅŸluklarÄ± temizle
                    miktar_text = miktar_text.replace(" ", "")
                    # EÄŸer hem virgÃ¼l hem nokta varsa
                    if ',' in miktar_text and '.' in miktar_text:
                        # Son noktadan Ã¶nceki kÄ±smÄ± kontrol et
                        last_dot = miktar_text.rfind('.')
                        last_comma = miktar_text.rfind(',')
                        if last_dot > last_comma:
                            # Nokta ondalÄ±k ayÄ±rÄ±cÄ± (Ä°ngilizce format: 1,234.56)
                            # VirgÃ¼lleri kaldÄ±r, noktayÄ± koru
                            birim_miktar = float(miktar_text.replace(',', ''))
                        else:
                            # VirgÃ¼l ondalÄ±k ayÄ±rÄ±cÄ± (TÃ¼rkÃ§e format: 1.234,56)
                            # NoktalarÄ± kaldÄ±r, virgÃ¼lÃ¼ noktaya Ã§evir
                            birim_miktar = float(miktar_text.replace('.', '').replace(',', '.'))
                    elif ',' in miktar_text:
                        # Sadece virgÃ¼l var - TÃ¼rkÃ§e format (ondalÄ±k ayÄ±rÄ±cÄ±)
                        birim_miktar = float(miktar_text.replace(',', '.'))
                    elif '.' in miktar_text:
                        # Sadece nokta var - kontrol et
                        # EÄŸer birden fazla nokta varsa, son nokta ondalÄ±k, diÄŸerleri binlik
                        dot_count = miktar_text.count('.')
                        if dot_count > 1:
                            # Son noktadan Ã¶nceki noktalarÄ± kaldÄ±r
                            last_dot = miktar_text.rfind('.')
                            before_last = miktar_text[:last_dot].replace('.', '')
                            after_last = miktar_text[last_dot:]
                            birim_miktar = float(before_last + after_last)
                        else:
                            # Tek nokta - ondalÄ±k ayÄ±rÄ±cÄ±
                            birim_miktar = float(miktar_text)
                    else:
                        # Sadece sayÄ±
                        birim_miktar = float(miktar_text)
                except (ValueError, AttributeError) as e:
                    print(f"Birim miktar parse hatasÄ±: {miktar_text} -> {e}")
                    birim_miktar = 0.0
            
            birim = birim_text if birim_text else ""
            
            # Birim fiyat parse - TÃ¼rkÃ§e ve Ä°ngilizce format desteÄŸi
            birim_fiyat = 0.0
            if fiyat_text:
                try:
                    # Ã–nce boÅŸluklarÄ± temizle
                    fiyat_text = fiyat_text.replace(" ", "")
                    # EÄŸer hem virgÃ¼l hem nokta varsa
                    if ',' in fiyat_text and '.' in fiyat_text:
                        # Son noktadan Ã¶nceki kÄ±smÄ± kontrol et
                        last_dot = fiyat_text.rfind('.')
                        last_comma = fiyat_text.rfind(',')
                        if last_dot > last_comma:
                            # Nokta ondalÄ±k ayÄ±rÄ±cÄ± (Ä°ngilizce format: 19,100.00)
                            # VirgÃ¼lleri kaldÄ±r, noktayÄ± koru
                            birim_fiyat = float(fiyat_text.replace(',', ''))
                        else:
                            # VirgÃ¼l ondalÄ±k ayÄ±rÄ±cÄ± (TÃ¼rkÃ§e format: 19.100,00)
                            # NoktalarÄ± kaldÄ±r, virgÃ¼lÃ¼ noktaya Ã§evir
                            birim_fiyat = float(fiyat_text.replace('.', '').replace(',', '.'))
                    elif ',' in fiyat_text:
                        # Sadece virgÃ¼l var - TÃ¼rkÃ§e format (ondalÄ±k ayÄ±rÄ±cÄ±)
                        birim_fiyat = float(fiyat_text.replace(',', '.'))
                    elif '.' in fiyat_text:
                        # Sadece nokta var - kontrol et
                        # EÄŸer birden fazla nokta varsa, son nokta ondalÄ±k, diÄŸerleri binlik
                        dot_count = fiyat_text.count('.')
                        if dot_count > 1:
                            # Son noktadan Ã¶nceki noktalarÄ± kaldÄ±r
                            last_dot = fiyat_text.rfind('.')
                            before_last = fiyat_text[:last_dot].replace('.', '')
                            after_last = fiyat_text[last_dot:]
                            birim_fiyat = float(before_last + after_last)
                        else:
                            # Tek nokta - ondalÄ±k ayÄ±rÄ±cÄ±
                            birim_fiyat = float(fiyat_text)
                    else:
                        # Sadece sayÄ±
                        birim_fiyat = float(fiyat_text)
                except (ValueError, AttributeError) as e:
                    print(f"Birim fiyat parse hatasÄ±: {fiyat_text} -> {e}")
                    birim_fiyat = 0.0
            
            # Toplam hesapla
            toplam = birim_miktar * birim_fiyat
            
            # VeritabanÄ±nÄ± gÃ¼ncelle
            success = self.db.update_ihale_kalem(kalem_id, birim_miktar=birim_miktar, birim=birim, birim_fiyat=birim_fiyat, toplam=toplam)
            
            if success:
                # itemChanged sinyalini blokla (sadece toplam sÃ¼tununu gÃ¼ncellerken)
                self.ihale_kalem_table.blockSignals(True)
                try:
                    # KullanÄ±cÄ±nÄ±n yazdÄ±ÄŸÄ± deÄŸerleri KORU - hiÃ§bir ÅŸey yapma
                    # Birim miktar, birim ve birim fiyat sÃ¼tunlarÄ± kullanÄ±cÄ±nÄ±n yazdÄ±ÄŸÄ± gibi kalacak
                    
                    # Sadece toplam sÃ¼tununu gÃ¼ncelle
                    toplam_item = QTableWidgetItem(f"{toplam:,.2f} â‚º")
                    toplam_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    toplam_item.setFlags(toplam_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.ihale_kalem_table.setItem(row, 6, toplam_item)
                    
                    # Genel toplamÄ± gÃ¼ncelle
                    self.update_ihale_total()
                finally:
                    # Sinyali tekrar aÃ§
                    self.ihale_kalem_table.blockSignals(False)
            else:
                QMessageBox.warning(self, "Hata", "VeritabanÄ± gÃ¼ncellemesi baÅŸarÄ±sÄ±z oldu")
            
        except ValueError:
            QMessageBox.warning(self, "Hata", "GeÃ§ersiz sayÄ± formatÄ±")
        except Exception as e:
            print(f"Ä°hale kalemi gÃ¼ncelleme hatasÄ±: {e}")
            import traceback
            traceback.print_exc()
    
    def update_ihale_total(self) -> None:
        """Ä°hale toplamÄ±nÄ± gÃ¼ncelle"""
        if not hasattr(self, 'ihale_kalem_table') or not hasattr(self, 'ihale_total_label'):
            return
        
        try:
            toplam = 0.0
            for row in range(self.ihale_kalem_table.rowCount()):
                # Toplam sÃ¼tunundan oku (6. sÃ¼tun)
                toplam_item = self.ihale_kalem_table.item(row, 6)
                if toplam_item:
                    toplam_text = toplam_item.text().replace("â‚º", "").strip()
                    try:
                        # TÃ¼rkÃ§e ve Ä°ngilizce format desteÄŸi
                        toplam_text = toplam_text.replace(" ", "")
                        if ',' in toplam_text and '.' in toplam_text:
                            # Son noktadan Ã¶nceki kÄ±smÄ± kontrol et
                            last_dot = toplam_text.rfind('.')
                            last_comma = toplam_text.rfind(',')
                            if last_dot > last_comma:
                                # Nokta ondalÄ±k ayÄ±rÄ±cÄ± (Ä°ngilizce format: 19,100.00)
                                toplam += float(toplam_text.replace(',', ''))
                            else:
                                # VirgÃ¼l ondalÄ±k ayÄ±rÄ±cÄ± (TÃ¼rkÃ§e format: 19.100,00)
                                toplam += float(toplam_text.replace('.', '').replace(',', '.'))
                        elif ',' in toplam_text:
                            # Sadece virgÃ¼l var - TÃ¼rkÃ§e format (ondalÄ±k ayÄ±rÄ±cÄ±)
                            toplam += float(toplam_text.replace(',', '.'))
                        elif '.' in toplam_text:
                            # Sadece nokta var - kontrol et
                            dot_count = toplam_text.count('.')
                            if dot_count > 1:
                                # Son noktadan Ã¶nceki noktalarÄ± kaldÄ±r
                                last_dot = toplam_text.rfind('.')
                                before_last = toplam_text[:last_dot].replace('.', '')
                                after_last = toplam_text[last_dot:]
                                toplam += float(before_last + after_last)
                            else:
                                # Tek nokta - ondalÄ±k ayÄ±rÄ±cÄ±
                                toplam += float(toplam_text)
                        else:
                            # Sadece sayÄ±
                            toplam += float(toplam_text)
                    except (ValueError, AttributeError) as e:
                        print(f"Toplam parse hatasÄ± (satÄ±r {row}): {toplam_text} -> {e}")
                        # Alternatif: Birim miktar ve birim fiyattan hesapla
                        try:
                            miktar_item = self.ihale_kalem_table.item(row, 3)
                            fiyat_item = self.ihale_kalem_table.item(row, 5)
                            if miktar_item and fiyat_item:
                                miktar_text = miktar_item.text().strip()
                                fiyat_text = fiyat_item.text().replace("â‚º", "").strip()
                                
                                # Miktar parse
                                miktar_val = 0.0
                                if miktar_text:
                                    miktar_text = miktar_text.replace(" ", "")
                                    if ',' in miktar_text and '.' in miktar_text:
                                        last_dot = miktar_text.rfind('.')
                                        last_comma = miktar_text.rfind(',')
                                        if last_dot > last_comma:
                                            miktar_val = float(miktar_text.replace(',', ''))
                                        else:
                                            miktar_val = float(miktar_text.replace('.', '').replace(',', '.'))
                                    elif ',' in miktar_text:
                                        miktar_val = float(miktar_text.replace(',', '.'))
                                    else:
                                        miktar_val = float(miktar_text.replace(',', '.'))
                                
                                # Fiyat parse
                                fiyat_val = 0.0
                                if fiyat_text:
                                    fiyat_text = fiyat_text.replace(" ", "")
                                    if ',' in fiyat_text and '.' in fiyat_text:
                                        last_dot = fiyat_text.rfind('.')
                                        last_comma = fiyat_text.rfind(',')
                                        if last_dot > last_comma:
                                            fiyat_val = float(fiyat_text.replace(',', ''))
                                        else:
                                            fiyat_val = float(fiyat_text.replace('.', '').replace(',', '.'))
                                    elif ',' in fiyat_text:
                                        fiyat_val = float(fiyat_text.replace(',', '.'))
                                    else:
                                        fiyat_val = float(fiyat_text.replace(',', '.'))
                                
                                # Ã‡arp ve ekle
                                toplam += miktar_val * fiyat_val
                        except:
                            pass
            
            # KDV hesaplama
            kdv_rate_text = self.ihale_kdv_rate.currentText().replace("%", "")
            kdv_rate = float(kdv_rate_text)
            kdv_hesap = self.calculator.calculate_with_kdv(toplam, kdv_rate)
            
            self.ihale_total_label.setText(f"Toplam (KDV HariÃ§): {toplam:,.2f} â‚º")
            self.ihale_total_kdv_label.setText(f"Toplam (KDV %{kdv_rate_text} Dahil): {kdv_hesap['kdv_dahil']:,.2f} â‚º")
        except Exception as e:
            print(f"Ä°hale toplam gÃ¼ncelleme hatasÄ±: {e}")
            import traceback
            traceback.print_exc()
    
    def edit_ihale_tanim(self) -> None:
        """SeÃ§ili kalemin tanÄ±mÄ±nÄ± dÃ¼zelt"""
        current_row = self.ihale_kalem_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen dÃ¼zeltmek istediÄŸiniz bir satÄ±rÄ± seÃ§in")
            return
        
        kalem_id_item = self.ihale_kalem_table.item(current_row, 0)
        if not kalem_id_item:
            return
        
        kalem_id = kalem_id_item.data(Qt.ItemDataRole.UserRole)
        if not kalem_id:
            return
        
        # Mevcut tanÄ±mÄ± al
        tanim_item = self.ihale_kalem_table.item(current_row, 2)
        mevcut_tanim = tanim_item.text() if tanim_item else ""
        poz_no_item = self.ihale_kalem_table.item(current_row, 1)
        poz_no = poz_no_item.text() if poz_no_item else ""
        
        # Yeni tanÄ±m gir
        from PyQt6.QtWidgets import QInputDialog
        yeni_tanim, ok = QInputDialog.getMultiLineText(
            self,
            "TanÄ±m DÃ¼zelt",
            f"Poz {poz_no} iÃ§in yeni tanÄ±mÄ± girin:",
            mevcut_tanim
        )
        
        if ok and yeni_tanim.strip():
            # VeritabanÄ±nÄ± gÃ¼ncelle
            success = self.db.update_ihale_kalem(kalem_id, poz_tanim=yeni_tanim.strip())
            if success:
                # Tabloyu gÃ¼ncelle
                tanim_item.setText(yeni_tanim.strip())
                tanim_item.setToolTip(yeni_tanim.strip())
                # SatÄ±r yÃ¼ksekliÄŸini gÃ¼ncelle
                if len(yeni_tanim.strip()) > 80:
                    self.ihale_kalem_table.setRowHeight(current_row, max(40, min(100, len(yeni_tanim.strip()) // 40 * 20)))
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"Poz {poz_no} iÃ§in tanÄ±m gÃ¼ncellendi")
            else:
                QMessageBox.warning(self, "Hata", "TanÄ±m gÃ¼ncellenirken bir hata oluÅŸtu")
    
    def delete_ihale_kalem(self) -> None:
        """Ä°hale kalemini sil"""
        if not self.current_ihale_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir ihale seÃ§in")
            return
        
        current_row = self.ihale_kalem_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen silmek istediÄŸiniz kalemi seÃ§in")
            return
        
        kalem_id_item = self.ihale_kalem_table.item(current_row, 0)
        if not kalem_id_item:
            return
        
        kalem_id = kalem_id_item.data(Qt.ItemDataRole.UserRole)
        poz_no = self.ihale_kalem_table.item(current_row, 1).text()
        
        reply = QMessageBox.question(
            self, "Kalem Sil",
            f"'{poz_no}' kalemini silmek istediÄŸinize emin misiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            if self.db.delete_ihale_kalem(kalem_id):
                self.load_ihale_kalemleri()
                self.statusBar().showMessage("Kalem silindi")
    
    def export_ihale_pdf(self) -> None:
        """Ä°hale dosyasÄ±nÄ± PDF olarak export et"""
        if not self.current_ihale_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir ihale seÃ§in")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "PDF Ä°hale DosyasÄ± OluÅŸtur", "", "PDF DosyalarÄ± (*.pdf)"
        )
        
        if file_path:
            try:
                ihale = self.db.get_ihale(self.current_ihale_id)
                kalemler = self.db.get_ihale_kalemleri(self.current_ihale_id)
                
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import cm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                
                # TÃ¼rkÃ§e karakter desteÄŸi iÃ§in font yÃ¼kle (eÄŸer yoksa varsayÄ±lan font kullanÄ±lÄ±r)
                font_name = 'Helvetica'
                font_bold_name = 'Helvetica-Bold'
                try:
                    # Windows'ta Arial fontunu kullan
                    import platform
                    if platform.system() == 'Windows':
                        arial_path = 'C:/Windows/Fonts/arial.ttf'
                        arial_bold_path = 'C:/Windows/Fonts/arialbd.ttf'
                        if Path(arial_path).exists():
                            pdfmetrics.registerFont(TTFont('Arial', arial_path))
                            font_name = 'Arial'
                            # Arial Bold fontunu da yÃ¼kle
                            if Path(arial_bold_path).exists():
                                try:
                                    pdfmetrics.registerFont(TTFont('Arial-Bold', arial_bold_path))
                                    font_bold_name = 'Arial-Bold'
                                except:
                                    # Arial-Bold yÃ¼klenemezse Helvetica-Bold kullan
                                    font_bold_name = 'Helvetica-Bold'
                            else:
                                # Arial Bold yoksa, Helvetica-Bold kullan
                                font_bold_name = 'Helvetica-Bold'
                        else:
                            font_name = 'Helvetica'
                            font_bold_name = 'Helvetica-Bold'
                    else:
                        font_name = 'Helvetica'
                        font_bold_name = 'Helvetica-Bold'
                except Exception as e:
                    print(f"Font yÃ¼kleme hatasÄ±: {e}")
                    font_name = 'Helvetica'
                    font_bold_name = 'Helvetica-Bold'
                
                # Logo yolu kontrolÃ¼
                logo_path = Path(__file__).parent.parent.parent / "assets" / "logo.png"
                has_logo = logo_path.exists()
                
                doc = SimpleDocTemplate(str(file_path), pagesize=A4)
                story = []
                styles = getSampleStyleSheet()
                
                # Logo ekle
                if has_logo:
                    try:
                        from reportlab.platypus import Image
                        from reportlab.lib.units import inch
                        logo = Image(str(logo_path), width=2*inch, height=0.8*inch)
                        logo.hAlign = 'CENTER'
                        story.append(logo)
                        story.append(Spacer(1, 0.2*inch))
                    except Exception as e:
                        print(f"Logo yÃ¼kleme hatasÄ±: {e}")
                
                # BaÅŸlÄ±k
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontName=font_name,
                    fontSize=18,
                    textColor=colors.HexColor('#1a1a2e'),
                    spaceAfter=30,
                    alignment=1
                )
                story.append(Paragraph(f"Ä°HALE DOSYASI - {ihale.get('ad', '')}", title_style))
                story.append(Spacer(1, 0.5*cm))
                
                # Ä°hale bilgileri
                info_data = [
                    ['Ä°hale AdÄ±', ihale.get('ad', '')],
                    ['AÃ§Ä±klama', ihale.get('aciklama', '')],
                    ['OluÅŸturulma Tarihi', ihale.get('olusturma_tarihi', '')[:10] if ihale.get('olusturma_tarihi') else ''],
                ]
                
                info_table = Table(info_data, colWidths=[6*cm, 6*cm])
                info_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#16213e')),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
                    ('BACKGROUND', (1, 0), (1, -1), colors.white),
                    ('TEXTCOLOR', (1, 0), (1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), font_bold_name),
                    ('FONTNAME', (1, 0), (1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ]))
                story.append(info_table)
                story.append(Spacer(1, 0.5*cm))
                
                # Kalem listesi
                if kalemler:
                    heading2_style = ParagraphStyle(
                        'CustomHeading2',
                        parent=styles['Heading2'],
                        fontName=font_name,
                        fontSize=14,
                        textColor=colors.HexColor('#1a1a2e'),
                        spaceAfter=20
                    )
                    story.append(Paragraph("Ä°hale Kalem Listesi", heading2_style))
                    kalem_data = [['SÄ±ra', 'Poz No', 'TanÄ±m', 'Miktar', 'Birim', 'Birim Fiyat', 'Toplam']]
                    
                    toplam_genel = 0.0
                    for kalem in kalemler:
                        toplam_genel += kalem.get('toplam', 0)
                        kalem_data.append([
                            str(kalem.get('sira_no', '')),
                            kalem.get('poz_no', ''),
                            kalem.get('poz_tanim', '')[:40],
                            f"{kalem.get('birim_miktar', 0):,.2f}",
                            kalem.get('birim', ''),
                            f"{kalem.get('birim_fiyat', 0):,.2f} TL",
                            f"{kalem.get('toplam', 0):,.2f} TL"
                        ])
                    
                    kalem_table = Table(kalem_data, colWidths=[1*cm, 2*cm, 5*cm, 2*cm, 1.5*cm, 2*cm, 2*cm])
                    kalem_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16213e')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), f'{font_name}-Bold' if font_name == 'Arial' else 'Helvetica-Bold'),
                        ('FONTNAME', (0, 1), (-1, -1), font_name),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('FONTSIZE', (0, 1), (-1, -1), 9),
                        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
                    ]))
                    story.append(kalem_table)
                    story.append(Spacer(1, 0.5*cm))
                    
                    # Toplam
                    toplam_data = [['GENEL TOPLAM', f"{toplam_genel:,.2f} TL"]]
                    toplam_table = Table(toplam_data, colWidths=[10*cm, 4*cm])
                    toplam_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#16213e')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ]))
                    story.append(toplam_table)
                
                doc.build(story)
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"Ä°hale dosyasÄ± PDF'e aktarÄ±ldÄ±:\n{file_path}")
                self.statusBar().showMessage(f"PDF ihale dosyasÄ± oluÅŸturuldu: {file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"PDF oluÅŸturulurken hata oluÅŸtu:\n{str(e)}")
    
    def export_ihale_excel(self) -> None:
        """Ä°hale dosyasÄ±nÄ± Excel olarak export et"""
        if not self.current_ihale_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir ihale seÃ§in")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Excel Ä°hale DosyasÄ± OluÅŸtur", "", "Excel DosyalarÄ± (*.xlsx)"
        )
        
        if file_path:
            try:
                import pandas as pd
                from openpyxl import load_workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                
                ihale = self.db.get_ihale(self.current_ihale_id)
                kalemler = self.db.get_ihale_kalemleri(self.current_ihale_id)
                
                # Excel writer
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # Ä°hale bilgileri
                    info_data = {
                        'Bilgi': ['Ä°hale AdÄ±', 'AÃ§Ä±klama', 'OluÅŸturulma Tarihi'],
                        'DeÄŸer': [
                            ihale.get('ad', ''),
                            ihale.get('aciklama', ''),
                            ihale.get('olusturma_tarihi', '')[:10] if ihale.get('olusturma_tarihi') else ''
                        ]
                    }
                    df_info = pd.DataFrame(info_data)
                    df_info.to_excel(writer, sheet_name='Ä°hale Bilgileri', index=False)
                    
                    # Kalem listesi
                    if kalemler:
                        kalem_data = {
                            'SÄ±ra': [k.get('sira_no', '') for k in kalemler],
                            'Poz No': [k.get('poz_no', '') for k in kalemler],
                            'TanÄ±m': [k.get('poz_tanim', '') for k in kalemler],
                            'Birim Miktar': [k.get('birim_miktar', 0) for k in kalemler],
                            'Birim': [k.get('birim', '') for k in kalemler],
                            'Birim Fiyat': [f"{k.get('birim_fiyat', 0):,.2f} TL" for k in kalemler],
                            'Toplam': [f"{k.get('toplam', 0):,.2f} TL" for k in kalemler]
                        }
                        df_kalem = pd.DataFrame(kalem_data)
                        df_kalem.to_excel(writer, sheet_name='Kalem Listesi', index=False)
                        
                        # Toplam satÄ±rÄ±
                        toplam_genel = sum(k.get('toplam', 0) for k in kalemler)
                        toplam_row = pd.DataFrame({
                            'SÄ±ra': [''],
                            'Poz No': [''],
                            'TanÄ±m': ['GENEL TOPLAM'],
                            'Birim Miktar': [''],
                            'Birim': [''],
                            'Birim Fiyat': [''],
                            'Toplam': [f"{toplam_genel:,.2f} TL"]
                        })
                        df_kalem = pd.concat([df_kalem, toplam_row], ignore_index=True)
                        df_kalem.to_excel(writer, sheet_name='Kalem Listesi', index=False)
                
                # Stil ayarlarÄ±
                wb = load_workbook(file_path)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    header_fill = PatternFill(start_color='16213e', end_color='16213e', fill_type='solid')
                    for cell in ws[1]:
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                wb.save(file_path)
                
                QMessageBox.information(self, "BaÅŸarÄ±lÄ±", f"Ä°hale dosyasÄ± Excel'e aktarÄ±ldÄ±:\n{file_path}")
                self.statusBar().showMessage(f"Excel ihale dosyasÄ± oluÅŸturuldu: {file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Excel oluÅŸturulurken hata oluÅŸtu:\n{str(e)}")
    
    def show_unit_converter(self) -> None:
        """Birim dÃ¶nÃ¼ÅŸtÃ¼rÃ¼cÃ¼ dialogu gÃ¶ster"""
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QDoubleSpinBox, QComboBox, QLabel
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Birim DÃ¶nÃ¼ÅŸtÃ¼rÃ¼cÃ¼")
        dialog.setGeometry(300, 300, 400, 200)
        
        layout = QVBoxLayout(dialog)
        
        form_layout = QFormLayout()
        
        # DeÄŸer giriÅŸi
        value_input = QDoubleSpinBox()
        value_input.setRange(0, 999999999)
        value_input.setDecimals(4)
        value_input.setValue(1.0)
        form_layout.addRow("DeÄŸer:", value_input)
        
        # Kaynak birim
        from_unit_combo = QComboBox()
        from_unit_combo.setEditable(True)
        from_unit_combo.addItems(['m', 'mÂ²', 'mÂ³', 'kg', 't', 'cm', 'cmÂ²', 'cmÂ³', 'mm', 'km', 'l', 'ml'])
        form_layout.addRow("Kaynak Birim:", from_unit_combo)
        
        # Hedef birim
        to_unit_combo = QComboBox()
        to_unit_combo.setEditable(True)
        to_unit_combo.addItems(['m', 'mÂ²', 'mÂ³', 'kg', 't', 'cm', 'cmÂ²', 'cmÂ³', 'mm', 'km', 'l', 'ml'])
        form_layout.addRow("Hedef Birim:", to_unit_combo)
        
        # SonuÃ§
        result_label = QLabel("0.0000")
        result_label.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        result_label.setStyleSheet("color: #00BFFF; padding: 10px;")
        form_layout.addRow("SonuÃ§:", result_label)
        
        layout.addLayout(form_layout)
        
        def calculate():
            try:
                value = value_input.value()
                from_unit = from_unit_combo.currentText().strip()
                to_unit = to_unit_combo.currentText().strip()
                
                result = self.calculator.convert_unit(value, from_unit, to_unit)
                result_label.setText(f"{result:,.4f}")
            except Exception as e:
                result_label.setText(f"Hata: {str(e)}")
                result_label.setStyleSheet("color: #c9184a; padding: 10px;")
        
        value_input.valueChanged.connect(calculate)
        from_unit_combo.currentTextChanged.connect(calculate)
        to_unit_combo.currentTextChanged.connect(calculate)
        
        btn_layout = QHBoxLayout()
        btn_close = QPushButton("Kapat")
        btn_close.clicked.connect(dialog.close)
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)
        
        calculate()  # Ä°lk hesaplama
        dialog.exec()
    
    def calculate_auto_fire_rates(self) -> None:
        """TÃ¼m pozlar iÃ§in otomatik fire oranÄ± hesapla"""
        if not self.current_project_id:
            QMessageBox.warning(self, "UyarÄ±", "LÃ¼tfen Ã¶nce bir proje seÃ§in")
            return
        
        reply = QMessageBox.question(
            self, "Onay",
            "TÃ¼m metraj kalemleri iÃ§in kategori bazlÄ± otomatik fire oranÄ± hesaplanacak.\n"
            "Mevcut fire oranlarÄ± gÃ¼ncellenecek. Devam etmek istiyor musunuz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        try:
            metraj_items = self.db.get_project_metraj(self.current_project_id)
            updated_count = 0
            
            for item in metraj_items:
                kategori = item.get('kategori', '')
                auto_fire = self.calculator.get_auto_fire_rate(kategori)
                
                # Poz varsa fire oranÄ±nÄ± gÃ¼ncelle
                poz_no = item.get('poz_no', '')
                if poz_no:
                    poz = self.db.get_poz(poz_no)
                    if poz and poz.get('fire_orani', 0.05) != auto_fire:
                        self.db.update_poz(
                            poz_no=poz_no,
                            fire_orani=auto_fire
                        )
                        updated_count += 1
            
            QMessageBox.information(
                self, "BaÅŸarÄ±lÄ±",
                f"{updated_count} poz iÃ§in fire oranÄ± otomatik olarak gÃ¼ncellendi.\n"
                f"Kategori bazlÄ± fire oranlarÄ± uygulandÄ±."
            )
            self.statusBar().showMessage(f"{updated_count} poz iÃ§in fire oranÄ± gÃ¼ncellendi")
        except Exception as e:
            QMessageBox.critical(
                self, "Hata",
                f"Fire oranÄ± hesaplama sÄ±rasÄ±nda hata oluÅŸtu:\n{str(e)}"
            )
    
    def show_about(self) -> None:
        """HakkÄ±nda dialogu"""
        QMessageBox.about(
            self, "HakkÄ±nda",
            "InsaatMetrajPro v1.0.0\n\n"
            "Ä°nÅŸaat sektÃ¶rÃ¼ iÃ§in profesyonel metraj uygulamasÄ±\n"
            "Python ve PyQt6 ile geliÅŸtirilmiÅŸtir.\n\n"
            "Offline-First yaklaÅŸÄ±m ile Ã§alÄ±ÅŸÄ±r.\n\n"
            "Konut yapÄ±sÄ± iÃ§in 150+ iÅŸ kalemi iÃ§erir."
        )

